"""
Data loading and preprocessing for analysis.

Converts JSON pipeline outputs to analysis-ready DataFrames with:
- Reach-level data (one row per reach)
- Session-level data (one row per video)
- Mouse-level data (aggregated per mouse)

All DataFrames include a 'mouse_id' column for easy joining with external
datasets (connectome, histology, etc.).
"""

import json
import re
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Optional, List, Union
from datetime import datetime


def parse_video_name(video_name: str) -> dict:
    """
    Parse metadata from video filename.

    Expected format: YYYYMMDD_CNTxxxx_[Type][Run]
    Example: 20250624_CNT0115_P2

    Where:
        - YYYYMMDD = date
        - CNTxxxx = mouse ID
        - Type = Tray type (P, E, or F)
        - Run = Run number for that tray type that day (1, 2, 3, etc.)

    Tray types:
        - P = Standard side-facing tray
        - E = Extended/alternate configuration
        - F = Front-facing tray (often unsupported)

    Returns dict with: date, mouse_id, tray_type, run_num, session_id
    """
    # Remove common suffixes
    clean_name = video_name
    for suffix in ['_features', '_reaches', '_pellet_outcomes', '_segments', 'DLC_']:
        clean_name = clean_name.replace(suffix, '')

    # Try standard pattern: DATE_MOUSEID_TYPERUN
    match = re.match(r'(\d{8})_(CNT\d+)_([PEF])(\d+)', clean_name)
    if match:
        date_str, mouse_id, tray_type, run_num = match.groups()
        return {
            'date': datetime.strptime(date_str, '%Y%m%d'),
            'mouse_id': mouse_id,
            'tray_type': tray_type,  # P, E, or F
            'run_num': int(run_num),  # Which run of that type (1, 2, 3, etc.)
            'session_id': f"{mouse_id}_{date_str}",  # Unique per mouse per day
            'run_id': f"{mouse_id}_{date_str}_{tray_type}{run_num}",  # Unique per video
            'video_name': clean_name
        }

    # Fallback - try to extract what we can
    return {
        'date': None,
        'mouse_id': None,
        'tray_type': None,
        'run_num': None,
        'session_id': clean_name,
        'run_id': clean_name,
        'video_name': clean_name
    }


class ReachDataFrame:
    """
    Analysis-ready reach data with filtering and export capabilities.

    Wrapper around pandas DataFrame with convenience methods for:
    - Filtering by mouse, phase, outcome, exclusion status
    - Grouping for statistical analysis
    - Export to various formats
    """

    def __init__(self, df: pd.DataFrame):
        self.df = df

    def __len__(self):
        return len(self.df)

    def __repr__(self):
        n_mice = self.df['mouse_id'].nunique() if 'mouse_id' in self.df.columns else 0
        n_sessions = self.df['session_id'].nunique() if 'session_id' in self.df.columns else 0
        return f"ReachDataFrame({len(self.df)} reaches, {n_mice} mice, {n_sessions} sessions)"

    @property
    def mice(self) -> List[str]:
        """List of unique mouse IDs."""
        return sorted(self.df['mouse_id'].dropna().unique().tolist())

    @property
    def tray_types(self) -> List[str]:
        """List of unique tray types (P, E, F)."""
        if 'tray_type' in self.df.columns:
            return sorted(self.df['tray_type'].dropna().unique().tolist())
        return []

    @property
    def timepoints(self) -> List[str]:
        """List of unique experimental timepoints (if metadata loaded)."""
        if 'timepoint' in self.df.columns:
            return sorted(self.df['timepoint'].dropna().unique().tolist())
        return []

    @property
    def outcomes(self) -> List[str]:
        """List of unique outcomes."""
        if 'outcome' in self.df.columns:
            return sorted(self.df['outcome'].dropna().unique().tolist())
        return []

    def filter(
        self,
        mouse_id: Optional[Union[str, List[str]]] = None,
        tray_type: Optional[Union[str, List[str]]] = None,
        timepoint: Optional[Union[str, List[str]]] = None,
        outcome: Optional[Union[str, List[str]]] = None,
        exclude_flagged: bool = True,
        min_confidence: Optional[float] = None,
        date_range: Optional[tuple] = None
    ) -> 'ReachDataFrame':
        """
        Filter reaches by criteria.

        Args:
            mouse_id: Single mouse ID or list of IDs
            tray_type: Tray type ('P', 'E', 'F') or list
            timepoint: Experimental timepoint (if metadata loaded)
            outcome: Single outcome or list (e.g., 'retrieved')
            exclude_flagged: If True, exclude reaches marked as excluded
            min_confidence: Minimum DLC confidence to include
            date_range: Tuple of (start_date, end_date) to filter by date

        Returns:
            New ReachDataFrame with filtered data
        """
        mask = pd.Series(True, index=self.df.index)

        if mouse_id is not None:
            if isinstance(mouse_id, str):
                mouse_id = [mouse_id]
            mask &= self.df['mouse_id'].isin(mouse_id)

        if tray_type is not None and 'tray_type' in self.df.columns:
            if isinstance(tray_type, str):
                tray_type = [tray_type]
            mask &= self.df['tray_type'].isin(tray_type)

        if timepoint is not None and 'timepoint' in self.df.columns:
            if isinstance(timepoint, str):
                timepoint = [timepoint]
            mask &= self.df['timepoint'].isin(timepoint)

        if outcome is not None and 'outcome' in self.df.columns:
            if isinstance(outcome, str):
                outcome = [outcome]
            mask &= self.df['outcome'].isin(outcome)

        if exclude_flagged and 'exclude_from_analysis' in self.df.columns:
            mask &= ~self.df['exclude_from_analysis'].fillna(False)

        if min_confidence is not None and 'mean_likelihood' in self.df.columns:
            mask &= self.df['mean_likelihood'] >= min_confidence

        if date_range is not None and 'date' in self.df.columns:
            start_date, end_date = date_range
            mask &= (self.df['date'] >= pd.to_datetime(start_date)) & (self.df['date'] <= pd.to_datetime(end_date))

        return ReachDataFrame(self.df[mask].copy())

    def get_kinematic_features(self) -> pd.DataFrame:
        """
        Extract kinematic feature columns suitable for PCA/statistical analysis.

        Returns DataFrame with only numeric kinematic columns, NaN-filtered.
        """
        kinematic_cols = [
            'duration_frames',
            'max_extent_mm',
            'max_extent_ruler',
            'peak_velocity_px_per_frame',
            'mean_velocity_px_per_frame',
            'trajectory_straightness',
            'trajectory_smoothness',
            'hand_angle_at_apex_deg',
            'hand_rotation_total_deg',
        ]

        available_cols = [c for c in kinematic_cols if c in self.df.columns]
        return self.df[available_cols].dropna()

    def get_feature_matrix(self, standardize: bool = True) -> tuple:
        """
        Get feature matrix for PCA/ML analysis.

        Args:
            standardize: If True, z-score normalize features

        Returns:
            (X, feature_names, index) tuple where:
            - X: numpy array of shape (n_samples, n_features)
            - feature_names: list of column names
            - index: DataFrame index for mapping back to reaches
        """
        features_df = self.get_kinematic_features()
        feature_names = features_df.columns.tolist()
        X = features_df.values

        if standardize:
            X = (X - X.mean(axis=0)) / X.std(axis=0)

        return X, feature_names, features_df.index

    def group_by_session(self) -> pd.DataFrame:
        """Aggregate reaches to session level (one row per mouse per day)."""
        agg_funcs = {
            'reach_id': 'count',
            'duration_frames': 'mean',
            'max_extent_mm': 'mean',
            'peak_velocity_px_per_frame': 'mean',
        }

        # Only use columns that exist
        available_agg = {k: v for k, v in agg_funcs.items() if k in self.df.columns}

        # Determine groupby columns based on what exists
        group_cols = ['mouse_id', 'session_id', 'date']
        group_cols = [c for c in group_cols if c in self.df.columns]

        if not group_cols:
            return pd.DataFrame()

        if not available_agg:
            return self.df.groupby(group_cols).size().reset_index(name='n_reaches')

        grouped = self.df.groupby(group_cols).agg(available_agg)
        grouped = grouped.rename(columns={'reach_id': 'n_reaches'})
        return grouped.reset_index()

    def group_by_mouse(self) -> pd.DataFrame:
        """Aggregate reaches to mouse level (across all sessions)."""
        agg_funcs = {
            'reach_id': 'count',
            'duration_frames': 'mean',
            'max_extent_mm': 'mean',
            'peak_velocity_px_per_frame': 'mean',
        }

        available_agg = {k: v for k, v in agg_funcs.items() if k in self.df.columns}

        if not available_agg:
            return self.df.groupby('mouse_id').size().reset_index(name='n_reaches')

        grouped = self.df.groupby('mouse_id').agg(available_agg)
        grouped = grouped.rename(columns={'reach_id': 'n_reaches'})
        return grouped.reset_index()

    def to_csv(self, path: Path):
        """Export to CSV."""
        self.df.to_csv(path, index=False)
        print(f"Exported {len(self.df)} reaches to {path}")

    def to_excel(self, path: Path):
        """Export to Excel with multiple sheets."""
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            self.df.to_excel(writer, sheet_name='Reaches', index=False)
            self.group_by_session().to_excel(writer, sheet_name='Sessions', index=False)
            self.group_by_mouse().to_excel(writer, sheet_name='Mice', index=False)
        print(f"Exported to {path}")


def load_reaches_from_json(reaches_path: Path, outcomes_path: Optional[Path] = None) -> List[dict]:
    """
    Load reaches from a single JSON file and flatten to row dicts.

    Args:
        reaches_path: Path to *_reaches.json file
        outcomes_path: Optional path to *_pellet_outcomes.json for outcome data

    Returns:
        List of dicts, one per reach
    """
    with open(reaches_path) as f:
        reaches_data = json.load(f)

    # Load outcomes if available
    outcomes_by_segment = {}
    if outcomes_path and outcomes_path.exists():
        with open(outcomes_path) as f:
            outcomes_data = json.load(f)
        for seg in outcomes_data.get('segments', []):
            outcomes_by_segment[seg['segment_num']] = seg.get('outcome', 'unknown')

    # Parse video metadata
    video_name = reaches_data.get('video_name', reaches_path.stem.replace('_reaches', ''))
    meta = parse_video_name(video_name)

    rows = []
    for segment in reaches_data.get('segments', []):
        seg_num = segment['segment_num']
        outcome = outcomes_by_segment.get(seg_num, 'unknown')
        seg_flagged = segment.get('flagged_for_review', False)
        seg_flag_reason = segment.get('flag_reason', '')

        for reach in segment.get('reaches', []):
            row = {
                # Identifiers
                'video_name': video_name,
                'mouse_id': meta['mouse_id'],
                'date': meta['date'],
                'tray_type': meta.get('tray_type'),  # P, E, or F
                'run_num': meta.get('run_num'),      # Which run of that tray type
                'session_id': meta['session_id'],
                'run_id': meta.get('run_id'),
                'segment_num': seg_num,
                'reach_id': reach['reach_id'],
                'reach_num': reach['reach_num'],

                # Temporal
                'start_frame': reach['start_frame'],
                'apex_frame': reach['apex_frame'],
                'end_frame': reach['end_frame'],
                'duration_frames': reach['duration_frames'],

                # Spatial
                'max_extent_pixels': reach.get('max_extent_pixels'),
                'max_extent_ruler': reach.get('max_extent_ruler'),

                # Outcome
                'outcome': outcome,

                # Quality flags
                'source': reach.get('source', 'algorithm'),
                'human_corrected': reach.get('human_corrected', False),
                'exclude_from_analysis': reach.get('exclude_from_analysis', False),
                'exclude_reason': reach.get('exclude_reason'),
                'review_note': reach.get('review_note'),
                'segment_flagged': seg_flagged,
                'segment_flag_reason': seg_flag_reason,
            }
            rows.append(row)

    return rows


def load_features_from_json(features_path: Path) -> List[dict]:
    """
    Load features from a *_features.json file and flatten to row dicts.

    Features files contain computed kinematic metrics per reach.
    """
    with open(features_path) as f:
        features_data = json.load(f)

    video_name = features_data.get('video_name', features_path.stem.replace('_features', ''))
    meta = parse_video_name(video_name)

    rows = []
    for segment in features_data.get('segments', []):
        seg_num = segment['segment_num']
        outcome = segment.get('outcome', 'unknown')

        for reach in segment.get('reaches', []):
            row = {
                # Identifiers
                'video_name': video_name,
                'mouse_id': meta['mouse_id'],
                'date': meta['date'],
                'tray_type': meta.get('tray_type'),  # P, E, or F
                'run_num': meta.get('run_num'),      # Which run of that tray type
                'session_id': meta['session_id'],
                'run_id': meta.get('run_id'),
                'segment_num': seg_num,
                'reach_id': reach.get('reach_id'),

                # Temporal
                'start_frame': reach.get('start_frame'),
                'apex_frame': reach.get('apex_frame'),
                'end_frame': reach.get('end_frame'),
                'duration_frames': reach.get('duration_frames'),

                # Spatial
                'max_extent_mm': reach.get('max_extent_mm'),
                'max_extent_ruler': reach.get('max_extent_ruler'),

                # Velocity
                'peak_velocity_px_per_frame': reach.get('peak_velocity_px_per_frame'),
                'velocity_at_apex_mm_per_sec': reach.get('velocity_at_apex_mm_per_sec'),
                'mean_velocity_px_per_frame': reach.get('mean_velocity_px_per_frame'),

                # Trajectory
                'trajectory_straightness': reach.get('trajectory_straightness'),
                'trajectory_smoothness': reach.get('trajectory_smoothness'),

                # Orientation
                'hand_angle_at_apex_deg': reach.get('hand_angle_at_apex_deg'),
                'hand_rotation_total_deg': reach.get('hand_rotation_total_deg'),

                # Body/posture
                'head_width_at_apex_mm': reach.get('head_width_at_apex_mm'),
                'nose_to_slit_at_apex_mm': reach.get('nose_to_slit_at_apex_mm'),
                'head_angle_at_apex_deg': reach.get('head_angle_at_apex_deg'),
                'head_angle_change_deg': reach.get('head_angle_change_deg'),

                # Confidence
                'mean_likelihood': reach.get('mean_likelihood'),
                'frames_low_confidence': reach.get('frames_low_confidence'),

                # Outcome
                'outcome': outcome,
                'is_causal': reach.get('causal_reach', False),

                # Quality flags
                'exclude_from_analysis': reach.get('exclude_from_analysis', False),
                'exclude_reason': reach.get('exclude_reason'),
                'human_corrected': reach.get('human_corrected', False),
                'source': reach.get('source', 'algorithm'),
            }
            rows.append(row)

    return rows


def load_all_data(
    data_dir: Path,
    use_features: bool = True,
    exclude_flagged: bool = True
) -> ReachDataFrame:
    """
    Load all reach data from a directory.

    Args:
        data_dir: Directory containing pipeline output files
        use_features: If True, load *_features.json (has kinematic metrics).
                      If False, load *_reaches.json (has raw detections).
        exclude_flagged: If True, automatically exclude flagged reaches

    Returns:
        ReachDataFrame with all loaded data
    """
    data_dir = Path(data_dir)
    all_rows = []

    if use_features:
        # Load from features files (has computed kinematics)
        for fpath in sorted(data_dir.glob('*_features.json')):
            try:
                rows = load_features_from_json(fpath)
                all_rows.extend(rows)
            except Exception as e:
                print(f"Warning: Could not load {fpath.name}: {e}")
    else:
        # Load from reaches files (raw detections)
        for rpath in sorted(data_dir.glob('*_reaches.json')):
            try:
                # Try to find matching outcomes file
                base_name = rpath.stem.replace('_reaches', '')
                outcomes_path = data_dir / f"{base_name}_pellet_outcomes.json"
                rows = load_reaches_from_json(rpath, outcomes_path)
                all_rows.extend(rows)
            except Exception as e:
                print(f"Warning: Could not load {rpath.name}: {e}")

    if not all_rows:
        print(f"No data found in {data_dir}")
        return ReachDataFrame(pd.DataFrame())

    df = pd.DataFrame(all_rows)

    # Convert date column
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date'])

    print(f"Loaded {len(df)} reaches from {df['video_name'].nunique()} videos")
    print(f"  Mice: {df['mouse_id'].nunique()}")
    if 'tray_type' in df.columns:
        tray_types = sorted(df['tray_type'].dropna().unique())
        print(f"  Tray types: {tray_types}")

    result = ReachDataFrame(df)

    if exclude_flagged:
        n_before = len(result)
        result = result.filter(exclude_flagged=True)
        n_excluded = n_before - len(result)
        if n_excluded > 0:
            print(f"  Excluded {n_excluded} flagged reaches")

    return result


# =============================================================================
# EXPERIMENTAL METADATA INTEGRATION
# =============================================================================

# Experimental timeline phases - maps Test_Phase names to broader categories
TIMEPOINT_MAPPING = {
    # Training phases
    'Training_Flat_1': 'Training',
    'Training_Flat_2': 'Training',
    'Training_Flat_3': 'Training',
    'Training_Pillar_1': 'Training',
    'Training_Pillar_2': 'Training',
    'Training_Pillar_3': 'Training',
    'Training_Pillar_4': 'Training',
    'Training_Pillar_5': 'Training',
    'Training_Pillar_6': 'Training',
    'Training_Pillar_7': 'Training',
    # Pre-injury test (baseline)
    'Pre-Injury_Test_Pillar_1': 'Pre-Injury',
    'Pre-Injury_Test_Pillar_2': 'Pre-Injury',
    'Pre-Injury_Test_Pillar_3': 'Pre-Injury',
    # Post-injury tests
    'Post-Injury_Test_1': 'Post-Injury',
    'Post-Injury_Test_2': 'Post-Injury',
    'Post-Injury_Test_3': 'Post-Injury',
    'Post-Injury_Test_4': 'Post-Injury',
    # Rehab phases
    'Rehab_Easy_1': 'Rehab_Easy',
    'Rehab_Easy_2': 'Rehab_Easy',
    'Rehab_Easy_3': 'Rehab_Easy',
    'Rehab_Easy_4': 'Rehab_Easy',
    'Rehab_Easy_5': 'Rehab_Easy',
    'Rehab_Easy_6': 'Rehab_Easy',
    'Rehab_Easy_7': 'Rehab_Easy',
    'Rehab_Easy_8': 'Rehab_Easy',
    'Rehab_Easy_9': 'Rehab_Easy',
    'Rehab_Flat_1': 'Rehab_Flat',
    'Rehab_Flat_2': 'Rehab_Flat',
    'Rehab_Flat_3': 'Rehab_Flat',
    'Rehab_Flat_4': 'Rehab_Flat',
    'Rehab_Flat_5': 'Rehab_Flat',
    'Rehab_Flat_6': 'Rehab_Flat',
    'Rehab_Flat_7': 'Rehab_Flat',
    'Rehab_Pillar_1': 'Rehab_Pillar',
    'Rehab_Pillar_2': 'Rehab_Pillar',
    'Rehab_Pillar_3': 'Rehab_Pillar',
    'Rehab_Pillar_4': 'Rehab_Pillar',
    'Rehab_Pillar_5': 'Rehab_Pillar',
    'Rehab_Pillar_6': 'Rehab_Pillar',
    'Rehab_Pillar_7': 'Rehab_Pillar',
}


def categorize_timepoint(test_phase: str) -> str:
    """
    Map detailed Test_Phase to broader experimental timepoint category.

    Args:
        test_phase: Detailed phase name from tracking sheet (e.g., 'Pre-Injury_Test_Pillar_2')

    Returns:
        Category: 'Training', 'Pre-Injury', 'Post-Injury', 'Rehab_Easy', 'Rehab_Flat', 'Rehab_Pillar'
    """
    if pd.isna(test_phase):
        return None
    return TIMEPOINT_MAPPING.get(str(test_phase), None)


def load_tracking_metadata(
    tracking_file: Path,
    sheet_name: str = '3b_Manual_Tray'
) -> pd.DataFrame:
    """
    Load experimental metadata from a cohort tracking spreadsheet.

    The tracking spreadsheet contains manual pellet scores and experimental phase info.

    Args:
        tracking_file: Path to Connectome_XX_Animal_Tracking.xlsx file
        sheet_name: Sheet name containing tray data (default: '3b_Manual_Tray')

    Returns:
        DataFrame with columns: Date, Animal, Sex, Weight, Test_Phase, Tray_Type, Tray_Num,
        plus manual pellet scores and calculated metrics.
    """
    tracking_file = Path(tracking_file)
    if not tracking_file.exists():
        raise FileNotFoundError(f"Tracking file not found: {tracking_file}")

    # Load the manual tray sheet
    df = pd.read_excel(tracking_file, sheet_name=sheet_name)

    # Standardize column names
    df.columns = df.columns.astype(str).str.strip()

    # Ensure Date is datetime
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'])

    # Parse tray type/number (e.g., "P2" -> tray_type="P", tray_num=2)
    if 'Tray Type/Number' in df.columns:
        def parse_tray(val):
            if pd.isna(val) or len(str(val)) < 2:
                return None, None
            s = str(val)
            return s[0], int(s[1:]) if s[1:].isdigit() else None

        df['tray_type'], df['tray_num'] = zip(*df['Tray Type/Number'].apply(parse_tray))

    # Normalize Animal/mouse ID column
    if 'Animal' in df.columns:
        df['mouse_id'] = df['Animal'].str.replace('_', '')  # CNT_05_01 -> CNT0501

    # Add timepoint category
    if 'Test_Phase' in df.columns:
        df['timepoint'] = df['Test_Phase'].apply(categorize_timepoint)

    return df


def load_all_tracking_metadata(tracking_dir: Path) -> pd.DataFrame:
    """
    Load and combine metadata from all tracking spreadsheets in a directory.

    Args:
        tracking_dir: Directory containing Connectome_XX_Animal_Tracking.xlsx files

    Returns:
        Combined DataFrame with all cohorts' metadata
    """
    tracking_dir = Path(tracking_dir)
    all_dfs = []

    for fpath in sorted(tracking_dir.glob('Connectome_*_Animal_Tracking*.xlsx')):
        # Skip temp files and duplicates
        if fpath.name.startswith('~') or '_fixed' in fpath.name or '(' in fpath.name:
            continue

        try:
            df = load_tracking_metadata(fpath)
            df['source_file'] = fpath.name
            all_dfs.append(df)
            print(f"  Loaded {len(df)} rows from {fpath.name}")
        except Exception as e:
            print(f"  Warning: Could not load {fpath.name}: {e}")

    if not all_dfs:
        return pd.DataFrame()

    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"Total: {len(combined)} metadata rows from {len(all_dfs)} files")
    return combined


def merge_with_metadata(
    reach_data: ReachDataFrame,
    metadata_df: pd.DataFrame
) -> ReachDataFrame:
    """
    Merge reach data with experimental metadata from tracking spreadsheets.

    Matches on: mouse_id + date + tray_type + run_num

    This adds columns like:
    - Test_Phase: Detailed phase (e.g., 'Pre-Injury_Test_Pillar_1')
    - timepoint: Category (e.g., 'Pre-Injury')
    - Weight, Sex from tracking sheet

    Args:
        reach_data: ReachDataFrame with video-derived data
        metadata_df: DataFrame from load_tracking_metadata()

    Returns:
        ReachDataFrame with merged metadata columns
    """
    df = reach_data.df.copy()

    # Ensure consistent date format
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date']).dt.normalize()
    if 'Date' in metadata_df.columns:
        metadata_df['Date'] = pd.to_datetime(metadata_df['Date']).dt.normalize()

    # Normalize mouse IDs for matching
    if 'mouse_id' in df.columns and 'Animal' in metadata_df.columns:
        # Video filenames use CNT0501, tracking sheets use CNT_05_01
        df['_mouse_norm'] = df['mouse_id'].str.upper()
        metadata_df['_mouse_norm'] = metadata_df['Animal'].str.replace('_', '').str.upper()

    # Build merge keys
    merge_cols_left = ['_mouse_norm', 'date', 'tray_type', 'run_num']
    merge_cols_right = ['_mouse_norm', 'Date', 'tray_type', 'tray_num']

    # Check which columns exist
    available_left = [c for c in merge_cols_left if c in df.columns]
    available_right = [c for c in merge_cols_right if c in metadata_df.columns]

    if len(available_left) < 2 or len(available_right) < 2:
        print("Warning: Insufficient columns for metadata merge")
        return reach_data

    # Select metadata columns to merge
    meta_cols = ['Test_Phase', 'timepoint', 'Sex', 'Weight', 'Weight %']
    meta_cols = [c for c in meta_cols if c in metadata_df.columns]

    # Create merge subset
    meta_for_merge = metadata_df[available_right + meta_cols].drop_duplicates()

    # Rename for merge
    rename_map = {'Date': 'date', 'tray_num': 'run_num'}
    meta_for_merge = meta_for_merge.rename(columns=rename_map)

    # Perform merge
    merged = df.merge(
        meta_for_merge,
        on=[c for c in available_left if c in meta_for_merge.columns],
        how='left',
        suffixes=('', '_meta')
    )

    # Clean up temp columns
    if '_mouse_norm' in merged.columns:
        merged = merged.drop(columns=['_mouse_norm'])

    n_matched = merged['Test_Phase'].notna().sum() if 'Test_Phase' in merged.columns else 0
    print(f"Merged metadata: {n_matched}/{len(merged)} reaches matched ({100*n_matched/len(merged):.1f}%)")

    return ReachDataFrame(merged)


def load_data_with_metadata(
    data_dir: Path,
    tracking_dir: Optional[Path] = None,
    use_features: bool = True,
    exclude_flagged: bool = True
) -> ReachDataFrame:
    """
    Load reach data and merge with experimental metadata.

    Convenience function that combines load_all_data() with metadata loading.

    Args:
        data_dir: Directory containing pipeline output files
        tracking_dir: Directory containing tracking spreadsheets (optional)
        use_features: If True, load *_features.json files
        exclude_flagged: If True, exclude flagged reaches

    Returns:
        ReachDataFrame with both kinematic features and experimental metadata
    """
    # Load reach data
    reach_data = load_all_data(data_dir, use_features=use_features, exclude_flagged=False)

    # Load and merge metadata if tracking_dir provided
    if tracking_dir is not None:
        tracking_dir = Path(tracking_dir)
        if tracking_dir.exists():
            print(f"\nLoading metadata from {tracking_dir}")
            metadata = load_all_tracking_metadata(tracking_dir)
            if len(metadata) > 0:
                reach_data = merge_with_metadata(reach_data, metadata)

    # Apply exclusion filter after merge
    if exclude_flagged:
        n_before = len(reach_data)
        reach_data = reach_data.filter(exclude_flagged=True)
        n_excluded = n_before - len(reach_data)
        if n_excluded > 0:
            print(f"  Excluded {n_excluded} flagged reaches")

    return reach_data


# =============================================================================
# SURGERY/MOUSE-LEVEL METADATA
# =============================================================================

def load_surgery_metadata(tracking_file: Path) -> pd.DataFrame:
    """
    Load surgery data from a cohort tracking spreadsheet.

    Extracts mouse-level data from:
    - 4_Contusion_Injury_Details (contusion surgery)
    - 5_SC_Injection_Details (viral injection)

    Args:
        tracking_file: Path to Connectome_XX_Animal_Tracking.xlsx file

    Returns:
        DataFrame with one row per mouse containing surgery details
    """
    import openpyxl

    tracking_file = Path(tracking_file)
    if not tracking_file.exists():
        raise FileNotFoundError(f"Tracking file not found: {tracking_file}")

    wb = openpyxl.load_workbook(tracking_file, data_only=True)

    surgery_data = {}  # mouse_id -> dict of all fields

    # === 4_Contusion_Injury_Details ===
    if '4_Contusion_Injury_Details' in wb.sheetnames:
        ws = wb['4_Contusion_Injury_Details']

        # Build header map
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = ws.cell(1, col).value
            if h:
                headers[str(h).strip()] = col

        # Find ID column
        id_col = None
        for name in ['Subject_ID', 'SubjectID', 'Animal', 'Mouse']:
            if name in headers:
                id_col = headers[name]
                break

        if id_col:
            # Field mappings (sheet column -> output column)
            field_map = {
                'Surgery_Date': 'surgery_date',
                'Surgery_Type': 'surgery_type',
                'Surgery_Severity': 'surgery_severity',
                'Contusion_Location': 'injury_location',
                'Subject_Weight (g)': 'surgery_weight_g',
                'Anesthetic': 'anesthetic',
                'Analgesic': 'analgesic',
                'Intended_kd': 'intended_kd',
                'Intended_Dwell': 'intended_dwell_ms',
                'Actual_kd': 'actual_kd',
                'Actual_displacement': 'actual_displacement_um',
                'Actual_Velocity': 'actual_velocity_mm_s',
                'Actual_Dwell': 'actual_dwell_ms',
                'Survived': 'survived_surgery',
            }

            for row in range(2, ws.max_row + 1):
                animal_id = ws.cell(row, id_col).value
                if not animal_id:
                    continue

                # Normalize ID (CNT_05_01 -> CNT0501)
                animal_id = str(animal_id).strip().replace('_', '')

                if animal_id not in surgery_data:
                    surgery_data[animal_id] = {}

                for sheet_field, data_field in field_map.items():
                    if sheet_field in headers:
                        val = ws.cell(row, headers[sheet_field]).value
                        surgery_data[animal_id][data_field] = val

    # === 5_SC_Injection_Details ===
    if '5_SC_Injection_Details' in wb.sheetnames:
        ws = wb['5_SC_Injection_Details']

        # Build header map
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = ws.cell(1, col).value
            if h:
                headers[str(h).strip()] = col

        # Find ID column
        id_col = None
        for name in ['Subject_ID', 'SubjectID', 'Animal', 'Mouse']:
            if name in headers:
                id_col = headers[name]
                break

        if id_col:
            field_map = {
                'Surgery_Date': 'injection_date',
                'Injected_Virus': 'virus',
                'Virus_Titer': 'virus_titer',
                'Injection_Target': 'injection_target',
                'Depths (D/V)': 'injection_depth_dv',
                'Coordinates (M/L)': 'injection_coord_ml',
                'Signal Post Perfusion': 'signal_post_perfusion',
            }

            for row in range(2, ws.max_row + 1):
                animal_id = ws.cell(row, id_col).value
                if not animal_id:
                    continue

                animal_id = str(animal_id).strip().replace('_', '')

                if animal_id not in surgery_data:
                    surgery_data[animal_id] = {}

                for sheet_field, data_field in field_map.items():
                    if sheet_field in headers:
                        val = ws.cell(row, headers[sheet_field]).value
                        surgery_data[animal_id][data_field] = val

    wb.close()

    if not surgery_data:
        return pd.DataFrame()

    # Convert to DataFrame
    rows = []
    for mouse_id, fields in surgery_data.items():
        row = {'mouse_id': mouse_id}
        row.update(fields)
        rows.append(row)

    df = pd.DataFrame(rows)

    # Convert surgery_date to datetime
    if 'surgery_date' in df.columns:
        df['surgery_date'] = pd.to_datetime(df['surgery_date'], errors='coerce')
    if 'injection_date' in df.columns:
        df['injection_date'] = pd.to_datetime(df['injection_date'], errors='coerce')

    return df


def load_all_surgery_metadata(tracking_dir: Path) -> pd.DataFrame:
    """
    Load and combine surgery metadata from all tracking spreadsheets.

    Args:
        tracking_dir: Directory containing Connectome_XX_Animal_Tracking.xlsx files

    Returns:
        Combined DataFrame with all cohorts' surgery metadata
    """
    tracking_dir = Path(tracking_dir)
    all_dfs = []

    for fpath in sorted(tracking_dir.glob('Connectome_*_Animal_Tracking*.xlsx')):
        # Skip temp files and duplicates
        if fpath.name.startswith('~') or '_fixed' in fpath.name or '(' in fpath.name:
            continue

        try:
            df = load_surgery_metadata(fpath)
            if len(df) > 0:
                # Extract cohort from filename
                import re
                match = re.search(r'Connectome_(\d+)_', fpath.name)
                if match:
                    df['cohort'] = f'CNT_{int(match.group(1)):02d}'
                df['source_file'] = fpath.name
                all_dfs.append(df)
                print(f"  Loaded {len(df)} mice from {fpath.name}")
        except Exception as e:
            print(f"  Warning: Could not load surgery data from {fpath.name}: {e}")

    if not all_dfs:
        return pd.DataFrame()

    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"Total: {len(combined)} mice with surgery data")
    return combined


# =============================================================================
# MOUSE ID NORMALIZATION
# =============================================================================

def normalize_mouse_id(mouse_id: str) -> str:
    """
    Normalize mouse ID to canonical format: CNT0102.

    Handles various formats:
        CNT_01_02       -> CNT0102  (tracking sheets, brainglobe)
        CNT_01_02/349   -> CNT0102  (brainglobe with brain number)
        CNT0102         -> CNT0102  (already canonical)
        CNT01_02        -> CNT0102  (partial underscore)

    Args:
        mouse_id: Raw mouse ID string

    Returns:
        Normalized ID in CNT0102 format
    """
    if pd.isna(mouse_id) or not mouse_id:
        return None

    s = str(mouse_id).strip()

    # Remove brain number suffix (e.g., /349)
    if '/' in s:
        s = s.split('/')[0]

    # Remove all underscores
    s = s.replace('_', '')

    # Uppercase
    s = s.upper()

    return s


def extract_cohort_from_mouse_id(mouse_id: str) -> str:
    """
    Extract cohort number from normalized mouse ID.

    Examples:
        CNT0102 -> CNT_01
        CNT0501 -> CNT_05

    Returns cohort in CNT_XX format.
    """
    if not mouse_id or len(mouse_id) < 5:
        return None

    # Normalize first
    norm_id = normalize_mouse_id(mouse_id)
    if not norm_id or not norm_id.startswith('CNT'):
        return None

    # Extract cohort digits (positions 3-4 after CNT)
    try:
        cohort_num = norm_id[3:5]
        return f"CNT_{cohort_num}"
    except (IndexError, ValueError):
        return None


# =============================================================================
# BRAINGLOBE CONNECTOMICS DATA
# =============================================================================

def load_brainglobe_data(
    region_counts_path: Path = None,
    normalize_ids: bool = True
) -> pd.DataFrame:
    """
    Load BrainGlobe nuclei detection results (cell counts per brain region).

    The region_counts.csv file contains one row per brain with:
    - Identification: brain, subject, cohort, project_code
    - Detection parameters: det_preset, ball_xy, ball_z, soma_diameter, threshold
    - Cell counts: total_cells + ~400 region-specific columns (region_AAA, region_GRN, etc.)

    Args:
        region_counts_path: Path to region_counts.csv file (from BrainGlobe nuclei detection)
            Must be provided - no default path (external to MouseReach)
        normalize_ids: If True, normalize subject column to canonical mouse ID format

    Returns:
        DataFrame with one row per mouse, region columns as cell counts
    """
    # No default - this is external data that must be explicitly provided
    if region_counts_path is None:
        print("  BrainGlobe region_counts_path not provided (optional external data)")
        return pd.DataFrame()

    region_counts_path = Path(region_counts_path)

    if not region_counts_path.exists():
        print(f"  BrainGlobe data not found: {region_counts_path}")
        return pd.DataFrame()

    try:
        df = pd.read_csv(region_counts_path)
    except Exception as e:
        print(f"  Error loading BrainGlobe data: {e}")
        return pd.DataFrame()

    print(f"  Loaded {len(df)} brains from {region_counts_path.name}")

    # Normalize subject IDs
    if normalize_ids and 'subject' in df.columns:
        df['mouse_id'] = df['subject'].apply(normalize_mouse_id)

        # Show mapping for verification
        unique_mappings = df[['subject', 'mouse_id']].drop_duplicates()
        print(f"  Subject ID mappings: {len(unique_mappings)} unique")

    # Identify region columns (start with 'region_')
    region_cols = [c for c in df.columns if c.startswith('region_')]
    print(f"  Region columns: {len(region_cols)}")

    # Add prefix to region columns to distinguish from other data
    # (rename region_GRN -> brainglobe_GRN for clarity)
    rename_map = {c: f"brainglobe_{c.replace('region_', '')}" for c in region_cols}
    df = df.rename(columns=rename_map)

    # Select columns to keep
    id_cols = ['mouse_id', 'brain', 'brain_id', 'cohort', 'total_cells']
    id_cols = [c for c in id_cols if c in df.columns]

    brainglobe_cols = [c for c in df.columns if c.startswith('brainglobe_')]

    result_df = df[id_cols + brainglobe_cols].copy()

    return result_df


def get_reaching_region_summary(df: pd.DataFrame) -> dict:
    """
    Summarize cell counts in reaching-relevant brain regions.

    Based on regions defined in the nuclei detection pipeline as relevant
    for skilled reaching behavior.

    Key regions:
    - Motor: MOp (primary motor), MOs (secondary motor)
    - Brainstem reticular: GRN (gigantocellular), MRN (midbrain reticular)
    - Red nucleus: RN
    - Spinal: (if present)

    Args:
        df: DataFrame with brainglobe_ columns

    Returns:
        Dict with summarized counts for reaching-relevant regions
    """
    # Region groupings for reaching (from nuclei detection pipeline)
    reaching_regions = {
        'motor_cortex': ['brainglobe_MOp1', 'brainglobe_MOp2/3', 'brainglobe_MOp5',
                         'brainglobe_MOp6a', 'brainglobe_MOp6b',
                         'brainglobe_MOs1', 'brainglobe_MOs2/3', 'brainglobe_MOs5',
                         'brainglobe_MOs6a'],
        'reticular_formation': ['brainglobe_GRN', 'brainglobe_MRN', 'brainglobe_IRN',
                                 'brainglobe_MDRNd', 'brainglobe_MDRNv'],
        'red_nucleus': ['brainglobe_RN'],
        'rubrospinal': ['brainglobe_rust'],
    }

    summary = {}
    for group_name, region_list in reaching_regions.items():
        available = [r for r in region_list if r in df.columns]
        if available:
            summary[f'{group_name}_total'] = df[available].sum(axis=1).sum()
            summary[f'{group_name}_regions'] = available

    return summary


# =============================================================================
# SESSION AND SEGMENT CONTEXT
# =============================================================================

def add_session_context(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add session-level derived statistics to each reach row.

    Adds columns:
    - session_total_reaches: Total reaches in this session (mouse+date)
    - session_reach_position: This reach's position (1-indexed) in session
    - run_total_reaches: Total reaches in this run (single video)
    - run_reach_position: This reach's position in the run
    - session_total_segments: Total segments in session
    - session_segment_position: This segment's position in session

    Args:
        df: DataFrame with reach-level data

    Returns:
        DataFrame with session context columns added
    """
    df = df.copy()

    # Session-level counts (mouse + date)
    if 'session_id' in df.columns:
        session_counts = df.groupby('session_id').agg(
            session_total_reaches=('reach_id', 'count'),
            session_total_segments=('segment_num', 'nunique')
        ).reset_index()
        df = df.merge(session_counts, on='session_id', how='left')

        # Position within session
        df['session_reach_position'] = df.groupby('session_id').cumcount() + 1

        # Segment position within session
        if 'segment_num' in df.columns:
            segment_order = df.groupby(['session_id', 'segment_num']).first().reset_index()[['session_id', 'segment_num']]
            segment_order['session_segment_position'] = segment_order.groupby('session_id').cumcount() + 1
            df = df.merge(segment_order, on=['session_id', 'segment_num'], how='left')

    # Run-level counts (single video)
    if 'run_id' in df.columns:
        run_counts = df.groupby('run_id').agg(
            run_total_reaches=('reach_id', 'count')
        ).reset_index()
        df = df.merge(run_counts, on='run_id', how='left')

        # Position within run
        df['run_reach_position'] = df.groupby('run_id').cumcount() + 1

    return df


def add_segment_context(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add segment-level context to each reach row.

    Adds columns:
    - segment_n_reaches: Total reaches in this segment
    - segment_reach_position: This reach's position in segment
    - segment_prior_outcomes: Outcomes of prior reaches in segment
    - segment_prior_misses: Count of misses before this reach in segment

    Args:
        df: DataFrame with reach-level data

    Returns:
        DataFrame with segment context columns added
    """
    df = df.copy()

    if 'run_id' not in df.columns or 'segment_num' not in df.columns:
        return df

    # Segment-level counts
    segment_key = ['run_id', 'segment_num']
    segment_counts = df.groupby(segment_key).agg(
        segment_n_reaches=('reach_id', 'count')
    ).reset_index()
    df = df.merge(segment_counts, on=segment_key, how='left')

    # Position within segment
    df['segment_reach_position'] = df.groupby(segment_key).cumcount() + 1

    # Is this the first/last reach in segment?
    df['is_first_reach_in_segment'] = df['segment_reach_position'] == 1
    df['is_last_reach_in_segment'] = df['segment_reach_position'] == df['segment_n_reaches']

    return df


# =============================================================================
# INDIVIDUAL REACH OUTCOME DERIVATION
# =============================================================================

"""
Reach Outcome Categories:
    CAUSAL REACHES (the reach during which interaction_frame occurred):
        - retrieved: Grabbed the pellet
        - displaced_sa: Knocked pellet into scoring area
        - displaced_outside: Knocked pellet outside

    NON-CAUSAL REACHES:
        - miss_on_pillar: Reach occurred BEFORE interaction, pellet was available (had a chance)
        - miss_off_pillar: Reach occurred AFTER interaction, pellet already gone (no chance)

    SPECIAL CASES:
        - untouched_segment: Segment where pellet was never contacted (all reaches are miss_on_pillar)
"""


def derive_reach_outcomes(df: pd.DataFrame) -> pd.DataFrame:
    """
    Derive individual reach outcomes based on interaction_frame timing.

    The segment-level pellet outcome (retrieved/displaced/untouched) applies to the
    whole segment, but each reach has its own outcome:

    - The CAUSAL REACH (reach during which interaction_frame occurred) gets the segment outcome
    - Reaches BEFORE the causal reach are 'miss_on_pillar' (pellet was available, they missed)
    - Reaches AFTER the causal reach are 'miss_off_pillar' (pellet was already displaced/retrieved)
    - In 'untouched' segments, all reaches are 'miss_on_pillar' (pellet was there, all missed)

    Args:
        df: DataFrame with columns: start_frame, end_frame, outcome (segment), interaction_frame

    Returns:
        DataFrame with new 'reach_outcome' column
    """
    df = df.copy()

    # First, identify causal reaches for each segment based on interaction_frame timing
    df['is_causal_reach_derived'] = False
    df['reach_outcome'] = 'miss_unknown'

    # Group by video+segment to process each segment
    group_cols = ['video_name', 'segment_num'] if 'video_name' in df.columns else ['segment_num']

    for group_key, segment_df in df.groupby(group_cols):
        segment_outcome = segment_df['outcome'].iloc[0] if 'outcome' in segment_df.columns else 'unknown'
        interaction_frame = segment_df['interaction_frame'].iloc[0] if 'interaction_frame' in segment_df.columns else None

        # Handle untouched segments - all reaches are miss_on_pillar
        if segment_outcome == 'untouched' or pd.isna(interaction_frame):
            df.loc[segment_df.index, 'reach_outcome'] = 'miss_on_pillar'
            continue

        # Find the causal reach (contains interaction_frame)
        causal_idx = None
        for idx, row in segment_df.iterrows():
            if row['start_frame'] <= interaction_frame <= row['end_frame']:
                causal_idx = idx
                break

        # If no reach contains interaction_frame, find closest reach ending before it
        if causal_idx is None:
            before_interaction = segment_df[segment_df['end_frame'] <= interaction_frame]
            if len(before_interaction) > 0:
                causal_idx = before_interaction.iloc[-1].name

        # Assign outcomes to each reach in segment
        for idx, row in segment_df.iterrows():
            if idx == causal_idx:
                # Causal reach gets the segment outcome
                df.loc[idx, 'reach_outcome'] = segment_outcome
                df.loc[idx, 'is_causal_reach_derived'] = True
            elif causal_idx is not None and row['end_frame'] <= df.loc[causal_idx, 'start_frame']:
                # Reach ended before causal reach started -> pellet was available
                df.loc[idx, 'reach_outcome'] = 'miss_on_pillar'
            else:
                # Reach started after causal reach or overlaps -> pellet was gone
                df.loc[idx, 'reach_outcome'] = 'miss_off_pillar'

    return df


def get_reach_outcome_summary(df: pd.DataFrame) -> dict:
    """
    Get summary statistics of reach outcomes.

    Args:
        df: DataFrame with 'reach_outcome' column

    Returns:
        Dict with counts and percentages for each outcome type
    """
    if 'reach_outcome' not in df.columns:
        return {}

    counts = df['reach_outcome'].value_counts()
    total = len(df)

    summary = {
        'total_reaches': total,
        'outcomes': {}
    }

    for outcome, count in counts.items():
        summary['outcomes'][outcome] = {
            'count': int(count),
            'percentage': float(count / total * 100)
        }

    # Derived metrics
    n_causal = df['is_causal_reach_derived'].sum() if 'is_causal_reach_derived' in df.columns else 0
    n_retrieved = counts.get('retrieved', 0)
    n_displaced = counts.get('displaced_sa', 0) + counts.get('displaced_outside', 0)
    n_miss_on_pillar = counts.get('miss_on_pillar', 0)
    n_miss_off_pillar = counts.get('miss_off_pillar', 0)

    summary['causal_reaches'] = int(n_causal)
    summary['retrieval_rate_of_causal'] = float(n_retrieved / n_causal * 100) if n_causal > 0 else 0
    summary['contact_rate_of_causal'] = float((n_retrieved + n_displaced) / n_causal * 100) if n_causal > 0 else 0
    summary['miss_on_pillar_total'] = int(n_miss_on_pillar)
    summary['miss_off_pillar_total'] = int(n_miss_off_pillar)

    return summary


# =============================================================================
# UNIFIED "GOD VIEW" DATA BUILDER
# =============================================================================

def build_unified_reach_data(
    data_dir: Path,
    tracking_dir: Optional[Path] = None,
    brainglobe_path: Optional[Path] = None,
    use_features: bool = True,
    exclude_flagged: bool = True,
    include_surgery: bool = True,
    include_brainglobe: bool = True
) -> ReachDataFrame:
    """
    Build the unified reach-centric dataset with ALL hierarchical data attached.

    This is the "god view" of reaches - each row contains:
    - Reach identification (video, segment, reach IDs)
    - Temporal bounds (start/apex/end frames)
    - Kinematic features (extent, velocity, trajectory, etc.)
    - Pellet outcome (miss/displaced/retrieved)
    - Segment context (reaches in segment, position, prior outcomes)
    - Session context (position in day's testing, totals)
    - Experimental metadata (Test_Phase, timepoint, Weight)
    - Mouse-level surgery data (injury details, severity, etc.)
    - Days post injury (computed from surgery date and session date)
    - BrainGlobe connectomics (cell counts per brain region for that mouse)

    Args:
        data_dir: Directory containing pipeline output files (*_features.json)
        tracking_dir: Directory containing Connectome_XX_Animal_Tracking.xlsx files
        brainglobe_path: Path to region_counts.csv (default: auto-detect)
        use_features: If True, load computed kinematics from *_features.json
        exclude_flagged: If True, exclude reaches flagged for exclusion
        include_surgery: If True, merge surgery/mouse-level metadata
        include_brainglobe: If True, merge BrainGlobe connectomics data

    Returns:
        ReachDataFrame with all data merged - the complete reach-centric dataset

    Example:
        >>> from mousereach.analysis import build_unified_reach_data
        >>> data = build_unified_reach_data(
        ...     data_dir='/path/to/MouseReach_Pipeline/Processing',
        ...     tracking_dir='/path/to/Animal_Cohorts'
        ... )
        >>> print(f"Loaded {len(data)} reaches with {len(data.df.columns)} columns")
        >>> data.to_csv('all_reaches_unified.csv')
    """
    data_dir = Path(data_dir)

    print("=" * 60)
    print("Building Unified Reach-Centric Dataset")
    print("=" * 60)

    # 1. Load base reach data from pipeline
    print("\n[1/5] Loading reach data from pipeline...")
    reach_data = load_all_data(data_dir, use_features=use_features, exclude_flagged=False)

    if len(reach_data) == 0:
        print("No reach data found.")
        return reach_data

    df = reach_data.df.copy()
    print(f"  Base data: {len(df)} reaches, {df['mouse_id'].nunique()} mice")

    # 2. Add segment context
    print("\n[2/5] Adding segment context...")
    df = add_segment_context(df)
    print(f"  Added: segment_n_reaches, segment_reach_position, is_first/last_reach")

    # 3. Add session context
    print("\n[3/5] Adding session context...")
    df = add_session_context(df)
    print(f"  Added: session_total_reaches, session_reach_position, run_total_reaches")

    # 4. Merge experimental metadata (Test_Phase, Weight, etc.)
    if tracking_dir is not None:
        tracking_dir = Path(tracking_dir)
        if tracking_dir.exists():
            print(f"\n[4/5] Loading experimental metadata from {tracking_dir}...")
            metadata = load_all_tracking_metadata(tracking_dir)
            if len(metadata) > 0:
                # Merge on mouse_id + date + tray_type + run_num
                reach_df_temp = ReachDataFrame(df)
                reach_df_temp = merge_with_metadata(reach_df_temp, metadata)
                df = reach_df_temp.df

                # Show what was merged
                if 'Test_Phase' in df.columns:
                    matched = df['Test_Phase'].notna().sum()
                    print(f"  Test_Phase matched: {matched}/{len(df)} reaches")
        else:
            print(f"\n[4/5] Tracking dir not found: {tracking_dir}")
    else:
        print("\n[4/5] No tracking_dir provided - skipping experimental metadata")

    # 5. Merge surgery/mouse-level metadata
    if include_surgery and tracking_dir is not None:
        tracking_dir = Path(tracking_dir)
        if tracking_dir.exists():
            print(f"\n[5/6] Loading surgery metadata...")
            surgery_df = load_all_surgery_metadata(tracking_dir)
            if len(surgery_df) > 0:
                # Normalize mouse_id for merge
                df['_mouse_norm'] = df['mouse_id'].apply(normalize_mouse_id)
                surgery_df['_mouse_norm'] = surgery_df['mouse_id'].apply(normalize_mouse_id)

                # Select surgery columns to merge (exclude duplicates)
                surgery_cols = [c for c in surgery_df.columns
                               if c not in ['mouse_id', 'source_file'] and c != '_mouse_norm']

                # Merge on normalized mouse_id
                df = df.merge(
                    surgery_df[['_mouse_norm'] + surgery_cols],
                    on='_mouse_norm',
                    how='left'
                )

                # Compute days post injury
                if 'surgery_date' in df.columns and 'date' in df.columns:
                    df['days_post_injury'] = (
                        pd.to_datetime(df['date']) - pd.to_datetime(df['surgery_date'])
                    ).dt.days
                    valid_dpi = df['days_post_injury'].notna().sum()
                    print(f"  days_post_injury computed for {valid_dpi} reaches")

                # Clean up temp columns
                df = df.drop(columns=['_mouse_norm'], errors='ignore')

                matched_surgery = df['surgery_date'].notna().sum() if 'surgery_date' in df.columns else 0
                print(f"  Surgery data matched: {matched_surgery}/{len(df)} reaches")
        else:
            print(f"\n[5/6] Tracking dir not found - skipping surgery metadata")
    else:
        print("\n[5/6] Surgery metadata disabled or no tracking_dir")

    # 6. Merge BrainGlobe connectomics data
    if include_brainglobe:
        print(f"\n[6/6] Loading BrainGlobe connectomics...")
        brainglobe_df = load_brainglobe_data(brainglobe_path, normalize_ids=True)

        if len(brainglobe_df) > 0:
            # Normalize mouse_id for merge
            df['_mouse_norm'] = df['mouse_id'].apply(normalize_mouse_id)

            # Select columns to merge (all brainglobe_ columns plus identifiers)
            brainglobe_cols = [c for c in brainglobe_df.columns
                              if c.startswith('brainglobe_') or c == 'total_cells']

            # Merge on normalized mouse_id
            df = df.merge(
                brainglobe_df[['mouse_id'] + brainglobe_cols].rename(columns={'mouse_id': '_mouse_norm'}),
                on='_mouse_norm',
                how='left'
            )

            # Clean up temp columns
            df = df.drop(columns=['_mouse_norm'], errors='ignore')

            matched_brainglobe = df['total_cells'].notna().sum() if 'total_cells' in df.columns else 0
            print(f"  BrainGlobe data matched: {matched_brainglobe}/{len(df)} reaches")
            print(f"  Region columns added: {len([c for c in df.columns if c.startswith('brainglobe_')])}")
        else:
            print(f"  No BrainGlobe data found")
    else:
        print("\n[6/6] BrainGlobe connectomics disabled")

    # Apply exclusion filter
    result = ReachDataFrame(df)
    if exclude_flagged:
        n_before = len(result)
        result = result.filter(exclude_flagged=True)
        n_excluded = n_before - len(result)
        if n_excluded > 0:
            print(f"\nExcluded {n_excluded} flagged reaches")

    # Summary
    print("\n" + "=" * 60)
    print("UNIFIED DATASET READY")
    print("=" * 60)
    print(f"  Total reaches: {len(result):,}")
    print(f"  Total columns: {len(result.df.columns)}")
    print(f"  Mice: {result.df['mouse_id'].nunique()}")
    if 'cohort' in result.df.columns:
        print(f"  Cohorts: {sorted(result.df['cohort'].dropna().unique())}")
    if 'timepoint' in result.df.columns:
        print(f"  Timepoints: {sorted(result.df['timepoint'].dropna().unique())}")

    return result


# =============================================================================
# DATABASE VERSIONING AND AUTO-UPDATE
# =============================================================================

DATABASE_VERSION = "1.0.0"  # Increment when schema changes


def get_source_fingerprint(data_dir: Path) -> dict:
    """
    Compute a fingerprint of all source files used for the database.

    Returns dict with file paths and their modification times.
    Used to determine if database needs rebuilding.
    """
    data_dir = Path(data_dir)
    fingerprint = {
        'version': DATABASE_VERSION,
        'data_dir': str(data_dir),
        'files': {}
    }

    # Find all relevant JSON files
    patterns = ['*_features.json', '*_reaches.json', '*_pellet_outcomes.json', '*_segments.json']

    for pattern in patterns:
        for fpath in data_dir.glob(pattern):
            # Store relative path and mtime
            rel_path = str(fpath.relative_to(data_dir))
            fingerprint['files'][rel_path] = fpath.stat().st_mtime

    fingerprint['n_files'] = len(fingerprint['files'])
    fingerprint['computed_at'] = datetime.now().isoformat()

    return fingerprint


def save_database_metadata(db_path: Path, fingerprint: dict, n_reaches: int, columns: list):
    """Save metadata sidecar file for the database."""
    meta_path = db_path.with_suffix('.meta.json')

    metadata = {
        'database_path': str(db_path),
        'database_version': DATABASE_VERSION,
        'built_at': datetime.now().isoformat(),
        'n_reaches': n_reaches,
        'n_columns': len(columns),
        'columns': columns,
        'source_fingerprint': fingerprint
    }

    with open(meta_path, 'w') as f:
        json.dump(metadata, f, indent=2)

    return meta_path


def load_database_metadata(db_path: Path) -> Optional[dict]:
    """Load metadata sidecar file for the database."""
    meta_path = db_path.with_suffix('.meta.json')

    if not meta_path.exists():
        return None

    try:
        with open(meta_path) as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return None


def check_database_current(db_path: Path, data_dir: Path) -> tuple:
    """
    Check if the database is current (doesn't need rebuilding).

    Returns:
        (is_current: bool, reason: str)
    """
    db_path = Path(db_path)
    data_dir = Path(data_dir)

    # Check if database exists
    if not db_path.exists():
        return False, "Database does not exist"

    # Load existing metadata
    metadata = load_database_metadata(db_path)
    if metadata is None:
        return False, "No metadata file found (rebuild to create)"

    # Check version
    if metadata.get('database_version') != DATABASE_VERSION:
        old_ver = metadata.get('database_version', 'unknown')
        return False, f"Version mismatch (database: {old_ver}, current: {DATABASE_VERSION})"

    # Compute current fingerprint
    current_fp = get_source_fingerprint(data_dir)
    stored_fp = metadata.get('source_fingerprint', {})

    # Compare file counts
    if current_fp['n_files'] != stored_fp.get('n_files', 0):
        return False, f"File count changed ({stored_fp.get('n_files', 0)} -> {current_fp['n_files']})"

    # Compare individual files
    stored_files = stored_fp.get('files', {})
    for rel_path, mtime in current_fp['files'].items():
        if rel_path not in stored_files:
            return False, f"New file: {rel_path}"
        if mtime > stored_files[rel_path]:
            return False, f"Modified: {rel_path}"

    # Check for deleted files
    for rel_path in stored_files:
        if rel_path not in current_fp['files']:
            return False, f"Deleted: {rel_path}"

    return True, f"Database current ({metadata.get('n_reaches', 0)} reaches, built {metadata.get('built_at', 'unknown')[:10]})"


def trigger_database_update(
    data_dir: Optional[Path] = None,
    tracking_dir: Optional[Path] = None,
    brainglobe_path: Optional[Path] = None,
    output_path: Optional[Path] = None,
    force: bool = False
) -> Optional[Path]:
    """
    Update the unified database if needed.

    This function is designed to be called after validation saves.
    It checks if the database needs rebuilding and only rebuilds if necessary.

    Args:
        data_dir: Pipeline data directory (default: from config)
        tracking_dir: Tracking spreadsheets directory (optional - provide explicitly)
        brainglobe_path: Path to BrainGlobe region_counts.csv (optional - auto-detect if available)
        output_path: Output database path (default: ./unified_reaches.parquet)
        force: Force rebuild even if current

    Returns:
        Path to database if rebuilt, None if skipped (already current)
    """
    from mousereach.config import Paths

    # Resolve paths
    data_dir = Path(data_dir) if data_dir else Paths.PROCESSING

    # Default output to current directory
    if output_path:
        output_path = Path(output_path)
    else:
        output_path = Path("./unified_reaches.parquet")

    # Tracking dir is optional - user must provide explicitly
    # No auto-detection of external directories for portability
    if tracking_dir:
        tracking_dir = Path(tracking_dir)

    # Check if rebuild needed
    if not force:
        is_current, reason = check_database_current(output_path, data_dir)
        if is_current:
            print(f"Database up-to-date: {reason}")
            return None
        else:
            print(f"Database needs rebuild: {reason}")

    # Build the database
    print("\nRebuilding unified database...")
    reach_data = build_unified_reach_data(
        data_dir=data_dir,
        tracking_dir=tracking_dir,
        brainglobe_path=brainglobe_path,
        use_features=True,
        exclude_flagged=True,
        include_surgery=True,
        include_brainglobe=True
    )

    if len(reach_data) == 0:
        print("No data found.")
        return None

    df = reach_data.df

    # Save database
    suffix = output_path.suffix.lower()
    if suffix == '.parquet':
        df.to_parquet(output_path, index=False)
    elif suffix == '.csv':
        df.to_csv(output_path, index=False)
    else:
        df.to_parquet(output_path.with_suffix('.parquet'), index=False)
        output_path = output_path.with_suffix('.parquet')

    # Save metadata
    fingerprint = get_source_fingerprint(data_dir)
    meta_path = save_database_metadata(output_path, fingerprint, len(df), df.columns.tolist())

    print(f"Database updated: {output_path}")
    print(f"Metadata saved: {meta_path}")
    print(f"  {len(df):,} reaches, {len(df.columns)} columns")

    return output_path
