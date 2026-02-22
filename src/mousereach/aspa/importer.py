"""
mousereach.aspa.importer - Import old ASPA Post-Processing xlsx files into ASPA.db.

Old xlsx files live in:
    Analyzed/{cohort}/Post-Processing/*.xlsx   (on NAS, e.g. X:/! DLC Output/Analyzed/H/Post-Processing/)

Each xlsx has columns:
    Swipe_num, Timestamp, Swipe Duration, Pellet #, Reach Outcome,
    s_idx, e_idx, plus kinematic columns (breadth_mm, reach_mm, distance_mm,
    speed_mm_s, area_mm2, pillar_visible)

CLI:
    mousereach-aspa-import [--cohort H] [--all] [--db-path ASPA.db]
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Optional

from mousereach.aspa.database import ensure_tables, get_connection, get_db_path
from mousereach.config import Paths


# ---------------------------------------------------------------------------
# Outcome normalisation
# ---------------------------------------------------------------------------

_OUTCOME_MAP = {
    # ASPA raw outcome -> canonical label
    "Pellet Eaten":             "retrieved",
    "Swipe Missed":             "missed",
    "Swiping with no Pellet":   "no_pellet",
    "No Pellet":                "no_pellet",
    "Pellet Displaced":         "displaced",
    "Pellet Dropped":           "dropped",
}


def normalise_outcome(raw: str) -> str:
    """Convert raw ASPA outcome string to canonical label."""
    if not isinstance(raw, str):
        return "unknown"
    raw_stripped = raw.strip()
    return _OUTCOME_MAP.get(raw_stripped, raw_stripped.lower().replace(" ", "_"))


# ---------------------------------------------------------------------------
# Video ID parsing from xlsx filename
# ---------------------------------------------------------------------------

def parse_video_id_from_filename(filename: str) -> str:
    """Extract video_id from xlsx filename.

    Expected patterns (best-effort):
        20210315_H0101_P1_PostProcessed.xlsx  -> 20210315_H0101_P1
        H0101_20210315_P1.xlsx                -> H0101_20210315_P1
        H0101.xlsx                            -> H0101

    Falls back to stem (filename without extension).
    """
    stem = Path(filename).stem
    # Strip common suffixes that are not part of the video ID
    for suffix in ("_PostProcessed", "_post_processed", "_PP", "_Results",
                   "_results", "_Analysis", "_analysis"):
        if stem.endswith(suffix):
            stem = stem[: -len(suffix)]
            break
    return stem


# ---------------------------------------------------------------------------
# xlsx importer
# ---------------------------------------------------------------------------

_KINEMATIC_COLUMNS = {
    # xlsx column name -> db column name
    "breadth_mm":     "breadth_mm",
    "reach_mm":       "reach_mm",
    "distance_mm":    "distance_mm",
    "speed_mm_s":     "speed_mm_s",
    "area_mm2":       "area_mm2",
    "pillar_visible": "pillar_visible",
    # common alternate spellings
    "Breadth_mm":     "breadth_mm",
    "Reach_mm":       "reach_mm",
    "Distance_mm":    "distance_mm",
    "Speed_mm_s":     "speed_mm_s",
    "Area_mm2":       "area_mm2",
    "Pillar_Visible": "pillar_visible",
}


def _safe_float(val) -> Optional[float]:
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _safe_int(val) -> Optional[int]:
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def import_xlsx(xlsx_path: Path, cohort: str, db_path: Path = None,
                dry_run: bool = False) -> int:
    """Import a single xlsx file into ASPA.db.

    Args:
        xlsx_path: Path to the .xlsx file.
        cohort:    Cohort label (e.g. "H").
        db_path:   Override DB path. If None, uses get_db_path().
        dry_run:   If True, parse but do not write to DB.

    Returns:
        Number of reach rows imported (or would have been imported in dry_run).

    Raises:
        ImportError if openpyxl is not installed.
        FileNotFoundError if xlsx_path does not exist.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required to import xlsx files.\n"
            "Install it with: pip install openpyxl"
        )

    if not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx not found: {xlsx_path}")

    video_id = parse_video_id_from_filename(xlsx_path.name)

    # Parse animal_id from video_id (first token after date if present)
    parts = video_id.split("_")
    animal_id = "UNKNOWN"
    for part in parts:
        if re.match(r"^[A-Za-z]+\d+", part):
            animal_id = part
            break

    # Parse session_date - look for YYYYMMDD token
    session_date = None
    for part in parts:
        if re.match(r"^\d{8}$", part):
            session_date = part
            break

    # Parse tray_type and position
    tray_type = None
    position = None
    for part in parts:
        m = re.match(r"^([PEF])(\d+)$", part)
        if m:
            tray_type = m.group(1)
            position = int(m.group(2))
            break

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        wb.close()
        print(f"  [!] Empty xlsx, skipping: {xlsx_path.name}")
        return 0

    headers = [str(h).strip() if h is not None else "" for h in header_row]

    def col_idx(name: str) -> Optional[int]:
        try:
            return headers.index(name)
        except ValueError:
            return None

    # Map columns
    idx_swipe_num    = col_idx("Swipe_num")
    idx_timestamp    = col_idx("Timestamp")
    idx_duration     = col_idx("Swipe Duration")
    idx_pellet_num   = col_idx("Pellet #")
    idx_outcome      = col_idx("Reach Outcome")
    idx_s_idx        = col_idx("s_idx")
    idx_e_idx        = col_idx("e_idx")

    # Kinematic columns (optional)
    kinematic_idxs = {}
    for xlsx_col, db_col in _KINEMATIC_COLUMNS.items():
        i = col_idx(xlsx_col)
        if i is not None and db_col not in kinematic_idxs:
            kinematic_idxs[db_col] = i

    rows_to_insert = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        def get(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        outcome_raw = str(get(idx_outcome)).strip() if get(idx_outcome) is not None else None
        outcome     = normalise_outcome(outcome_raw) if outcome_raw else None

        reach_row = {
            "video_id":       video_id,
            "cohort":         cohort,
            "animal_id":      animal_id,
            "reach_num":      _safe_int(get(idx_swipe_num)),
            "start_frame":    _safe_int(get(idx_s_idx)),
            "end_frame":      _safe_int(get(idx_e_idx)),
            "duration_s":     _safe_float(get(idx_duration)),
            "pellet_num":     _safe_int(get(idx_pellet_num)),
            "outcome":        outcome,
            "outcome_raw":    outcome_raw,
            "breadth_mm":     _safe_float(row[kinematic_idxs["breadth_mm"]])     if "breadth_mm"  in kinematic_idxs else None,
            "reach_mm":       _safe_float(row[kinematic_idxs["reach_mm"]])       if "reach_mm"    in kinematic_idxs else None,
            "distance_mm":    _safe_float(row[kinematic_idxs["distance_mm"]])    if "distance_mm" in kinematic_idxs else None,
            "speed_mm_s":     _safe_float(row[kinematic_idxs["speed_mm_s"]])     if "speed_mm_s"  in kinematic_idxs else None,
            "area_mm2":       _safe_float(row[kinematic_idxs["area_mm2"]])       if "area_mm2"    in kinematic_idxs else None,
            "pillar_visible": _safe_int(row[kinematic_idxs["pillar_visible"]])   if "pillar_visible" in kinematic_idxs else None,
        }
        rows_to_insert.append(reach_row)

    wb.close()

    if not rows_to_insert:
        print(f"  [!] No data rows found in: {xlsx_path.name}")
        return 0

    if dry_run:
        print(f"  [DRY-RUN] Would import {len(rows_to_insert)} reaches from {xlsx_path.name}")
        return len(rows_to_insert)

    conn = get_connection(db_path)
    try:
        with conn:
            # Upsert video record
            conn.execute(
                """
                INSERT INTO videos (video_id, cohort, animal_id, session_date,
                                    tray_type, position, has_aspa_results)
                VALUES (?, ?, ?, ?, ?, ?, 1)
                ON CONFLICT(video_id) DO UPDATE SET
                    has_aspa_results = 1,
                    cohort           = excluded.cohort,
                    animal_id        = excluded.animal_id,
                    session_date     = COALESCE(excluded.session_date, session_date),
                    tray_type        = COALESCE(excluded.tray_type,    tray_type),
                    position         = COALESCE(excluded.position,     position)
                """,
                (video_id, cohort, animal_id, session_date, tray_type, position),
            )

            # Delete existing aspa_reaches for this video to allow re-import
            conn.execute("DELETE FROM aspa_reaches WHERE video_id = ?", (video_id,))

            # Bulk insert
            conn.executemany(
                """
                INSERT INTO aspa_reaches
                    (video_id, cohort, animal_id, reach_num, start_frame, end_frame,
                     duration_s, pellet_num, outcome, outcome_raw,
                     breadth_mm, reach_mm, distance_mm, speed_mm_s, area_mm2, pillar_visible)
                VALUES
                    (:video_id, :cohort, :animal_id, :reach_num, :start_frame, :end_frame,
                     :duration_s, :pellet_num, :outcome, :outcome_raw,
                     :breadth_mm, :reach_mm, :distance_mm, :speed_mm_s, :area_mm2, :pillar_visible)
                """,
                rows_to_insert,
            )
    finally:
        conn.close()

    return len(rows_to_insert)


# ---------------------------------------------------------------------------
# Discovery helpers
# ---------------------------------------------------------------------------

def find_post_processing_dirs(nas_root: Path, cohort: Optional[str] = None):
    """Yield (cohort, post_processing_dir) for all cohorts under nas_root/Analyzed/.

    Args:
        nas_root: Root of NAS output tree (contains Analyzed/ subfolder).
        cohort:   If given, restrict to that cohort only.
    """
    analyzed = nas_root / "Analyzed"
    if not analyzed.exists():
        print(f"[!] Analyzed directory not found: {analyzed}")
        return

    if cohort:
        candidate = analyzed / cohort / "Post-Processing"
        if candidate.exists():
            yield cohort, candidate
        else:
            print(f"[!] Post-Processing dir not found for cohort {cohort}: {candidate}")
    else:
        for cohort_dir in sorted(analyzed.iterdir()):
            if not cohort_dir.is_dir():
                continue
            pp = cohort_dir / "Post-Processing"
            if pp.exists():
                yield cohort_dir.name, pp


# ---------------------------------------------------------------------------
# CLI main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Import old ASPA Post-Processing xlsx files into ASPA.db."
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--cohort", metavar="COHORT",
                       help="Import single cohort (e.g. H)")
    group.add_argument("--all", action="store_true",
                       help="Import all cohorts found under Analyzed/")

    parser.add_argument("--db-path", metavar="PATH",
                        help="Override ASPA.db path (default: ASPA_DB_PATH env or Y:/2_Connectome/Behavior/MouseReach_Pipeline/ASPA.db)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse xlsx files but do not write to database")

    args = parser.parse_args()

    db_path = Path(args.db_path) if args.db_path else get_db_path()

    if not args.dry_run:
        ensure_tables(db_path)

    from mousereach.config import Paths, require_nas_drive
    try:
        nas_root = require_nas_drive() / "! DLC Output"
    except Exception as e:
        print(f"[FAIL] Could not resolve NAS root: {e}")
        print("       Set MouseReach_NAS_DRIVE or run mousereach-setup.")
        sys.exit(1)

    cohort_filter = args.cohort if not args.all else None
    total_files  = 0
    total_rows   = 0
    errors       = 0

    for cohort, pp_dir in find_post_processing_dirs(nas_root, cohort_filter):
        xlsx_files = sorted(pp_dir.glob("*.xlsx"))
        if not xlsx_files:
            print(f"  [!] No xlsx files in {pp_dir}")
            continue

        print(f"\nCohort {cohort}: {len(xlsx_files)} xlsx file(s) in {pp_dir}")

        for xlsx_path in xlsx_files:
            try:
                n = import_xlsx(xlsx_path, cohort, db_path=db_path,
                                dry_run=args.dry_run)
                print(f"  [OK] {xlsx_path.name} -> {n} reaches")
                total_files += 1
                total_rows  += n
            except Exception as exc:
                print(f"  [FAIL] {xlsx_path.name}: {exc}")
                errors += 1

    print(f"\nDone. Files processed: {total_files}, Reach rows: {total_rows}, Errors: {errors}")
    if args.dry_run:
        print("(dry-run: no data written)")


if __name__ == "__main__":
    main()
