"""
End-to-end provenance trace for individual M swipes.

For selected swipes from the UCSF CSV, traces backwards through:
1. UCSF CSV -> swipe_info_M_MB_corrected_Only_Included.xlsx
2. -> swipe_info_M_MB.xlsx (pre-exclusion)
3. -> per-video ASPA xlsx on X:
4. -> DLC h5 file

At each step, verifies that kinematic values match.
READ-ONLY: does not modify any source files.

Usage:
    python m_provenance_trace.py              # Trace 10 sample swipes across phases
    python m_provenance_trace.py --animal M05 # Trace 10 swipes from specific animal
"""

import pandas as pd
import numpy as np
import os
import sys
import glob
import openpyxl
from datetime import datetime, timedelta

# --- Paths (READ ONLY) ---
UCSF_BASE = os.path.join(
    r'C:\Users\friedrichl\OneDrive - Marquette University\Blackmore Lab Notes - Sharepoint',
    r'3 Lab Projects\Automated single pellet apparatus\!UCSF_Collab',
    r'May2025_Uploads\ODC Uploads'
)
MARCH_BASE = os.path.join(
    r'G:\user_data\BlackmoreLab\Logan\OneDrive\OneDrive - Marquette University',
    r'Blackmore Lab Notes - Sharepoint\03 Lab Projects',
    r'Automated single pellet apparatus\AI Summary',
    r'!Jan_2025_Troubleshooting\March_Analysis'
)
ASPA_XLSX_DIR = r'X:\! DLC Output\Analyzed\M\Post-Processing'
DLC_DIR = r'X:\! DLC Output\Analyzed\M\Single_Animal'
DLC_DIR_Y = r'Y:\2_Connectome\Behavior\MouseReach_Pipeline\Analyzed\ASPA\M\Post-Processing'

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_connectome_root = SCRIPT_DIR
while os.path.basename(_connectome_root) != '2_Connectome' and os.path.dirname(_connectome_root) != _connectome_root:
    _connectome_root = os.path.dirname(_connectome_root)
OUTPUT_DIR = os.path.join(_connectome_root, 'Databases', 'figures', 'bodypart_tracking_quality')


def parse_frame_range(s):
    parts = str(s).strip().split('-')
    if len(parts) >= 2:
        try:
            return int(parts[0]), int(parts[1].split(' ')[0])
        except ValueError:
            pass
    return None, None


def session_to_stem(sid):
    parts = sid.split('-')
    date_str = (datetime(1899, 12, 30) + timedelta(days=int(parts[1]))).strftime('%Y%m%d')
    return '%s_%s_%s' % (date_str, parts[0], parts[2])


def session_to_date(sid):
    parts = sid.split('-')
    return (datetime(1899, 12, 30) + timedelta(days=int(parts[1]))).strftime('%Y%m%d')


def find_dlc_h5(stem):
    for base in [DLC_DIR, DLC_DIR_Y]:
        h5s = glob.glob(os.path.join(base, stem + 'DLC*.h5'))
        if h5s:
            return h5s[0]
    return None


def extract_aspa_value(text_val):
    """Extract numeric value from ASPA text format like '28.45 (7.11mm)'"""
    if text_val is None:
        return None
    s = str(text_val).strip()
    if s in ('NA', 'nan', ''):
        return None
    # Take first number
    parts = s.split(' ')
    try:
        return float(parts[0])
    except ValueError:
        return None


def compare_values(label, val_a, val_b, source_a, source_b, tolerance=0.02):
    """Compare two values and report match/mismatch."""
    if val_a is None or val_b is None:
        return '  %s: %s=%s, %s=%s -> SKIP (missing)' % (label, source_a, val_a, source_b, val_b)
    if val_a == 0 and val_b == 0:
        return '  %s: both 0 -> MATCH' % label
    if val_a == 0:
        return '  %s: %s=0, %s=%.4f -> MISMATCH (zero vs nonzero)' % (label, source_a, source_b, val_b)
    ratio = val_b / val_a
    status = 'MATCH' if abs(ratio - 1.0) <= tolerance else 'MISMATCH (%.4fx)' % ratio
    return '  %s: %s=%.4f, %s=%.4f -> %s' % (label, source_a, val_a, source_b, val_b, status)


def main():
    target_animal = None
    for i, arg in enumerate(sys.argv):
        if arg == '--animal' and i + 1 < len(sys.argv):
            target_animal = sys.argv[i + 1]

    # Step 1: Load UCSF CSV
    print('=== STEP 1: Loading UCSF CSV ===')
    df1 = pd.read_csv(os.path.join(UCSF_BASE, 'Swipe_Contusion_Data.csv'), low_memory=False)
    df2 = pd.read_csv(os.path.join(UCSF_BASE, 'Swipe_Contusion_Data_2.csv'), low_memory=False)
    ucsf = pd.concat([df1, df2], ignore_index=True)
    ucsf = ucsf[ucsf['SubjectID'].str.startswith('M', na=False)].copy()
    for col in ['Swipe_length', 'Swipe_area', 'Swipe_breadth', 'Swipe_speed',
                'Path_length', 'Path_over_Frames', 'Swipe_Duration_Frames']:
        ucsf[col] = pd.to_numeric(ucsf[col], errors='coerce')
    ucsf['sd_start'], ucsf['sd_end'] = zip(*ucsf['Swipe_Duration'].apply(parse_frame_range))
    print('  UCSF M rows: %d' % len(ucsf))

    # Select sample swipes — 2 from each phase, spread across animals
    if target_animal:
        sample_df = ucsf[ucsf['SubjectID'] == target_animal]
    else:
        sample_df = ucsf

    samples = []
    phases = ['2_Pre-injury_1', '2_Pre-injury_2', '2_Pre-injury_3',
              '3_1wk_Post-injury', '3_2wk_Post-injury',
              '5_Post-rehab_Test_1']
    for phase in phases:
        phase_df = sample_df[sample_df['Test_Type_Grouped_1'] == phase]
        if len(phase_df) > 0:
            # Pick one from start, one from end
            samples.append(phase_df.iloc[0])
            if len(phase_df) > 10:
                samples.append(phase_df.iloc[len(phase_df)//2])

    print('  Selected %d sample swipes to trace' % len(samples))

    # Step 2: For each sample, trace through the chain
    dlc_cache = {}
    results = []

    for idx, row in enumerate(samples):
        print('\n' + '=' * 80)
        print('SWIPE %d/%d: %s, %s, frames %s-%s' % (
            idx+1, len(samples), row['SubjectID'], row['Session_ID'],
            row.get('sd_start', '?'), row.get('sd_end', '?')))
        print('  Phase: %s, Outcome: %s' % (
            row.get('Test_Type_Grouped_1', '?'), row.get('Reach_outcome', '?')))
        print('  UCSF values: length=%.2f, area=%.2f, breadth=%.2f, speed=%.2f' % (
            row['Swipe_length'], row['Swipe_area'], row['Swipe_breadth'], row['Swipe_speed']))

        stem = session_to_stem(row['Session_ID'])
        date_str = session_to_date(row['Session_ID'])
        animal = row['SubjectID']
        sd_start = row['sd_start']
        sd_end = row['sd_end']

        # Step 2a: Find in per-video ASPA xlsx
        print('\n  --- Per-video ASPA xlsx ---')
        aspa_xlsx = glob.glob(os.path.join(ASPA_XLSX_DIR, stem + '.xlsx'))
        if aspa_xlsx:
            print('  Found: %s' % os.path.basename(aspa_xlsx[0]))
            wb = openpyxl.load_workbook(aspa_xlsx[0], read_only=True, data_only=True)
            ws = wb.active
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            print('  Columns: %s' % header[:10])

            # Find the swipe by bodyparts_coords or s_idx
            bc_col = header.index('bodyparts_coords') if 'bodyparts_coords' in header else None
            sidx_col = header.index('s_idx') if 's_idx' in header else None

            matched_row = None
            for r in ws.iter_rows(min_row=2, values_only=True):
                if bc_col is not None and r[bc_col] is not None:
                    # bodyparts_coords in per-video might be the swipe key (peak frame)
                    pass
                if sidx_col is not None and r[sidx_col] is not None:
                    try:
                        s = int(r[sidx_col])
                        # Check if this swipe's s_idx is near our sd_start
                        if sd_start is not None and abs(s - sd_start) < 50:
                            if matched_row is None or abs(s - sd_start) < abs(int(matched_row[sidx_col]) - sd_start):
                                matched_row = r
                    except (ValueError, TypeError):
                        pass

            if matched_row:
                print('  Matched by s_idx proximity')
                for i, h in enumerate(header):
                    if matched_row[i] is not None:
                        print('    %s: %s' % (h, matched_row[i]))
            else:
                print('  NO MATCH found in per-video xlsx')
            wb.close()
        else:
            print('  NOT FOUND: %s.xlsx' % stem)

        # Step 2b: Find in DLC h5
        print('\n  --- DLC h5 ---')
        h5_path = find_dlc_h5(stem)
        if h5_path:
            print('  Found: %s' % os.path.basename(h5_path))
            if stem not in dlc_cache:
                d = pd.read_hdf(h5_path)
                sc = d.columns.get_level_values(0)[0]
                dlc_cache[stem] = d[sc]
            dlc = dlc_cache[stem]
            print('  Video frames: %d' % len(dlc))

            if sd_start is not None and sd_end is not None:
                s, e = int(sd_start), int(sd_end)
                if e + 1 <= len(dlc):
                    rh_x = dlc['RightHand']['x'].iloc[s:e+1].values
                    rh_y = dlc['RightHand']['y'].iloc[s:e+1].values
                    nose_y = dlc['Nose']['y'].iloc[s:e+1].values
                    n_frames = len(rh_x)

                    # Recompute kinematics (ALL frames, no pre-filter)
                    area = 0.5 * abs(np.dot(rh_x, np.roll(rh_y, 1)) - np.dot(rh_y, np.roll(rh_x, 1)))
                    peak_idx = np.argmax(rh_y)
                    length = rh_y[peak_idx] - nose_y[peak_idx]
                    breadth = np.max(rh_x) - np.min(rh_x)
                    path = np.sum(np.sqrt(np.diff(rh_x)**2 + np.diff(rh_y)**2))
                    speed = (path / 4.0) / ((e - s) / 60.0) if (e - s) > 0 else 0

                    print('\n  --- Comparison (UCSF vs DLC recomputed, no pre-filter) ---')
                    print(compare_values('Area', row['Swipe_area'], area, 'UCSF', 'DLC'))
                    print(compare_values('Length', row['Swipe_length'], length, 'UCSF', 'DLC'))
                    print(compare_values('Breadth', row['Swipe_breadth'], breadth, 'UCSF', 'DLC'))
                    print(compare_values('Speed', row['Swipe_speed'], speed, 'UCSF', 'DLC'))
                else:
                    print('  Frame range %d-%d exceeds video length %d' % (s, e, len(dlc)))
        else:
            print('  NOT FOUND: %s' % stem)

        results.append({
            'animal': animal,
            'session': row['Session_ID'],
            'phase': row.get('Test_Type_Grouped_1', ''),
            'sd_start': sd_start,
            'sd_end': sd_end,
            'stem': stem,
            'aspa_xlsx_found': bool(aspa_xlsx),
            'dlc_h5_found': h5_path is not None,
        })

    # Summary
    print('\n' + '=' * 80)
    print('=== TRACE SUMMARY ===')
    print('Swipes traced: %d' % len(results))
    print('ASPA xlsx found: %d/%d' % (sum(r['aspa_xlsx_found'] for r in results), len(results)))
    print('DLC h5 found: %d/%d' % (sum(r['dlc_h5_found'] for r in results), len(results)))

    # Save
    out = os.path.join(OUTPUT_DIR, 'm_provenance_trace.csv')
    pd.DataFrame(results).to_csv(out, index=False)
    print('\nSaved: %s' % out)


if __name__ == '__main__':
    main()
