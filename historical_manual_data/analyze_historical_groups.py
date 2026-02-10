"""
Historical Behavioral Data Analysis - Manual Pellet Scores

Computes Eaten % and Contacted % from raw manual pellet scores
for all old injury groups (non-CNT, non-ENCR) across four time windows:
  - Final 3: Last 3 training days before injury
  - Immediate Post: 1 Week Post-Injury test(s)
  - 2-4 Post: Weeks 2-4 Post-Injury
  - Last 2 Days: Last 2 Pillar tray days in continuous rehab block

Learner criterion: Final 3 avg eaten% > 5% (excludes non-learners).
"""

import os
import re
import numpy as np
import openpyxl
from collections import defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Group files and their descriptive injury types
GROUP_FILES = {
    'ABS3': ('ABS3.xlsx', 'Early Study'),
    'G': ('G - Transection.xlsx', 'Transection'),
    'H': ('H - Transection.xlsx', 'Transection'),
    'K': ('K - Contusion 70kd.xlsx', 'Contusion 70kd'),
    'L': ('L - Contusion - 50kd.xlsx', 'Contusion 50kd'),
    'M': ('M - Contusion - 60kd.xlsx', 'Contusion 60kd'),
    'OptD': ('OptD - Rehab 1 - pyramidotomy.xlsx', 'Pyramidotomy'),
}

# Tray column offsets: each tray block is 24 columns (tray#, 20 pellets, displaced, eaten, hit)
# Pellet positions start at col 7 for Tray 1
TRAY_OFFSETS_4 = [
    (7, 27),    # Tray 1: pellets at cols 7-26
    (31, 51),   # Tray 2: pellets at cols 31-50
    (55, 75),   # Tray 3: pellets at cols 55-74
    (79, 99),   # Tray 4: pellets at cols 79-98
]

TRAY_OFFSETS_2 = [
    (7, 27),    # Tray 1: pellets at cols 7-26
    (31, 51),   # Tray 2: pellets at cols 31-50
]

# Learner criterion: exclude animals with Final 3 avg eaten% <= this threshold
LEARNER_EATEN_THRESHOLD = 5.0


def score_tray(pellet_values):
    """Compute Eaten% and Contacted% from 20 pellet values.

    Scoring: 0=untouched, 1=displaced/contacted, 2=eaten
    Contacted = value >= 1 (displaced + eaten)
    Eaten = value == 2
    """
    valid = [v for v in pellet_values if v is not None and v != 'N/A' and isinstance(v, (int, float))]
    if len(valid) == 0:
        return None, None
    n_pellets = len(valid)
    contacted = sum(1 for v in valid if v >= 1)
    eaten = sum(1 for v in valid if v >= 2)
    return (eaten / n_pellets) * 100, (contacted / n_pellets) * 100


def is_valid_tray(pellet_values):
    """Check if a tray has valid data (not all N/A or None)."""
    for v in pellet_values:
        if v is not None and v != 'N/A' and isinstance(v, (int, float)):
            return True
    return False


def classify_test_type(test_type_name, all_test_types_sorted, first_post_idx):
    """Classify a test type into a time window.

    Returns one of: 'final_3', 'immediate_post', '2_4_post', 'last_2', or None
    """
    tt_lower = test_type_name.lower()

    # Check if it's a post-injury test
    if '1 week post' in tt_lower:
        return 'immediate_post'

    week_match = re.search(r'(\d+)\s*week\s*post', tt_lower)
    if week_match:
        week_num = int(week_match.group(1))
        if 2 <= week_num <= 4:
            return '2_4_post'
        # Weeks > 4 don't fall into our defined windows
        return None

    return None


def get_time_windows(test_types_sorted, test_meta):
    """Given sorted test types and metadata, determine which belong to each time window.

    Last 2: Uses last 2 Pillar tray days in continuous rehab block (before any >5-day gap).
    """
    windows = {
        'final_3': [],
        'immediate_post': [],
        '2_4_post': [],
        'last_2': [],
    }

    # Find the index of the first post-injury test
    first_post_idx = None
    for i, tt in enumerate(test_types_sorted):
        if 'post' in tt.lower() and 'injury' in tt.lower():
            first_post_idx = i
            break

    if first_post_idx is None:
        return windows  # No post-injury data

    # Final 3: last 3 training test types before post-injury
    pre_injury_types = test_types_sorted[:first_post_idx]
    windows['final_3'] = pre_injury_types[-3:] if len(pre_injury_types) >= 3 else pre_injury_types

    # Classify post-injury tests
    for tt in test_types_sorted[first_post_idx:]:
        tt_lower = tt.lower()
        if '1 week post' in tt_lower:
            windows['immediate_post'].append(tt)
        else:
            week_match = re.search(r'(\d+)\s*week\s*post', tt_lower)
            if week_match:
                week_num = int(week_match.group(1))
                if 2 <= week_num <= 4:
                    windows['2_4_post'].append(tt)

    # Last 2: find last 2 Pillar days in continuous rehab (no gap > 5 days)
    rehab_tts = [tt for tt in test_types_sorted if 'rehab' in tt.lower()]

    if not rehab_tts:
        windows['last_2'] = test_types_sorted[-2:]
        return windows

    # Split rehab into continuous blocks (gap > 5 days = new block)
    blocks = [[rehab_tts[0]]]
    for i in range(1, len(rehab_tts)):
        prev_d = test_meta.get(rehab_tts[i-1], {}).get('date')
        curr_d = test_meta.get(rehab_tts[i], {}).get('date')
        if prev_d and curr_d:
            try:
                gap = (curr_d - prev_d).days
            except (TypeError, AttributeError):
                gap = 1
            if 5 < gap < 100:  # Real breaks; ignore date-entry errors
                blocks.append([rehab_tts[i]])
                continue
        blocks[-1].append(rehab_tts[i])

    # Use the first (main) rehab block - find last 2 Pillar days in it
    main_block = blocks[0]
    pillar_in_main = [tt for tt in main_block
                      if test_meta.get(tt, {}).get('tray', '').lower().startswith('p')]

    if len(pillar_in_main) >= 2:
        windows['last_2'] = pillar_in_main[-2:]
    elif len(pillar_in_main) == 1:
        windows['last_2'] = pillar_in_main
    else:
        windows['last_2'] = main_block[-2:] if len(main_block) >= 2 else main_block

    return windows


def read_group_data(filepath, group_name):
    """Read all data from a group's ENTER DATA HERE sheet.

    Returns:
        data: {animal_id: {test_type: {'eaten_pct': float, 'contacted_pct': float}}}
        sorted test types list
        test_meta: {test_type: {'date': datetime, 'tray': str}}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['1 - ENTER DATA HERE']

    # Detect number of columns to determine tray layout
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    n_cols = len(headers)

    if n_cols > 80:
        tray_offsets = TRAY_OFFSETS_4
    else:
        tray_offsets = TRAY_OFFSETS_2

    # Read all data rows
    data = defaultdict(lambda: defaultdict(list))  # animal -> test_type -> list of (eaten%, contacted%)
    all_test_types = set()
    test_meta = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if vals[0] is None:
            continue

        test_type = str(vals[1])
        animal = vals[5]
        all_test_types.add(test_type)

        if test_type not in test_meta:
            test_meta[test_type] = {'date': vals[0], 'tray': str(vals[2])}

        # Compute metrics from raw pellet scores across valid trays
        tray_eaten = []
        tray_contacted = []

        for start, end in tray_offsets:
            if start >= len(vals):
                continue
            pellets = vals[start:end]
            if not is_valid_tray(pellets):
                continue
            eaten_pct, contacted_pct = score_tray(pellets)
            if eaten_pct is not None:
                tray_eaten.append(eaten_pct)
                tray_contacted.append(contacted_pct)

        if tray_eaten:
            avg_eaten = np.mean(tray_eaten)
            avg_contacted = np.mean(tray_contacted)
            data[animal][test_type] = {'eaten_pct': avg_eaten, 'contacted_pct': avg_contacted}

    wb.close()
    return data, sorted(all_test_types), test_meta


def get_learners(data, final3_tts):
    """Return set of animals that meet the learner criterion (Final 3 eaten% > threshold)."""
    learners = set()
    for animal, tests in data.items():
        if final3_tts:
            e_vals = [tests[tt]['eaten_pct'] for tt in final3_tts if tt in tests]
            if e_vals and np.mean(e_vals) > LEARNER_EATEN_THRESHOLD:
                learners.add(animal)
        else:
            learners.add(animal)
    return learners


def compute_window_stats(data, test_types_in_window, learners=None):
    """Compute mean +/- SEM across animals for a time window.

    For each animal: average across all test days in the window.
    Then compute mean and SEM across animals.
    If learners is provided, only includes animals in that set.
    """
    if not test_types_in_window:
        return None, None, None, None, 0

    animal_eaten = []
    animal_contacted = []

    for animal, tests in data.items():
        if learners is not None and animal not in learners:
            continue
        eaten_vals = []
        contacted_vals = []
        for tt in test_types_in_window:
            if tt in tests:
                eaten_vals.append(tests[tt]['eaten_pct'])
                contacted_vals.append(tests[tt]['contacted_pct'])

        if eaten_vals:
            animal_eaten.append(np.mean(eaten_vals))
            animal_contacted.append(np.mean(contacted_vals))

    if not animal_eaten:
        return None, None, None, None, 0

    n = len(animal_eaten)
    eaten_mean = np.mean(animal_eaten)
    eaten_sem = np.std(animal_eaten, ddof=1) / np.sqrt(n) if n > 1 else 0
    contacted_mean = np.mean(animal_contacted)
    contacted_sem = np.std(animal_contacted, ddof=1) / np.sqrt(n) if n > 1 else 0

    return eaten_mean, eaten_sem, contacted_mean, contacted_sem, n


def main():
    print("=" * 110)
    print("HISTORICAL BEHAVIORAL DATA ANALYSIS - Manual Pellet Scores")
    print(f"  Learner criterion: Final 3 avg eaten% > {LEARNER_EATEN_THRESHOLD}%")
    print("=" * 110)

    all_results = []

    for group_name, (filename, injury_type) in GROUP_FILES.items():
        filepath = os.path.join(SCRIPT_DIR, filename)
        if not os.path.exists(filepath):
            print(f"\nWARNING: {filename} not found, skipping {group_name}")
            continue

        print(f"\n{'-' * 110}")
        print(f"Processing {group_name} ({injury_type}): {filename}")

        data, test_types, test_meta = read_group_data(filepath, group_name)
        windows = get_time_windows(test_types, test_meta)
        learners = get_learners(data, windows['final_3'])
        n_excluded = len(data) - len(learners)

        print(f"  Animals: {len(data)} total, {len(learners)} learners, {n_excluded} excluded")
        print(f"  Test types: {len(test_types)}")
        print(f"  Time window mapping:")
        for wname, ttypes in windows.items():
            print(f"    {wname:20s}: {', '.join(ttypes) if ttypes else '(none)'}")

        for window_name, window_label in [
            ('final_3', 'Final 3'),
            ('immediate_post', 'Immediate Post'),
            ('2_4_post', '2-4 Post'),
            ('last_2', 'Last 2 Days'),
        ]:
            eaten_mean, eaten_sem, contacted_mean, contacted_sem, n = compute_window_stats(
                data, windows[window_name], learners=learners
            )
            all_results.append({
                'group': group_name,
                'injury_type': injury_type,
                'window': window_label,
                'eaten_mean': eaten_mean,
                'eaten_sem': eaten_sem,
                'contacted_mean': contacted_mean,
                'contacted_sem': contacted_sem,
                'n_animals': n,
                'test_types_used': '; '.join(windows[window_name]),
            })

    # Print summary table
    print(f"\n\n{'=' * 110}")
    print("SUMMARY RESULTS")
    print(f"{'=' * 110}")
    print(f"{'Group':6s} {'Injury Type':15s} {'Window':16s} {'Eaten % (M +/- SEM)':22s} {'Contacted % (M +/- SEM)':26s} {'N':4s}")
    print(f"{'-' * 6} {'-' * 15} {'-' * 16} {'-' * 22} {'-' * 26} {'-' * 4}")

    for r in all_results:
        if r['eaten_mean'] is not None:
            eaten_str = f"{r['eaten_mean']:5.1f} +/- {r['eaten_sem']:4.1f}"
            contacted_str = f"{r['contacted_mean']:5.1f} +/- {r['contacted_sem']:4.1f}"
        else:
            eaten_str = "  N/A"
            contacted_str = "  N/A"
        print(f"{r['group']:6s} {r['injury_type']:15s} {r['window']:16s} {eaten_str:22s} {contacted_str:26s} {r['n_animals']:4d}")

    # Save to CSV
    csv_path = os.path.join(SCRIPT_DIR, 'historical_group_analysis.csv')
    with open(csv_path, 'w') as f:
        f.write('Group,Injury_Type,Window,Eaten_Mean,Eaten_SEM,Contacted_Mean,Contacted_SEM,N_Animals,Test_Types_Used\n')
        for r in all_results:
            em = f"{r['eaten_mean']:.2f}" if r['eaten_mean'] is not None else ''
            es = f"{r['eaten_sem']:.2f}" if r['eaten_sem'] is not None else ''
            cm = f"{r['contacted_mean']:.2f}" if r['contacted_mean'] is not None else ''
            cs = f"{r['contacted_sem']:.2f}" if r['contacted_sem'] is not None else ''
            f.write(f"{r['group']},{r['injury_type']},{r['window']},{em},{es},{cm},{cs},{r['n_animals']},\"{r['test_types_used']}\"\n")

    print(f"\nResults saved to: {csv_path}")

    # Also print a grouped view by window
    print(f"\n\n{'=' * 110}")
    print("GROUPED BY TIME WINDOW")
    print(f"{'=' * 110}")

    for window_label in ['Final 3', 'Immediate Post', '2-4 Post', 'Last 2 Days']:
        print(f"\n  {window_label}:")
        print(f"  {'Group':6s} {'Injury Type':15s} {'Eaten %':22s} {'Contacted %':26s} {'N':4s}")
        print(f"  {'-' * 6} {'-' * 15} {'-' * 22} {'-' * 26} {'-' * 4}")
        for r in all_results:
            if r['window'] == window_label and r['eaten_mean'] is not None:
                eaten_str = f"{r['eaten_mean']:5.1f} +/- {r['eaten_sem']:4.1f}"
                contacted_str = f"{r['contacted_mean']:5.1f} +/- {r['contacted_sem']:4.1f}"
                print(f"  {r['group']:6s} {r['injury_type']:15s} {eaten_str:22s} {contacted_str:26s} {r['n_animals']:4d}")


if __name__ == '__main__':
    main()
