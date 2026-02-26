"""
L Group Diagnostic Analysis
============================
Investigates anomalous inverted behavioral trend in Group L (Contusion 50kd).

Cross-references manual Excel pellet scores (ground truth) with L_Investigation
findings from ASPA/UCSF swipe-level data.

Outputs (to .../behavior_historical_manual/L_diagnostic/):
  L_individual_timelines.png  - Per-animal session-by-session trajectories
  L_session_completeness.png  - Heatmap of data presence per animal x session
  L_leave_one_out.png         - Group trajectory sensitivity to animal removal
  L_baseline_sensitivity.png  - Effect of varying learner inclusion threshold
  L_post_rehab_pillar.png     - L vs G vs H on pillar days 3+
  L_anomaly_summary.png       - Synthesis table with verdicts
  L_raw_scores.txt            - Raw pellet dumps for anomalous animals

Usage:
  cd Y:\\2_Connectome\\Behavior\\MouseReach\\historical_manual_data
  python diagnose_L_group.py
"""

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
import sys
from collections import defaultdict
from scipy import stats

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from plot_historical_groups import (
    read_group_data, get_time_windows, get_learners,
    get_animal_window_values, build_pellet_dataframe,
    TRAY_OFFSETS_4, TRAY_OFFSETS_2, LEARNER_EATEN_THRESHOLD,
    WINDOW_KEYS, WINDOW_LABELS, GROUP_FILES,
)

# --- Output directory ---
_connectome_root = SCRIPT_DIR
while (os.path.basename(_connectome_root) != '2_Connectome' and
       os.path.dirname(_connectome_root) != _connectome_root):
    _connectome_root = os.path.dirname(_connectome_root)

# Use Databases path (where existing figures actually live)
OUTPUT_DIR = os.path.join(
    _connectome_root, 'Databases', 'figures',
    'behavior_historical_manual', 'L_diagnostic'
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# --- L_Investigation verdicts (from UCSF_Collab/L_Investigation) ---
L_VERDICTS = {
    'L10': ('TRAINING ARTIFACT', 'Never acquired task (0% all training)'),
    'L09': ('LIKELY MILD INJURY', 'Ceiling pre-injury 94%, male 28.1g'),
    'L12': ('LIKELY MILD INJURY', 'Ceiling pre-injury 78.5%, male 29.1g'),
    'L08': ('NOT ANOMALOUS', 'Expected decline, male 27.5g'),
    'L01': ('UNRELIABLE', '30 post-injury swipes'),
    'L02': ('UNRELIABLE', '8 post-injury swipes, double impaction'),
    'L11': ('UNRELIABLE', '41 post-injury swipes, 2 training sessions'),
    'L13': ('OK', '293 post-injury swipes'),
}

VERDICT_COLORS = {
    'TRAINING ARTIFACT': '#d62728',
    'LIKELY MILD INJURY': '#ff7f0e',
    'NOT ANOMALOUS': '#2ca02c',
    'UNRELIABLE': '#7f7f7f',
    'OK': '#1f77b4',
}

VERDICT_BG = {
    'TRAINING ARTIFACT': '#ffcccc',
    'LIKELY MILD INJURY': '#ffe0b2',
    'NOT ANOMALOUS': '#c8e6c9',
    'UNRELIABLE': '#e0e0e0',
    'OK': '#bbdefb',
}

FLAGGED_ANIMALS = ['L01', 'L02', 'L08', 'L09', 'L10', 'L11', 'L12', 'L13']
MOST_ANOMALOUS = ['L09', 'L10', 'L12']
REFERENCE_GROUPS = ['G', 'H']

# Cumulative removal sets for progressive cleaning
REMOVAL_SETS = [
    ('Remove UNRELIABLE\n(L01, L02, L11)', {'L01', 'L02', 'L11'}),
    ('+ Remove TRAINING\nARTIFACT (L10)', {'L01', 'L02', 'L11', 'L10'}),
    ('+ Remove MILD\nINJURY (L09, L12)', {'L01', 'L02', 'L11', 'L10', 'L09', 'L12'}),
]


def _animal_str(animal_id):
    """Normalize animal ID to string for verdict lookup."""
    s = str(animal_id).strip()
    # Handle cases like "10" -> "L10" if needed
    if s.isdigit() and len(s) <= 2:
        s = f'L{s.zfill(2)}'
    return s


def _get_verdict(animal_id):
    """Look up L_Investigation verdict for an animal."""
    key = _animal_str(animal_id)
    return L_VERDICTS.get(key, ('N/A', ''))


def _get_verdict_color(animal_id):
    verdict = _get_verdict(animal_id)[0]
    return VERDICT_COLORS.get(verdict, '#888888')


def _find_injury_index(test_types):
    """Find index of first post-injury test type."""
    for i, tt in enumerate(test_types):
        if 'post' in tt.lower() and 'injury' in tt.lower():
            return i
    return None


def _find_post_rehab_pillar_tts(test_types, test_meta):
    """Find pillar tray days 3+ in rehab block."""
    rehab_tts = [tt for tt in test_types if 'rehab' in tt.lower()]
    pillar_rehab = [
        tt for tt in rehab_tts
        if test_meta.get(tt, {}).get('tray', '').lower().startswith('p')
    ]
    return pillar_rehab[2:] if len(pillar_rehab) > 2 else []


def _learners_at_threshold(data, final3_tts, threshold):
    """Get learners at a specific eaten% threshold."""
    learners = set()
    for animal, tests in data.items():
        if final3_tts:
            e_vals = [tests[tt]['eaten'] for tt in final3_tts if tt in tests]
            if e_vals and np.mean(e_vals) > threshold:
                learners.add(animal)
        else:
            learners.add(animal)
    return learners


# =============================================================================
# 1. Individual Timelines
# =============================================================================

def plot_individual_timelines(data, test_types, test_meta, windows):
    """Full session-by-session trajectory for each L animal."""
    animals = sorted(data.keys(), key=lambda a: _animal_str(a))
    n_animals = len(animals)

    injury_idx = _find_injury_index(test_types)
    post_rehab_tts = set(_find_post_rehab_pillar_tts(test_types, test_meta))

    ncols = 4
    nrows = max(1, (n_animals + ncols - 1) // ncols)
    fig, axes = plt.subplots(nrows, ncols, figsize=(24, 5 * nrows), squeeze=False)
    fig.suptitle(
        'L Group (Contusion 50kd): Per-Animal Session Timeline\n'
        'All test types shown | Red dashed = injury | Green shade = post-rehab (pillar 3+)',
        fontsize=15, fontweight='bold'
    )

    for idx, animal in enumerate(animals):
        row_i, col_i = idx // ncols, idx % ncols
        ax = axes[row_i, col_i]

        animal_data = data[animal]
        xs_e, ys_e = [], []
        xs_c, ys_c = [], []

        for i, tt in enumerate(test_types):
            if tt in animal_data:
                xs_e.append(i)
                ys_e.append(animal_data[tt]['eaten'])
                xs_c.append(i)
                ys_c.append(animal_data[tt]['contacted'])

        ax.plot(xs_e, ys_e, 'o-', color='#d62728', markersize=4,
                linewidth=1.2, label='Retrieved', alpha=0.8)
        ax.plot(xs_c, ys_c, 's-', color='#1f77b4', markersize=4,
                linewidth=1.2, label='Contacted', alpha=0.8)

        if injury_idx is not None:
            ax.axvline(x=injury_idx - 0.5, color='red', linestyle='--',
                       linewidth=2, alpha=0.7)

        if post_rehab_tts:
            pr_indices = [i for i, tt in enumerate(test_types) if tt in post_rehab_tts]
            if pr_indices:
                ax.axvspan(min(pr_indices) - 0.5, max(pr_indices) + 0.5,
                           alpha=0.12, color='green')

        verdict, detail = _get_verdict(animal)
        vcolor = _get_verdict_color(animal)

        ax.set_title(f'{_animal_str(animal)}: {verdict}', fontsize=11,
                     fontweight='bold', color=vcolor)
        ax.text(0.02, 0.98, detail, transform=ax.transAxes, fontsize=7,
                va='top', ha='left', style='italic',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.7))

        ax.set_ylim(-2, 105)
        ax.set_xlim(-0.5, len(test_types) - 0.5)
        ax.set_ylabel('% Pellets')
        if idx == 0:
            ax.legend(fontsize=7, loc='upper right')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        step = max(1, len(test_types) // 8)
        tick_idx = list(range(0, len(test_types), step))
        ax.set_xticks(tick_idx)
        ax.set_xticklabels([test_types[i][:12] for i in tick_idx],
                           rotation=45, fontsize=6, ha='right')

    for idx in range(n_animals, nrows * ncols):
        axes[idx // ncols, idx % ncols].set_visible(False)

    plt.tight_layout()
    out = os.path.join(OUTPUT_DIR, 'L_individual_timelines.png')
    plt.savefig(out, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  Saved: {out}")


# =============================================================================
# 2. Session Completeness Heatmap
# =============================================================================

def plot_session_completeness(data, test_types, test_meta):
    """Heatmap of data presence per animal x test_type."""
    animals = sorted(data.keys(), key=lambda a: _animal_str(a))

    matrix = np.zeros((len(animals), len(test_types)))
    for i, animal in enumerate(animals):
        for j, tt in enumerate(test_types):
            if tt in data[animal]:
                matrix[i, j] = 1

    tray_types = [test_meta.get(tt, {}).get('tray', '?')[:4] for tt in test_types]

    fig_w = max(16, len(test_types) * 0.45)
    fig_h = max(5, len(animals) * 0.45)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))

    im = ax.imshow(matrix, aspect='auto', cmap='RdYlGn', vmin=0, vmax=1,
                   interpolation='nearest')

    ax.set_yticks(range(len(animals)))
    ylabels = []
    for animal in animals:
        n_sess = sum(1 for tt in test_types if tt in data[animal])
        verdict = _get_verdict(animal)[0]
        ylabels.append(f'{_animal_str(animal)} [{n_sess}/{len(test_types)}] ({verdict})')
    ax.set_yticklabels(ylabels, fontsize=8)

    for i, animal in enumerate(animals):
        color = _get_verdict_color(animal)
        ax.get_yticklabels()[i].set_color(color)
        ax.get_yticklabels()[i].set_fontweight('bold')

    ax.set_xticks(range(len(test_types)))
    ax.set_xticklabels(
        [f'{tt[:15]}\n({tr})' for tt, tr in zip(test_types, tray_types)],
        rotation=90, fontsize=5.5
    )

    ax.set_title(
        'L Group: Session Data Completeness\n'
        'Green = data present, Red = missing | Y-axis colored by L_Investigation verdict',
        fontsize=13, fontweight='bold'
    )

    plt.tight_layout()
    out = os.path.join(OUTPUT_DIR, 'L_session_completeness.png')
    plt.savefig(out, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  Saved: {out}")


# =============================================================================
# 3. Leave-One-Out Sensitivity (THE key analysis)
# =============================================================================

def plot_leave_one_out(data, windows, learners, window_data):
    """Remove each learner, recompute group trajectory.

    Top row: raw % trajectory (eaten, contacted)
    Bottom row: normalized-to-baseline trajectory (eaten, contacted)
    Also shows cumulative removal sets.
    """
    from matplotlib.lines import Line2D

    fig, axes = plt.subplots(2, 2, figsize=(20, 14))
    fig.suptitle(
        'L Group: Leave-One-Out Sensitivity Analysis\n'
        'Which animal(s) removal normalizes the group trajectory?',
        fontsize=14, fontweight='bold'
    )

    for col, (metric, ylabel_raw, ylabel_norm) in enumerate([
        ('eaten', '% Pellets Retrieved', '% of Pre-Injury Baseline'),
        ('contacted', '% Pellets Contacted', '% of Pre-Injury Baseline'),
    ]):
        # --- RAW (top row) ---
        ax_raw = axes[0, col]

        full_means = [
            np.mean(window_data[wk][metric]) if len(window_data[wk][metric]) > 0 else 0
            for wk in WINDOW_KEYS
        ]
        x = np.arange(4)
        ax_raw.plot(x, full_means, 'ko-', linewidth=3, markersize=10,
                    label=f'Full group (N={len(learners)})', zorder=10)

        # Individual LOO
        loo_impacts = {}
        for animal in sorted(learners, key=lambda a: _animal_str(a)):
            reduced = learners - {animal}
            reduced_wd = get_animal_window_values(data, windows, reduced)
            loo_means = [
                np.mean(reduced_wd[wk][metric]) if len(reduced_wd[wk][metric]) > 0 else 0
                for wk in WINDOW_KEYS
            ]
            # Impact on rehab window
            impact = full_means[3] - loo_means[3]
            loo_impacts[animal] = impact

            astr = _animal_str(animal)
            is_flagged = astr in FLAGGED_ANIMALS
            vcolor = _get_verdict_color(animal)
            alpha = 0.8 if is_flagged else 0.2
            lw = 1.8 if is_flagged else 0.6

            label = f'w/o {astr} ({_get_verdict(animal)[0]})' if is_flagged else None
            ax_raw.plot(x, loo_means, 'o--', color=vcolor, alpha=alpha,
                        linewidth=lw, markersize=4, label=label)

        # Cumulative removal (dashed thick lines)
        cum_colors = ['#17becf', '#9467bd', '#e377c2']
        for ci, (label, remove_set) in enumerate(REMOVAL_SETS):
            reduced = {a for a in learners if _animal_str(a) not in remove_set}
            if len(reduced) == 0:
                continue
            reduced_wd = get_animal_window_values(data, windows, reduced)
            cum_means = [
                np.mean(reduced_wd[wk][metric]) if len(reduced_wd[wk][metric]) > 0 else 0
                for wk in WINDOW_KEYS
            ]
            ax_raw.plot(x, cum_means, 's-', color=cum_colors[ci], linewidth=2.5,
                        markersize=7, label=f'{label} (N={len(reduced)})', zorder=8)

        ax_raw.set_xticks(x)
        ax_raw.set_xticklabels(WINDOW_LABELS, fontsize=9)
        ax_raw.set_ylabel(ylabel_raw, fontsize=11)
        ax_raw.set_title(f'{metric.capitalize()} (Raw %)', fontsize=12, fontweight='bold')
        ax_raw.spines['top'].set_visible(False)
        ax_raw.spines['right'].set_visible(False)
        ax_raw.legend(fontsize=6, loc='upper right', ncol=1)

        # --- NORMALIZED (bottom row) ---
        ax_norm = axes[1, col]

        final3_tts = windows['final_3']

        def _compute_normalized_trajectory(animal_set):
            """Normalized group trajectory: each animal / own baseline, then average."""
            norm_trajs = {wi: [] for wi in range(4)}
            for animal in animal_set:
                if animal not in data:
                    continue
                # Baseline
                base_vals = [data[animal][tt][metric] for tt in final3_tts
                             if tt in data[animal]]
                if not base_vals:
                    continue
                baseline = np.mean(base_vals)
                if baseline <= 0:
                    continue
                # Each window
                for wi, wk in enumerate(WINDOW_KEYS):
                    wk_vals = [data[animal][tt][metric] for tt in windows[wk]
                               if tt in data[animal]]
                    if wk_vals:
                        norm_trajs[wi].append(np.mean(wk_vals) / baseline * 100)
            return norm_trajs

        # Full group normalized
        full_norm = _compute_normalized_trajectory(learners)
        full_norm_means = [np.mean(full_norm[wi]) if full_norm[wi] else 0 for wi in range(4)]
        ax_norm.plot(x, full_norm_means, 'ko-', linewidth=3, markersize=10,
                     label=f'Full group (N={len(learners)})', zorder=10)
        ax_norm.axhline(y=100, color='gray', linestyle='--', linewidth=1, alpha=0.5)

        # Individual LOO normalized
        for animal in sorted(learners, key=lambda a: _animal_str(a)):
            astr = _animal_str(animal)
            is_flagged = astr in FLAGGED_ANIMALS
            if not is_flagged:
                continue
            reduced = learners - {animal}
            red_norm = _compute_normalized_trajectory(reduced)
            red_norm_means = [np.mean(red_norm[wi]) if red_norm[wi] else 0 for wi in range(4)]

            vcolor = _get_verdict_color(animal)
            ax_norm.plot(x, red_norm_means, 'o--', color=vcolor, alpha=0.8,
                         linewidth=1.8, markersize=4,
                         label=f'w/o {astr} ({_get_verdict(animal)[0]})')

        # Cumulative removal normalized
        for ci, (label, remove_set) in enumerate(REMOVAL_SETS):
            reduced = {a for a in learners if _animal_str(a) not in remove_set}
            if len(reduced) == 0:
                continue
            red_norm = _compute_normalized_trajectory(reduced)
            red_norm_means = [np.mean(red_norm[wi]) if red_norm[wi] else 0 for wi in range(4)]
            ax_norm.plot(x, red_norm_means, 's-', color=cum_colors[ci], linewidth=2.5,
                         markersize=7, label=f'{label} (N={len(reduced)})', zorder=8)

        ax_norm.set_xticks(x)
        ax_norm.set_xticklabels(WINDOW_LABELS, fontsize=9)
        ax_norm.set_ylabel(ylabel_norm, fontsize=11)
        ax_norm.set_title(f'{metric.capitalize()} (Normalized to Baseline)',
                          fontsize=12, fontweight='bold')
        ax_norm.spines['top'].set_visible(False)
        ax_norm.spines['right'].set_visible(False)
        ax_norm.legend(fontsize=6, loc='upper right', ncol=1)

    plt.tight_layout()
    out = os.path.join(OUTPUT_DIR, 'L_leave_one_out.png')
    plt.savefig(out, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  Saved: {out}")

    # Print LOO impact summary
    print("    Leave-one-out rehab impact (eaten):")
    for animal in sorted(loo_impacts, key=lambda a: abs(loo_impacts[a]), reverse=True):
        impact = loo_impacts[animal]
        verdict = _get_verdict(animal)[0]
        print(f"      {_animal_str(animal):5s} ({verdict:20s}): rehab shifts {impact:+.2f}%")


# =============================================================================
# 4. Baseline Threshold Sensitivity
# =============================================================================

def plot_baseline_sensitivity(data, test_types, test_meta):
    """Show how varying the learner threshold changes L group composition."""
    thresholds = [5.0, 8.0, 10.0, 15.0]
    windows = get_time_windows(test_types, test_meta)
    final3_tts = windows['final_3']

    # Also load G and H for N-comparison
    ref_n = {}
    for ref_name in REFERENCE_GROUPS:
        filename, _ = GROUP_FILES[ref_name]
        filepath = os.path.join(SCRIPT_DIR, filename)
        if os.path.exists(filepath):
            rdata, rtt, rmeta, _ = read_group_data(filepath)
            rwindows = get_time_windows(rtt, rmeta)
            ref_n[ref_name] = {}
            for thr in thresholds:
                rlearners = _learners_at_threshold(rdata, rwindows['final_3'], thr)
                ref_n[ref_name][thr] = len(rlearners)

    fig, axes = plt.subplots(2, len(thresholds), figsize=(5.5 * len(thresholds), 10))
    fig.suptitle(
        'L Group: Baseline Threshold Sensitivity\n'
        'How does varying the learner criterion (Final 3 eaten%) affect the group?',
        fontsize=14, fontweight='bold'
    )

    for ti, threshold in enumerate(thresholds):
        learners = _learners_at_threshold(data, final3_tts, threshold)
        window_data = get_animal_window_values(data, windows, learners)
        excluded = sorted(set(data.keys()) - learners, key=lambda a: _animal_str(a))

        for row, (metric, ylabel) in enumerate([
            ('eaten', '% Retrieved'),
            ('contacted', '% Contacted'),
        ]):
            ax = axes[row, ti]

            means = [
                np.mean(window_data[wk][metric]) if len(window_data[wk][metric]) > 0 else 0
                for wk in WINDOW_KEYS
            ]
            sems = [
                stats.sem(window_data[wk][metric]) if len(window_data[wk][metric]) > 1 else 0
                for wk in WINDOW_KEYS
            ]

            x = np.arange(4)
            ax.bar(x, means, yerr=sems, color='#7f7f7f', capsize=4, alpha=0.8,
                   edgecolor='black', linewidth=0.8)

            # Ref group N comparison
            ref_str = '  '.join(
                f'{rn}: N={ref_n[rn].get(threshold, "?")}'
                for rn in REFERENCE_GROUPS if rn in ref_n
            )

            title_lines = f'Threshold: >{threshold}%\nL: N={len(learners)}'
            if ref_str:
                title_lines += f'  |  {ref_str}'
            ax.set_title(title_lines, fontsize=9, fontweight='bold')

            if excluded:
                excl_labels = [
                    f'{_animal_str(a)}({_get_verdict(a)[0][:4]})'
                    for a in excluded[:6]
                ]
                excl_str = ', '.join(excl_labels)
                if len(excluded) > 6:
                    excl_str += '...'
                ax.text(0.02, 0.98, f'Excluded: {excl_str}',
                        transform=ax.transAxes, fontsize=5.5, va='top', ha='left',
                        bbox=dict(facecolor='lightyellow', alpha=0.7))

            ax.set_xticks(x)
            ax.set_xticklabels(['Last 3', 'PI 1', 'PI 2-4', 'Rehab'], fontsize=8)
            ax.set_ylabel(ylabel, fontsize=9)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

    plt.tight_layout()
    out = os.path.join(OUTPUT_DIR, 'L_baseline_sensitivity.png')
    plt.savefig(out, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  Saved: {out}")


# =============================================================================
# 5. Post-Rehab Pillar (Days 3+) Comparison
# =============================================================================

def plot_post_rehab_pillar(l_data, l_test_types, l_test_meta, l_learners):
    """Compare L vs G vs H on pillar tray days 3+ in rehab."""
    groups_info = {
        'L': (l_data, l_test_types, l_test_meta, l_learners, '#2ca02c'),
    }

    for ref_name in REFERENCE_GROUPS:
        filename, _ = GROUP_FILES[ref_name]
        filepath = os.path.join(SCRIPT_DIR, filename)
        if not os.path.exists(filepath):
            continue
        rdata, rtt, rmeta, _ = read_group_data(filepath)
        rwindows = get_time_windows(rtt, rmeta)
        rlearners = get_learners(rdata, rwindows['final_3'])
        color = '#d62728' if ref_name == 'G' else '#1f77b4'
        groups_info[ref_name] = (rdata, rtt, rmeta, rlearners, color)

    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    fig.suptitle(
        'Post-Rehab Pillar Analysis (Pillar Days 3+)\n'
        'L (Contusion 50kd) vs G, H (Transection) -- different time periods',
        fontsize=14, fontweight='bold'
    )

    for col, (metric, ylabel) in enumerate([
        ('eaten', '% Pellets Retrieved'),
        ('contacted', '% Pellets Contacted'),
    ]):
        ax = axes[col]

        group_labels = []
        group_means = []
        group_sems = []
        group_colors = []
        group_individual = []

        for gname, (gdata, gtt, gmeta, glearners, gcolor) in groups_info.items():
            pr_tts = _find_post_rehab_pillar_tts(gtt, gmeta)
            if not pr_tts:
                print(f"    {gname}: no pillar 3+ sessions found")
                continue

            vals = []
            for animal in sorted(glearners):
                if animal not in gdata:
                    continue
                a_vals = [gdata[animal][tt][metric] for tt in pr_tts
                          if tt in gdata[animal]]
                if a_vals:
                    vals.append(np.mean(a_vals))

            if not vals:
                continue

            group_labels.append(f'{gname}\n(N={len(vals)}, {len(pr_tts)} sess)')
            group_means.append(np.mean(vals))
            group_sems.append(stats.sem(vals) if len(vals) > 1 else 0)
            group_colors.append(gcolor)
            group_individual.append(vals)

        if group_labels:
            x = np.arange(len(group_labels))
            ax.bar(x, group_means, yerr=group_sems, color=group_colors,
                   capsize=5, alpha=0.8, edgecolor='black', linewidth=1)

            np.random.seed(42)
            for i, vals in enumerate(group_individual):
                jitter = np.random.uniform(-0.15, 0.15, len(vals))
                ax.scatter(x[i] + jitter, vals, color='black', s=25,
                           zorder=5, alpha=0.5, edgecolor='white', linewidth=0.3)

            ax.set_xticks(x)
            ax.set_xticklabels(group_labels, fontsize=10)

        ax.set_ylabel(ylabel, fontsize=12)
        ax.set_title(f'{metric.capitalize()} -- Pillar Days 3+',
                     fontsize=12, fontweight='bold')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

    plt.tight_layout()
    out = os.path.join(OUTPUT_DIR, 'L_post_rehab_pillar.png')
    plt.savefig(out, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  Saved: {out}")


# =============================================================================
# 6. Anomaly Summary Table
# =============================================================================

def plot_anomaly_summary(data, windows, learners, window_data):
    """Table figure synthesizing L_Investigation verdicts with manual Excel evidence."""
    final3_tts = windows['final_3']
    ip_tts = windows['immediate_post']
    l2_tts = windows['last_2']

    fig, ax = plt.subplots(figsize=(18, max(6, len(data) * 0.45 + 2)))
    ax.axis('off')

    headers = [
        'Animal', 'Learner?', 'Final 3\nEaten%', 'Post Inj 1\nEaten%',
        'Rehab L2\nEaten%', 'Direction\n(Manual)', 'L_Inv\nVerdict', 'Detail'
    ]

    rows = []
    row_colors = []
    for animal in sorted(data.keys(), key=lambda a: _animal_str(a)):
        astr = _animal_str(animal)
        is_learner = animal in learners

        f3_vals = [data[animal][tt]['eaten'] for tt in final3_tts if tt in data[animal]]
        f3_mean = np.mean(f3_vals) if f3_vals else 0.0

        ip_vals = [data[animal][tt]['eaten'] for tt in ip_tts if tt in data[animal]]
        ip_mean = np.mean(ip_vals) if ip_vals else None

        l2_vals = [data[animal][tt]['eaten'] for tt in l2_tts if tt in data[animal]]
        l2_mean = np.mean(l2_vals) if l2_vals else None

        if l2_mean is not None and f3_mean > 0:
            if l2_mean > f3_mean * 0.8:
                direction = 'RECOVERED'
            elif ip_mean is not None and l2_mean > ip_mean:
                direction = 'IMPROVED'
            elif ip_mean is not None:
                direction = 'DECLINED'
            else:
                direction = 'INCOMPLETE'
        else:
            direction = 'INCOMPLETE'

        verdict, detail = _get_verdict(animal)

        rows.append([
            astr,
            'Yes' if is_learner else 'NO',
            f'{f3_mean:.1f}%',
            f'{ip_mean:.1f}%' if ip_mean is not None else '-',
            f'{l2_mean:.1f}%' if l2_mean is not None else '-',
            direction,
            verdict,
            detail[:45],
        ])
        row_colors.append(VERDICT_BG.get(verdict, 'white'))

    table = ax.table(cellText=rows, colLabels=headers, loc='center',
                     cellLoc='center', colLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1, 1.6)

    for i, bgcolor in enumerate(row_colors):
        for j in range(len(headers)):
            table[i + 1, j].set_facecolor(bgcolor)

    # Header styling
    for j in range(len(headers)):
        table[0, j].set_facecolor('#333333')
        table[0, j].set_text_props(color='white', fontweight='bold')

    ax.set_title(
        'L Group Anomaly Summary\n'
        'Manual Excel pellet data + L_Investigation (ASPA/UCSF) verdicts',
        fontsize=14, fontweight='bold', pad=20
    )

    plt.tight_layout()
    out = os.path.join(OUTPUT_DIR, 'L_anomaly_summary.png')
    plt.savefig(out, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  Saved: {out}")


# =============================================================================
# 7. Raw Score Dump
# =============================================================================

def dump_raw_scores(filepath, animals_of_interest):
    """Dump raw 0/1/2 pellet scores for anomalous animals to text file."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['1 - ENTER DATA HERE']

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    tray_offsets = TRAY_OFFSETS_4 if len(headers) > 80 else TRAY_OFFSETS_2

    target_strs = {_animal_str(a) for a in animals_of_interest}

    out_path = os.path.join(OUTPUT_DIR, 'L_raw_scores.txt')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write("L Group Raw Pellet Scores for Anomalous Animals\n")
        f.write("=" * 80 + "\n")
        f.write("Score key: 0=Miss, 1=Displaced, 2=Retrieved\n")
        f.write(f"Animals: {', '.join(sorted(target_strs))}\n")
        f.write("=" * 80 + "\n")

        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            if vals[0] is None:
                continue

            animal = vals[5]
            if _animal_str(animal) not in target_strs:
                continue

            test_type = str(vals[1])
            date_val = vals[0]
            tray = str(vals[2]) if vals[2] else '?'

            f.write(f"\n{_animal_str(animal)} | {date_val} | {test_type} | Tray: {tray}\n")
            f.write("-" * 60 + "\n")

            for tray_idx, (start, end) in enumerate(tray_offsets):
                if start >= len(vals):
                    continue
                pellets = vals[start:end]
                scores = []
                for v in pellets:
                    if v is not None and isinstance(v, (int, float)):
                        scores.append(str(int(v)))
                    else:
                        scores.append('.')
                f.write(f"  Tray {tray_idx + 1}: {' '.join(scores)}\n")

                n_scored = sum(1 for s in scores if s != '.')
                n_miss = sum(1 for s in scores if s == '0')
                n_disp = sum(1 for s in scores if s == '1')
                n_ret = sum(1 for s in scores if s == '2')
                if n_scored > 0:
                    f.write(
                        f"           [{n_scored} scored: {n_miss}M {n_disp}D {n_ret}R"
                        f" = {n_ret/n_scored*100:.0f}% eaten,"
                        f" {(n_disp+n_ret)/n_scored*100:.0f}% contacted]\n"
                    )

    wb.close()
    print(f"  Saved: {out_path}")


# =============================================================================
# Main
# =============================================================================

def main():
    print("=" * 70)
    print("L GROUP DIAGNOSTIC ANALYSIS")
    print("  Manual Excel pellet scores + L_Investigation cross-reference")
    print("  Post-rehab = pillar tray days 3+")
    print(f"  Learner criterion: Final 3 avg eaten% > {LEARNER_EATEN_THRESHOLD}%")
    print("=" * 70)

    l_filename, l_injury = GROUP_FILES['L']
    l_filepath = os.path.join(SCRIPT_DIR, l_filename)

    if not os.path.exists(l_filepath):
        print(f"ERROR: {l_filepath} not found")
        return

    print(f"\nReading L group data from: {l_filename}")
    data, test_types, test_meta, pellet_records = read_group_data(l_filepath)
    windows = get_time_windows(test_types, test_meta)
    learners = get_learners(data, windows['final_3'])
    window_data = get_animal_window_values(data, windows, learners)

    all_animals = sorted(data.keys(), key=lambda a: _animal_str(a))
    print(f"  Total animals: {len(data)}")
    print(f"  Learners (>{LEARNER_EATEN_THRESHOLD}% eaten): {len(learners)}")
    print(f"  Non-learners: {len(data) - len(learners)}")

    learner_strs = {_animal_str(a) for a in learners}
    flagged_in = sorted(f for f in FLAGGED_ANIMALS if f in learner_strs)
    flagged_out = sorted(f for f in FLAGGED_ANIMALS if f not in learner_strs)
    print(f"  Flagged in learners: {flagged_in}")
    print(f"  Flagged NOT learners: {flagged_out}")

    print(f"  Test types ({len(test_types)}): {test_types[:5]}...{test_types[-3:]}")
    print(f"\n  Windows:")
    for wk in WINDOW_KEYS:
        tts = windows[wk]
        n_a = len(window_data[wk]['animals'])
        print(f"    {wk:20s}: {n_a:2d} animals, {len(tts)} sessions "
              f"| {', '.join(tts[:3])}" + ("..." if len(tts) > 3 else ""))

    # Find post-rehab pillar (P3+)
    pr_tts = _find_post_rehab_pillar_tts(test_types, test_meta)
    print(f"\n  Post-rehab pillar (P3+): {len(pr_tts)} sessions")
    if pr_tts:
        print(f"    {', '.join(pr_tts)}")

    print(f"\nOutput directory: {OUTPUT_DIR}")

    # --- Run analyses ---
    print("\n[1/7] Individual timelines...")
    plot_individual_timelines(data, test_types, test_meta, windows)

    print("\n[2/7] Session completeness...")
    plot_session_completeness(data, test_types, test_meta)

    print("\n[3/7] Leave-one-out sensitivity...")
    plot_leave_one_out(data, windows, learners, window_data)

    print("\n[4/7] Baseline threshold sensitivity...")
    plot_baseline_sensitivity(data, test_types, test_meta)

    print("\n[5/7] Post-rehab pillar (days 3+)...")
    plot_post_rehab_pillar(data, test_types, test_meta, learners)

    print("\n[6/7] Anomaly summary table...")
    plot_anomaly_summary(data, windows, learners, window_data)

    print("\n[7/7] Raw score dump (L09, L10, L12)...")
    dump_raw_scores(l_filepath, MOST_ANOMALOUS)

    print(f"\n{'=' * 70}")
    print(f"All outputs saved to: {OUTPUT_DIR}")
    print(f"{'=' * 70}")


if __name__ == '__main__':
    main()
