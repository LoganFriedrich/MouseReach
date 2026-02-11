"""
Generate behavior bar plots matching fig5_behavior_layout4.png style
for all historical (non-CNT, non-ENCR) injury groups.

Creates per-group figures:
  - 2x2 bar panels (asterisks / p-values x Retrieved / Contacted)
  - Per-animal recovery panel (paired lines: baseline vs end-of-rehab)

Data is read directly from each group's "1 - ENTER DATA HERE" sheet,
computing Eaten% and Contacted% from raw pellet scores (0/1/2).

Stats: GEE (Generalized Estimating Equations) with binomial family,
exchangeable correlation structure, and animal as clustering variable.
Post-hoc: pairwise contrasts with Holm correction.
Fallback: chi-square / Fisher's exact when GEE can't converge (complete separation).
Last 2: Uses last 2 Pillar tray days in continuous rehab (before any >5-day gap).
"""

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import openpyxl
import os
import re
import warnings
from collections import defaultdict
from scipy import stats
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Binomial
from statsmodels.genmod.cov_struct import Exchangeable

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Output to MouseDB/figures for finalized data products
# Derive project root by walking up from script location
_connectome_root = SCRIPT_DIR
while os.path.basename(_connectome_root) != '2_Connectome' and os.path.dirname(_connectome_root) != _connectome_root:
    _connectome_root = os.path.dirname(_connectome_root)
OUTPUT_DIR = os.path.join(_connectome_root, 'MouseDB', 'figures', 'behavior_historical_manual')
os.makedirs(OUTPUT_DIR, exist_ok=True)

COLORS = {
    'bar': '#7f7f7f',
    'point': '#1f1f1f',
    'sig_bracket': '#333333',
}

GROUP_FILES = {
    'ABS3': ('ABS3.xlsx', 'Early Study'),
    'G': ('G - Transection.xlsx', 'Transection'),
    'H': ('H - Transection.xlsx', 'Transection'),
    'K': ('K - Contusion 70kd.xlsx', 'Contusion 70kd'),
    'L': ('L - Contusion - 50kd.xlsx', 'Contusion 50kd'),
    'M': ('M - Contusion - 60kd.xlsx', 'Contusion 60kd'),
    'OptD': ('OptD - Rehab 1 - pyramidotomy.xlsx', 'Pyramidotomy'),
}

TRAY_OFFSETS_4 = [(7, 27), (31, 51), (55, 75), (79, 99)]
TRAY_OFFSETS_2 = [(7, 27), (31, 51)]

WINDOW_LABELS = ['Last 3', 'Post\nInjury 1', 'Post\nInjury 2-4', 'Rehab\nPillar']
WINDOW_KEYS = ['final_3', 'immediate_post', '2_4_post', 'last_2']

# Learner criterion: exclude animals with Final 3 avg eaten% <= this threshold
LEARNER_EATEN_THRESHOLD = 5.0

# Bracket drawing order (shortest span first, for clean stacking)
BRACKET_ORDER = [
    (0, 1),  # Final 3 vs Post Injury 1 (span=1)
    (2, 3),  # Post Injury 2-4 vs Rehab Pillar (span=1)
    (0, 2),  # Final 3 vs Post Injury 2-4 (span=2)
    (1, 3),  # Post Injury 1 vs Rehab Pillar (span=2)
]


def get_significance_stars(p_value):
    if p_value < 0.0001:
        return '****'
    elif p_value < 0.001:
        return '***'
    elif p_value < 0.01:
        return '**'
    elif p_value < 0.05:
        return '*'
    else:
        return 'ns'


def add_stat_bracket(ax, x1, x2, y, h, p_val, show_pval=False):
    ax.plot([x1, x1, x2, x2], [y, y+h, y+h, y], lw=1.5, c=COLORS['sig_bracket'])
    if show_pval:
        text = '<0.0001' if p_val < 0.0001 else f'{p_val:.4f}'
    else:
        text = get_significance_stars(p_val)
    ax.text((x1+x2)/2, y+h, text, ha='center', va='bottom', fontsize=10, fontweight='bold')


def is_valid_tray(pellet_values):
    for v in pellet_values:
        if v is not None and v != 'N/A' and isinstance(v, (int, float)):
            return True
    return False


def read_group_data(filepath):
    """Read raw pellet data from Excel.

    Returns:
        data: {animal: {test_type: {eaten: float%, contacted: float%}}} - animal-level averages
        sorted test types list
        test_meta: {test_type: {date, tray}}
        pellet_records: list of (animal, test_type, eaten_01, contacted_01) - one per pellet
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['1 - ENTER DATA HERE']

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    tray_offsets = TRAY_OFFSETS_4 if len(headers) > 80 else TRAY_OFFSETS_2

    data = defaultdict(dict)
    pellet_records = []
    all_test_types = set()
    test_meta = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if vals[0] is None:
            continue

        test_type = str(vals[1])
        animal = vals[5]
        all_test_types.add(test_type)

        if test_type not in test_meta:
            test_meta[test_type] = {'date': vals[0], 'tray': str(vals[2])}

        tray_eaten, tray_contacted = [], []
        for start, end in tray_offsets:
            if start >= len(vals):
                continue
            pellets = vals[start:end]
            if not is_valid_tray(pellets):
                continue

            valid_in_tray = []
            for v in pellets:
                if v is not None and v != 'N/A' and isinstance(v, (int, float)):
                    valid_in_tray.append(v)
                    pellet_records.append((
                        animal, test_type,
                        1 if v >= 2 else 0,   # eaten binary
                        1 if v >= 1 else 0,   # contacted binary
                    ))

            if valid_in_tray:
                n_p = len(valid_in_tray)
                tray_eaten.append(sum(1 for v in valid_in_tray if v >= 2) / n_p * 100)
                tray_contacted.append(sum(1 for v in valid_in_tray if v >= 1) / n_p * 100)

        if tray_eaten:
            data[animal][test_type] = {
                'eaten': np.mean(tray_eaten),
                'contacted': np.mean(tray_contacted),
            }

    wb.close()
    return data, sorted(all_test_types), test_meta, pellet_records


def get_time_windows(test_types_sorted, test_meta):
    """Determine time windows, using dates and tray info for correct Last 2."""
    windows = {k: [] for k in WINDOW_KEYS}

    first_post_idx = None
    for i, tt in enumerate(test_types_sorted):
        if 'post' in tt.lower() and 'injury' in tt.lower():
            first_post_idx = i
            break

    if first_post_idx is None:
        return windows

    pre = test_types_sorted[:first_post_idx]
    windows['final_3'] = pre[-3:] if len(pre) >= 3 else pre

    for tt in test_types_sorted[first_post_idx:]:
        tt_lower = tt.lower()
        if '1 week post' in tt_lower:
            windows['immediate_post'].append(tt)
        else:
            m = re.search(r'(\d+)\s*week\s*post', tt_lower)
            if m and 2 <= int(m.group(1)) <= 4:
                windows['2_4_post'].append(tt)

    # Last 2: find last 2 Pillar days in continuous rehab (no gap > 5 days)
    rehab_tts = [tt for tt in test_types_sorted if 'rehab' in tt.lower()]

    if not rehab_tts:
        windows['last_2'] = test_types_sorted[-2:]
        return windows

    # Split rehab into continuous blocks (gap > 5 days = new block)
    blocks = [[rehab_tts[0]]]
    for i in range(1, len(rehab_tts)):
        prev_d = test_meta.get(rehab_tts[i-1], {}).get('date')
        curr_d = test_meta.get(rehab_tts[i], {}).get('date')
        if prev_d and curr_d:
            try:
                gap = (curr_d - prev_d).days
            except (TypeError, AttributeError):
                gap = 1
            if 5 < gap < 100:  # Real breaks (5-100 days); ignore date-entry errors
                blocks.append([rehab_tts[i]])
                continue
        blocks[-1].append(rehab_tts[i])

    # Use the first (main) rehab block - find last 2 Pillar days in it
    main_block = blocks[0]
    pillar_in_main = [tt for tt in main_block
                      if test_meta.get(tt, {}).get('tray', '').lower().startswith('p')]

    if len(pillar_in_main) >= 2:
        windows['last_2'] = pillar_in_main[-2:]
    elif len(pillar_in_main) == 1:
        windows['last_2'] = pillar_in_main
    else:
        windows['last_2'] = main_block[-2:] if len(main_block) >= 2 else main_block

    return windows


def get_learners(data, final3_tts):
    """Return set of animals meeting learner criterion (Final 3 avg eaten% > threshold)."""
    learners = set()
    for animal, tests in data.items():
        if final3_tts:
            e_vals = [tests[tt]['eaten'] for tt in final3_tts if tt in tests]
            if e_vals and np.mean(e_vals) > LEARNER_EATEN_THRESHOLD:
                learners.add(animal)
        else:
            learners.add(animal)
    return learners


def get_animal_window_values(data, windows, learners):
    """Compute per-animal averages for each time window (no pairing requirement).

    Each window independently includes all learners with data in that window.
    Returns dict: {window_key: {eaten: array, contacted: array, animals: list}}
    """
    result = {}
    for wkey in WINDOW_KEYS:
        tt_list = windows[wkey]
        if not tt_list:
            result[wkey] = {'eaten': np.array([]), 'contacted': np.array([]), 'animals': []}
            continue

        eaten_vals, contacted_vals, animal_ids = [], [], []
        for animal in sorted(learners):
            if animal not in data:
                continue
            e_vals, c_vals = [], []
            for tt in tt_list:
                if tt in data[animal]:
                    e_vals.append(data[animal][tt]['eaten'])
                    c_vals.append(data[animal][tt]['contacted'])
            if e_vals:
                eaten_vals.append(np.mean(e_vals))
                contacted_vals.append(np.mean(c_vals))
                animal_ids.append(animal)

        result[wkey] = {
            'eaten': np.array(eaten_vals),
            'contacted': np.array(contacted_vals),
            'animals': animal_ids,
        }

    return result


def build_pellet_dataframe(pellet_records, windows, learners):
    """Build pandas DataFrame from pellet records for GEE analysis.

    Each row is one pellet observation with columns:
        animal, window (int 0-3), eaten (0/1), contacted (0/1), animal_code (int)
    Filtered to learner animals and test types within defined windows.
    """
    tt_to_window = {}
    for wk_idx, wk in enumerate(WINDOW_KEYS):
        for tt in windows[wk]:
            tt_to_window[tt] = wk_idx

    rows = []
    for animal, test_type, eaten, contacted in pellet_records:
        if animal not in learners:
            continue
        if test_type not in tt_to_window:
            continue
        rows.append({
            'animal': animal,
            'window': tt_to_window[test_type],
            'eaten': eaten,
            'contacted': contacted,
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    animal_codes = {a: i for i, a in enumerate(sorted(df['animal'].unique()))}
    df['animal_code'] = df['animal'].map(animal_codes)
    return df


def _holm_correct(pvals):
    """Apply Holm-Bonferroni correction to a dict of p-values."""
    if not pvals:
        return pvals
    sorted_pairs = sorted(pvals.keys(), key=lambda k: pvals[k])
    m = len(sorted_pairs)
    corrected = {}
    for rank, pair in enumerate(sorted_pairs):
        corrected[pair] = min(1.0, pvals[pair] * (m - rank))
    # Enforce monotonicity
    prev_p = 0
    for pair in sorted_pairs:
        corrected[pair] = max(corrected[pair], prev_p)
        prev_p = corrected[pair]
    return corrected


def _chi_square_fallback(df, metric, comparisons, n_per_window):
    """Fallback to chi-square/Fisher's exact when GEE can't converge."""
    raw_pvals = {}
    for (i, j) in comparisons:
        di = df[df['window'] == i]
        dj = df[df['window'] == j]
        if len(di) == 0 or len(dj) == 0:
            continue
        a = int(di[metric].sum())
        b = len(di) - a
        c = int(dj[metric].sum())
        d_val = len(dj) - c
        table = np.array([[a, b], [c, d_val]])
        if table.min() < 5:
            _, p = stats.fisher_exact(table)
        else:
            _, p, _, _ = stats.chi2_contingency(table, correction=True)
        raw_pvals[(i, j)] = p

    raw_pvals = _holm_correct(raw_pvals)

    # Omnibus: chi-square on full contingency table across available windows
    available = [w for w in range(4) if n_per_window.get(w, 0) > 0]
    if len(available) >= 2:
        table_rows = []
        for w in available:
            subset = df[df['window'] == w]
            table_rows.append([int(subset[metric].sum()), len(subset) - int(subset[metric].sum())])
        table = np.array(table_rows)
        if table.sum() > 0:
            _, omnibus_p, _, _ = stats.chi2_contingency(table)
        else:
            omnibus_p = 1.0
    else:
        omnibus_p = 1.0

    return raw_pvals, omnibus_p, n_per_window, 'chi-sq/Fisher'


def run_gee_posthoc(df, metric):
    """Run GEE (binomial, exchangeable, animal cluster) + pairwise contrasts with Holm.

    Returns (pval_dict, omnibus_p, n_pellets_per_window, method_label)
    """
    comparisons = [(0, 1), (0, 2), (1, 3), (2, 3)]

    n_per_window = {}
    if not df.empty:
        for w in range(4):
            n_per_window[w] = int((df['window'] == w).sum())
    else:
        n_per_window = {w: 0 for w in range(4)}

    if df.empty or df['window'].nunique() < 2:
        return {}, 1.0, n_per_window, 'N/A'

    available_windows = sorted([w for w in range(4) if n_per_window.get(w, 0) > 0])
    if len(available_windows) < 2:
        return {}, 1.0, n_per_window, 'N/A'

    df_gee = df[df['window'].isin(available_windows)].copy()

    if df_gee['animal_code'].nunique() < 2:
        return _chi_square_fallback(df, metric, comparisons, n_per_window)

    # Check for complete separation in any window
    for w in available_windows:
        subset = df_gee[df_gee['window'] == w][metric]
        if len(subset) > 0 and (subset.sum() == 0 or subset.sum() == len(subset)):
            return _chi_square_fallback(df, metric, comparisons, n_per_window)

    try:
        ref_window = available_windows[0]  # 0 (final_3) if present
        df_gee = df_gee.sort_values('animal_code').reset_index(drop=True)

        formula = f'{metric} ~ C(window, Treatment({ref_window}))'
        model = GEE.from_formula(
            formula,
            groups='animal_code',
            data=df_gee,
            family=Binomial(),
            cov_struct=Exchangeable(),
        )

        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            result = model.fit(maxiter=100)

        # Omnibus Wald test for all window effects
        param_names = list(result.params.index)
        window_params = [p for p in param_names if 'C(window' in p]

        if len(window_params) >= 1:
            idx = [param_names.index(p) for p in window_params]
            beta_w = result.params.values[idx]
            vcov_w = result.cov_params().values[np.ix_(idx, idx)]
            try:
                wald_stat = float(beta_w @ np.linalg.inv(vcov_w) @ beta_w)
                omnibus_p = float(1 - stats.chi2.cdf(wald_stat, len(window_params)))
            except np.linalg.LinAlgError:
                omnibus_p = 1.0
        else:
            omnibus_p = 1.0

        # Pairwise contrasts
        params = result.params.values
        vcov = result.cov_params().values

        raw_pvals = {}
        for (i, j) in comparisons:
            if i not in available_windows or j not in available_windows:
                continue

            L = np.zeros(len(params))

            if i == ref_window:
                j_names = [p for p in param_names if f'T.{j}]' in p]
                if not j_names:
                    continue
                L[param_names.index(j_names[0])] = 1.0
            elif j == ref_window:
                i_names = [p for p in param_names if f'T.{i}]' in p]
                if not i_names:
                    continue
                L[param_names.index(i_names[0])] = -1.0
            else:
                i_names = [p for p in param_names if f'T.{i}]' in p]
                j_names = [p for p in param_names if f'T.{j}]' in p]
                if not i_names or not j_names:
                    continue
                L[param_names.index(j_names[0])] = 1.0
                L[param_names.index(i_names[0])] = -1.0

            est = float(L @ params)
            var_est = float(L @ vcov @ L)
            if var_est <= 0:
                continue
            z = est / np.sqrt(var_est)
            p = float(2 * (1 - stats.norm.cdf(abs(z))))
            raw_pvals[(i, j)] = p

        raw_pvals = _holm_correct(raw_pvals)

        return raw_pvals, omnibus_p, n_per_window, 'GEE'

    except Exception as e:
        print(f"    GEE failed ({e}), falling back to chi-square")
        return _chi_square_fallback(df, metric, comparisons, n_per_window)


def plot_group(group_name, injury_type, window_data, pellet_df, n_learners, output_dir):
    """Create 2x2 bar panel figure with GEE post-hoc tests.

    Bars/scatter = animal-level means and SEM (visual).
    Stats = GEE on pellet-level binary outcomes with animal clustering.
    """
    fig, axes = plt.subplots(2, 2, figsize=(14, 11))

    fig.suptitle(
        f'Behavior Performance (Manually Scored)\n'
        f'({group_name} - {injury_type}, N={n_learners} learners, '
        f'eaten >{LEARNER_EATEN_THRESHOLD}% at training)',
        fontsize=16, fontweight='bold'
    )

    for col, (metric_key, metric_name, ylabel) in enumerate([
        ('eaten', 'Retrieved', '% Pellets Retrieved'),
        ('contacted', 'Contacted', '% Pellets Contacted'),
    ]):
        data_by_phase = [window_data[wk][metric_key] for wk in WINDOW_KEYS]

        # GEE on pellet-level data
        posthoc_pvals, omnibus_p, n_pellets, method = run_gee_posthoc(pellet_df, metric_key)

        for row, show_pval in enumerate([False, True]):
            ax = axes[row, col]

            means = [np.mean(d) if len(d) > 0 else 0 for d in data_by_phase]
            sems = [stats.sem(d) if len(d) > 1 else 0 for d in data_by_phase]

            x = np.arange(len(WINDOW_KEYS))

            bars = ax.bar(x, means, yerr=sems, color=COLORS['bar'],
                         capsize=5, alpha=0.8, edgecolor='black', linewidth=1)

            np.random.seed(42)
            for i, d in enumerate(data_by_phase):
                if len(d) > 0:
                    jitter = np.random.uniform(-0.15, 0.15, len(d))
                    ax.scatter(x[i] + jitter, d, color=COLORS['point'],
                              s=40, zorder=5, alpha=0.6, edgecolor='white', linewidth=0.5)

            # Statistical brackets
            all_vals = [v for d in data_by_phase for v in d]
            data_ceil = max(all_vals) if all_vals else 1
            data_ceil = max(data_ceil, 5)

            bracket_h = 1.5
            bracket_step = 4.0
            first_bracket_y = data_ceil + 3

            bracket_idx = 0
            for (i, j) in BRACKET_ORDER:
                if (i, j) in posthoc_pvals:
                    y_pos = first_bracket_y + bracket_idx * bracket_step
                    add_stat_bracket(ax, x[i], x[j], y_pos, bracket_h,
                                    posthoc_pvals[(i, j)], show_pval=show_pval)
                    bracket_idx += 1

            # N labels on x-axis (pellet counts + animal counts)
            xlabels = []
            for i, wl in enumerate(WINDOW_LABELS):
                n_p = n_pellets.get(i, 0)
                n_a = len(data_by_phase[i])
                xlabels.append(f'{wl}\n({n_a} mice, {n_p} pel.)')

            if bracket_idx > 0:
                top_bracket = first_bracket_y + bracket_idx * bracket_step
                ax.set_ylim(0, top_bracket + bracket_step)
            else:
                ax.set_ylim(0, data_ceil * 1.2)

            ax.set_xticks(x)
            ax.set_xticklabels(xlabels, fontsize=8)
            ax.set_ylabel(ylabel, fontsize=12)

            omnibus_str = f'p={omnibus_p:.4f}' if omnibus_p >= 0.0001 else 'p<0.0001'
            if show_pval:
                title_suffix = f"({method} p-values + Holm | omnibus {omnibus_str})"
            else:
                title_suffix = f"(asterisks, {method} + Holm | omnibus {omnibus_str})"
            ax.set_title(f'{metric_name} {title_suffix}', fontsize=10, fontweight='bold')
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

    plt.tight_layout()
    out_path = os.path.join(output_dir, f'behavior_{group_name}.png')
    plt.savefig(out_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  Saved: {out_path}")


def plot_recovery(group_name, injury_type, window_data, output_dir):
    """Per-animal 4-point recovery plot: all windows for each subject.

    Shows individual animal trajectories across all 4 timepoints.
    Animals need pre + rehab at minimum; 1Wk and 2-4Wk shown when available.
    Recovery classification uses per-animal nadir (worst post-injury point).
    """
    from matplotlib.lines import Line2D

    f3 = window_data['final_3']
    ip = window_data['immediate_post']
    p24 = window_data['2_4_post']
    l2 = window_data['last_2']

    f3_idx = {a: i for i, a in enumerate(f3['animals'])}
    ip_idx = {a: i for i, a in enumerate(ip['animals'])}
    p24_idx = {a: i for i, a in enumerate(p24['animals'])}
    l2_idx = {a: i for i, a in enumerate(l2['animals'])}

    # Need at least pre + rehab
    paired_animals = sorted(set(f3['animals']) & set(l2['animals']))
    has_ip = set(ip['animals'])
    has_24 = set(p24['animals'])

    if len(paired_animals) < 2:
        return

    n = len(paired_animals)

    fig, axes = plt.subplots(1, 2, figsize=(14, 7))
    fig.suptitle(
        f'Per-Animal Recovery (Manually Scored): {group_name} - {injury_type} (N={n})\n'
        f'Full Trajectory: Pre-Injury through Rehab',
        fontsize=13, fontweight='bold'
    )

    for col, (metric_key, metric_name, ylabel) in enumerate([
        ('eaten', 'Retrieved', '% Pellets Retrieved'),
        ('contacted', 'Contacted', '% Pellets Contacted'),
    ]):
        ax = axes[col]
        mean_vals = {0: [], 1: [], 2: [], 3: []}

        for animal in paired_animals:
            pre = f3[metric_key][f3_idx[animal]]
            rehab = l2[metric_key][l2_idx[animal]]
            post_1wk = ip[metric_key][ip_idx[animal]] if animal in has_ip else None
            post_24wk = p24[metric_key][p24_idx[animal]] if animal in has_24 else None

            # Per-animal nadir
            post_vals = [v for v in [post_1wk, post_24wk] if v is not None]
            nadir = min(post_vals) if post_vals else None

            # Classify by recovery relative to nadir
            if nadir is not None:
                if rehab > pre * 0.8:
                    color = '#2ca02c'; alpha = 0.7
                elif rehab > nadir:
                    color = '#ff7f0e'; alpha = 0.6
                else:
                    color = '#d62728'; alpha = 0.5
            else:
                color = '#888888'; alpha = 0.5

            # Build trajectory points
            xs, ys = [0], [pre]
            mean_vals[0].append(pre)
            if post_1wk is not None:
                xs.append(1); ys.append(post_1wk)
                mean_vals[1].append(post_1wk)
            if post_24wk is not None:
                xs.append(2); ys.append(post_24wk)
                mean_vals[2].append(post_24wk)
            xs.append(3); ys.append(rehab)
            mean_vals[3].append(rehab)

            # If missing middle points, use dashed line for gap
            if post_1wk is not None and post_24wk is not None:
                ax.plot(xs, ys, 'o-', color=color, alpha=alpha,
                        markersize=5, linewidth=1.2, zorder=3)
            elif post_1wk is not None and post_24wk is None:
                ax.plot([0, 1], [pre, post_1wk], 'o-', color=color, alpha=alpha,
                        markersize=5, linewidth=1.2, zorder=3)
                ax.plot([1, 3], [post_1wk, rehab], 'o--', color=color, alpha=alpha * 0.7,
                        markersize=5, linewidth=0.8, zorder=3)
            elif post_1wk is None and post_24wk is not None:
                ax.plot([0, 2], [pre, post_24wk], 'o--', color=color, alpha=alpha * 0.7,
                        markersize=5, linewidth=0.8, zorder=3)
                ax.plot([2, 3], [post_24wk, rehab], 'o-', color=color, alpha=alpha,
                        markersize=5, linewidth=1.2, zorder=3)
            else:
                ax.plot([0, 3], [pre, rehab], 'o--', color=color, alpha=alpha * 0.5,
                        markersize=5, linewidth=0.8, zorder=3)

        # Group mean trajectory
        mean_x, mean_y = [], []
        for xi in [0, 1, 2, 3]:
            if mean_vals[xi]:
                mean_x.append(xi)
                mean_y.append(np.mean(mean_vals[xi]))
        ax.plot(mean_x, mean_y, 's-', color='black',
                markersize=10, linewidth=3, zorder=5)

        # Stats: paired t-test pre vs rehab
        pre_arr = np.array(mean_vals[0])
        rehab_arr = np.array(mean_vals[3])
        t_stat, p_val = stats.ttest_rel(pre_arr, rehab_arr)
        diff = rehab_arr - pre_arr
        mean_diff = np.mean(diff)
        sem_diff = stats.sem(diff)

        # Count recovery categories using nadir
        n_recovered = 0
        n_improved = 0
        n_none = 0
        for animal in paired_animals:
            pre = f3[metric_key][f3_idx[animal]]
            rehab = l2[metric_key][l2_idx[animal]]
            post_1wk = ip[metric_key][ip_idx[animal]] if animal in has_ip else None
            post_24wk = p24[metric_key][p24_idx[animal]] if animal in has_24 else None
            post_vals = [v for v in [post_1wk, post_24wk] if v is not None]
            nadir = min(post_vals) if post_vals else rehab

            if rehab > pre * 0.8:
                n_recovered += 1
            elif rehab > nadir:
                n_improved += 1
            else:
                n_none += 1

        stats_text = (
            f'Pre vs Rehab: p={p_val:.4f}\n'
            f'Mean diff: {mean_diff:+.1f} +/- {sem_diff:.1f}\n'
            f'Recovered (>80% pre): {n_recovered}/{n}\n'
            f'Improved from nadir: {n_improved}/{n}\n'
            f'No improvement: {n_none}/{n}'
        )
        ax.text(0.5, 0.97, stats_text, transform=ax.transAxes, fontsize=9,
                verticalalignment='top', horizontalalignment='center',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8),
                family='monospace')

        ax.set_xticks([0, 1, 2, 3])
        ax.set_xticklabels(['Pre-Injury\n(Final 3)', '1 Wk\nPost-Injury',
                            '2-4 Wk\nPost-Injury', 'Rehab\n(Last 2)'], fontsize=9)
        ax.set_ylabel(ylabel, fontsize=12)
        ax.set_xlim(-0.3, 3.3)
        ax.set_title(f'{metric_name}', fontsize=12, fontweight='bold')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        n_with_all = sum(1 for a in paired_animals if a in has_ip and a in has_24)
        legend_elements = [
            Line2D([0], [0], color='#2ca02c', marker='o', label='Recovered (>80% of pre)'),
            Line2D([0], [0], color='#ff7f0e', marker='o', label='Improved from nadir'),
            Line2D([0], [0], color='#d62728', marker='o', label='No improvement'),
            Line2D([0], [0], color='black', marker='s', linewidth=3, label='Group Mean'),
        ]
        if n_with_all < n:
            legend_elements.append(
                Line2D([0], [0], color='gray', linestyle='--', label='Missing timepoint'))
        ax.legend(handles=legend_elements, loc='upper right', fontsize=7)

    plt.tight_layout()
    out_path = os.path.join(output_dir, f'recovery_{group_name}.png')
    plt.savefig(out_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  Saved: {out_path}")


def plot_trajectory_waterfall(group_name, injury_type, window_data, output_dir):
    """Four-point trajectory + recovery waterfall revealing bimodal recovery.

    Left column: Pre -> 1Wk Post -> 2-4Wk Post -> Rehab individual trajectories
    Right column: Waterfall of recovery deltas (Rehab - 2-4 Post, sorted), showing
                  how non-responders mask the recovery of responders.
    """
    from matplotlib.lines import Line2D

    f3 = window_data['final_3']
    ip = window_data['immediate_post']
    p24 = window_data['2_4_post']
    l2 = window_data['last_2']

    f3_idx = {a: i for i, a in enumerate(f3['animals'])}
    ip_idx = {a: i for i, a in enumerate(ip['animals'])}
    p24_idx = {a: i for i, a in enumerate(p24['animals'])}
    l2_idx = {a: i for i, a in enumerate(l2['animals'])}

    # For trajectory: need at least pre + post + rehab; 2-4 post optional (dashed gap)
    core_animals = set(f3['animals']) & set(ip['animals']) & set(l2['animals'])
    has_24 = set(p24['animals'])
    traj_animals = sorted(core_animals)

    # For waterfall: compare rehab vs per-animal nadir (worst post-injury point)
    # Include any animal with at least one post-injury timepoint + rehab
    wf_animals_set = set(l2['animals']) & (set(ip['animals']) | set(p24['animals']))
    wf_animals = sorted(wf_animals_set)

    if len(traj_animals) < 2 and len(wf_animals) < 2:
        return

    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle(
        f'Recovery Analysis (Manually Scored): {group_name} - {injury_type}\n'
        f'Full Trajectory: Pre-Injury through Rehab',
        fontsize=14, fontweight='bold'
    )

    for row, (metric_key, metric_name, ylabel) in enumerate([
        ('eaten', 'Retrieved', '% Pellets Retrieved'),
        ('contacted', 'Contacted', '% Pellets Contacted'),
    ]):
        # === LEFT: Four-point trajectory ===
        ax_traj = axes[row, 0]

        if traj_animals:
            # Collect means for group trajectory
            mean_vals = {0: [], 1: [], 2: [], 3: []}

            for animal in traj_animals:
                pre = f3[metric_key][f3_idx[animal]]
                post = ip[metric_key][ip_idx[animal]]
                rehab = l2[metric_key][l2_idx[animal]]

                # Classify by final recovery outcome
                delta_rehab = rehab - post
                if rehab > pre * 0.8:
                    color = '#2ca02c'; alpha = 0.7
                elif delta_rehab > 0:
                    color = '#ff7f0e'; alpha = 0.6
                else:
                    color = '#d62728'; alpha = 0.5

                if animal in has_24:
                    p24_val = p24[metric_key][p24_idx[animal]]
                    # Full 4-point trajectory
                    ax_traj.plot([0, 1, 2, 3], [pre, post, p24_val, rehab], 'o-',
                                color=color, alpha=alpha, markersize=5, linewidth=1.2, zorder=3)
                    mean_vals[2].append(p24_val)
                else:
                    # Missing 2-4 post: solid pre→post, dashed post→rehab
                    ax_traj.plot([0, 1], [pre, post], 'o-',
                                color=color, alpha=alpha, markersize=5, linewidth=1.2, zorder=3)
                    ax_traj.plot([1, 3], [post, rehab], 'o--',
                                color=color, alpha=alpha * 0.7, markersize=5, linewidth=0.8, zorder=3)

                mean_vals[0].append(pre)
                mean_vals[1].append(post)
                mean_vals[3].append(rehab)

            # Group mean trajectory
            mean_x = []
            mean_y = []
            for xi in [0, 1, 2, 3]:
                if mean_vals[xi]:
                    mean_x.append(xi)
                    mean_y.append(np.mean(mean_vals[xi]))
            ax_traj.plot(mean_x, mean_y, 's-',
                        color='black', markersize=10, linewidth=3, zorder=5)

        ax_traj.set_xticks([0, 1, 2, 3])
        ax_traj.set_xticklabels(['Pre-Injury\n(Final 3)', '1 Wk\nPost-Injury',
                                  '2-4 Wk\nPost-Injury', 'Rehab\n(Last 2)'], fontsize=9)
        ax_traj.set_ylabel(ylabel, fontsize=11)
        n_with_24 = sum(1 for a in traj_animals if a in has_24)
        n_label = f'N={len(traj_animals)}'
        if n_with_24 < len(traj_animals):
            n_label += f' ({n_with_24} with 2-4wk)'
        ax_traj.set_title(f'{metric_name}: Individual Trajectories ({n_label})',
                         fontsize=11, fontweight='bold')
        ax_traj.spines['top'].set_visible(False)
        ax_traj.spines['right'].set_visible(False)
        ax_traj.set_xlim(-0.3, 3.3)

        traj_legend = [
            Line2D([0], [0], color='#2ca02c', marker='o', label='Recovered (>80% of pre)'),
            Line2D([0], [0], color='#ff7f0e', marker='o', label='Improved from post'),
            Line2D([0], [0], color='#d62728', marker='o', label='No improvement'),
            Line2D([0], [0], color='black', marker='s', linewidth=3, label='Group Mean'),
            Line2D([0], [0], color='gray', linestyle='--', label='Missing 2-4 wk'),
        ]
        ax_traj.legend(handles=traj_legend, loc='upper right', fontsize=7)

        # === RIGHT: Waterfall of recovery deltas ===
        ax_wf = axes[row, 1]

        # Per-animal nadir: worst of available post-injury timepoints
        deltas = []
        for animal in wf_animals:
            rehab_val = l2[metric_key][l2_idx[animal]]
            post_vals = []
            if animal in ip_idx:
                post_vals.append(ip[metric_key][ip_idx[animal]])
            if animal in p24_idx:
                post_vals.append(p24[metric_key][p24_idx[animal]])
            if not post_vals:
                continue
            nadir = min(post_vals)
            deltas.append(rehab_val - nadir)

        if not deltas:
            ax_wf.text(0.5, 0.5, 'Insufficient data', transform=ax_wf.transAxes,
                      ha='center', va='center', fontsize=12, color='gray')
            continue

        # Sort descending (most recovery first)
        sorted_idx = np.argsort(deltas)[::-1]
        sorted_deltas = [deltas[i] for i in sorted_idx]
        bar_colors = ['#2ca02c' if d > 0 else '#d62728' if d < 0 else '#888888'
                      for d in sorted_deltas]

        x_pos = np.arange(len(sorted_deltas))
        ax_wf.bar(x_pos, sorted_deltas, color=bar_colors, alpha=0.8,
                  edgecolor='black', linewidth=0.5)
        ax_wf.axhline(y=0, color='black', linewidth=1, linestyle='-')

        # Summary stats
        n_total = len(sorted_deltas)
        n_improved = sum(1 for d in sorted_deltas if d > 0)
        n_declined = sum(1 for d in sorted_deltas if d < 0)
        n_same = sum(1 for d in sorted_deltas if d == 0)
        mean_all = np.mean(sorted_deltas)

        # Responder-only stats
        pos_deltas = [d for d in sorted_deltas if d > 0]
        mean_responders = np.mean(pos_deltas) if pos_deltas else 0

        # Wilcoxon signed-rank on all deltas (is recovery significant overall?)
        p_str_all = 'N/A'
        if n_total >= 5:
            try:
                non_zero = [d for d in sorted_deltas if d != 0]
                if len(non_zero) >= 5:
                    _, wil_p = stats.wilcoxon(non_zero)
                    p_str_all = f'p={wil_p:.4f}' if wil_p >= 0.0001 else 'p<0.0001'
            except Exception:
                pass

        stats_text = (
            f'All animals (N={n_total}):\n'
            f'  Improved: {n_improved}  |  Declined: {n_declined}  |  Same: {n_same}\n'
            f'  Mean delta: {mean_all:+.1f}%  |  Wilcoxon {p_str_all}\n'
            f'Responders only (N={len(pos_deltas)}):\n'
            f'  Mean recovery: +{mean_responders:.1f}%'
        )
        ax_wf.text(0.98, 0.98, stats_text, transform=ax_wf.transAxes, fontsize=8,
                  verticalalignment='top', horizontalalignment='right',
                  bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8),
                  family='monospace')

        ax_wf.set_xlabel('Animals (sorted by recovery)', fontsize=10)
        ax_wf.set_ylabel(f'Change in {ylabel}\n(Rehab - Nadir)', fontsize=10)
        ax_wf.set_title(f'{metric_name}: Recovery Waterfall (from nadir)', fontsize=11, fontweight='bold')
        ax_wf.spines['top'].set_visible(False)
        ax_wf.spines['right'].set_visible(False)
        ax_wf.set_xticks([])

    plt.tight_layout()
    out_path = os.path.join(output_dir, f'trajectory_{group_name}.png')
    plt.savefig(out_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  Saved: {out_path}")


def plot_mega_cohort(all_group_data, output_dir):
    """Normalized mega-cohort analysis: all groups on common baseline.

    Each animal's values are expressed as % of their own pre-injury (Final 3) baseline.
    This allows pooling across groups with different raw performance levels.

    Figure 1 (mega_cohort_trajectory): 4-point normalized trajectories
      - Top-left: All animals pooled, individual lines + grand mean
      - Top-right: Waterfall of recovery (Rehab - nadir), using per-animal min post-injury
      - Bottom-left: Group mean trajectories overlaid, colored by injury type
      - Bottom-right: Violin/box of normalized rehab by injury type

    Figure 2 (mega_cohort_raw): Same layout but with raw % values (not normalized)
    """
    from matplotlib.lines import Line2D

    # Injury type color map
    injury_colors = {
        'Early Study': '#9467bd',
        'Transection': '#d62728',
        'Contusion 50kd': '#2ca02c',
        'Contusion 60kd': '#ff7f0e',
        'Contusion 70kd': '#e377c2',
        'Pyramidotomy': '#1f77b4',
    }

    # Collect per-animal 4-point data across all groups
    # Each record: (group, injury_type, animal_id, pre, post_1wk, post_24wk, rehab)
    records = []
    for group_name, gd in all_group_data.items():
        wd = gd['window_data']
        injury_type = gd['injury_type']

        f3 = wd['final_3']
        ip = wd['immediate_post']
        p24 = wd['2_4_post']
        l2 = wd['last_2']

        f3_idx = {a: i for i, a in enumerate(f3['animals'])}
        ip_idx = {a: i for i, a in enumerate(ip['animals'])}
        p24_idx = {a: i for i, a in enumerate(p24['animals'])}
        l2_idx = {a: i for i, a in enumerate(l2['animals'])}

        # Need at least pre + one post-injury + rehab
        for animal in f3['animals']:
            if animal not in l2_idx:
                continue
            pre = f3['eaten'][f3_idx[animal]]
            rehab = l2['eaten'][l2_idx[animal]]
            post_1wk = ip['eaten'][ip_idx[animal]] if animal in ip_idx else None
            post_24wk = p24['eaten'][p24_idx[animal]] if animal in p24_idx else None

            pre_c = f3['contacted'][f3_idx[animal]]
            rehab_c = l2['contacted'][l2_idx[animal]]
            post_1wk_c = ip['contacted'][ip_idx[animal]] if animal in ip_idx else None
            post_24wk_c = p24['contacted'][p24_idx[animal]] if animal in p24_idx else None

            if post_1wk is None and post_24wk is None:
                continue

            records.append({
                'group': group_name, 'injury': injury_type, 'animal': animal,
                'pre': pre, 'post_1wk': post_1wk, 'post_24wk': post_24wk, 'rehab': rehab,
                'pre_c': pre_c, 'post_1wk_c': post_1wk_c, 'post_24wk_c': post_24wk_c, 'rehab_c': rehab_c,
            })

    if not records:
        return

    # For each metric (eaten, contacted), generate the mega-cohort figure
    for metric_suffix, metric_label, ylabel_raw, ylabel_norm in [
        ('', 'Retrieved', '% Pellets Retrieved', '% of Pre-Injury Baseline'),
        ('_c', 'Contacted', '% Pellets Contacted', '% of Pre-Injury Baseline'),
    ]:
        pre_key = 'pre' + metric_suffix
        p1_key = 'post_1wk' + metric_suffix
        p24_key = 'post_24wk' + metric_suffix
        rehab_key = 'rehab' + metric_suffix

        fig, axes = plt.subplots(2, 2, figsize=(18, 14))
        fig.suptitle(
            f'Mega-Cohort Analysis (Manually Scored): {metric_label} — All Historical Groups\n'
            f'Normalized to Pre-Injury Baseline (N={len(records)} animals across {len(all_group_data)} groups)',
            fontsize=14, fontweight='bold'
        )

        # === Compute normalized values ===
        norm_records = []
        for r in records:
            pre_val = r[pre_key]
            if pre_val <= 0:
                continue  # Can't normalize if baseline is 0
            norm = {
                'group': r['group'], 'injury': r['injury'], 'animal': r['animal'],
                'pre_raw': pre_val,
                'pre': 100.0,  # normalized baseline
            }

            # Normalize each timepoint
            for tkey, nkey in [(p1_key, 'post_1wk'), (p24_key, 'post_24wk'), (rehab_key, 'rehab')]:
                raw_val = r[tkey]
                if raw_val is not None:
                    norm[nkey] = (raw_val / pre_val) * 100
                    norm[nkey + '_raw'] = raw_val
                else:
                    norm[nkey] = None
                    norm[nkey + '_raw'] = None

            # Per-animal nadir: min of available post-injury timepoints
            post_vals = [norm[k] for k in ['post_1wk', 'post_24wk'] if norm[k] is not None]
            norm['nadir'] = min(post_vals) if post_vals else None
            norm['nadir_raw'] = min(
                [r[k] for k in [p1_key, p24_key] if r[k] is not None]
            ) if post_vals else None

            norm_records.append(norm)

        if not norm_records:
            plt.close()
            continue

        # === TOP LEFT: All animals pooled — normalized trajectories ===
        ax = axes[0, 0]
        mean_vals = {0: [], 1: [], 2: [], 3: []}

        for nr in norm_records:
            xs, ys = [0], [nr['pre']]
            mean_vals[0].append(nr['pre'])
            if nr['post_1wk'] is not None:
                xs.append(1); ys.append(nr['post_1wk'])
                mean_vals[1].append(nr['post_1wk'])
            if nr['post_24wk'] is not None:
                xs.append(2); ys.append(nr['post_24wk'])
                mean_vals[2].append(nr['post_24wk'])
            if nr['rehab'] is not None:
                xs.append(3); ys.append(nr['rehab'])
                mean_vals[3].append(nr['rehab'])

            # Color by recovery outcome
            if nr['rehab'] is not None and nr['rehab'] > 80:
                color = '#2ca02c'; alpha = 0.25
            elif nr['rehab'] is not None and nr['nadir'] is not None and nr['rehab'] > nr['nadir']:
                color = '#ff7f0e'; alpha = 0.2
            else:
                color = '#d62728'; alpha = 0.2
            ax.plot(xs, ys, 'o-', color=color, alpha=alpha, markersize=3, linewidth=0.8, zorder=2)

        # Grand mean
        gmean_x, gmean_y = [], []
        for xi in [0, 1, 2, 3]:
            if mean_vals[xi]:
                gmean_x.append(xi)
                gmean_y.append(np.mean(mean_vals[xi]))
        ax.plot(gmean_x, gmean_y, 's-', color='black', markersize=12, linewidth=3.5, zorder=5)

        # SEM shading
        for xi in gmean_x:
            if len(mean_vals[xi]) > 1:
                sem = stats.sem(mean_vals[xi])
                m = np.mean(mean_vals[xi])
                ax.fill_between([xi - 0.1, xi + 0.1], m - sem, m + sem,
                               color='black', alpha=0.15, zorder=4)

        ax.axhline(y=100, color='gray', linestyle='--', linewidth=1, alpha=0.5)
        ax.set_xticks([0, 1, 2, 3])
        ax.set_xticklabels(['Pre-Injury\n(Final 3)', '1 Wk\nPost-Injury',
                            '2-4 Wk\nPost-Injury', 'Rehab\n(Last 2)'], fontsize=9)
        ax.set_ylabel(ylabel_norm, fontsize=11)
        ax.set_title(f'All Animals Pooled (N={len(norm_records)})', fontsize=12, fontweight='bold')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_xlim(-0.3, 3.3)

        legend_items = [
            Line2D([0], [0], color='#2ca02c', marker='o', alpha=0.5, label='Recovered (>80% of pre)'),
            Line2D([0], [0], color='#ff7f0e', marker='o', alpha=0.5, label='Improved from nadir'),
            Line2D([0], [0], color='#d62728', marker='o', alpha=0.5, label='No improvement'),
            Line2D([0], [0], color='black', marker='s', linewidth=3, label='Grand Mean'),
        ]
        ax.legend(handles=legend_items, loc='upper right', fontsize=7)

        # === TOP RIGHT: Recovery waterfall (Rehab - Nadir, normalized) ===
        ax = axes[0, 1]

        deltas = []
        delta_injuries = []
        for nr in norm_records:
            if nr['nadir'] is not None and nr['rehab'] is not None:
                deltas.append(nr['rehab'] - nr['nadir'])
                delta_injuries.append(nr['injury'])

        if deltas:
            sorted_idx = np.argsort(deltas)[::-1]
            sorted_deltas = [deltas[i] for i in sorted_idx]
            sorted_injuries = [delta_injuries[i] for i in sorted_idx]
            bar_colors = [injury_colors.get(inj, '#888888') for inj in sorted_injuries]

            x_pos = np.arange(len(sorted_deltas))
            ax.bar(x_pos, sorted_deltas, color=bar_colors, alpha=0.8,
                   edgecolor='none', linewidth=0)
            ax.axhline(y=0, color='black', linewidth=1)

            n_total = len(sorted_deltas)
            n_improved = sum(1 for d in sorted_deltas if d > 0)
            n_declined = sum(1 for d in sorted_deltas if d < 0)
            mean_all = np.mean(sorted_deltas)

            pos_deltas = [d for d in sorted_deltas if d > 0]
            mean_resp = np.mean(pos_deltas) if pos_deltas else 0

            p_str = 'N/A'
            try:
                non_zero = [d for d in sorted_deltas if d != 0]
                if len(non_zero) >= 5:
                    _, wil_p = stats.wilcoxon(non_zero)
                    p_str = f'p={wil_p:.4f}' if wil_p >= 0.0001 else 'p<0.0001'
            except Exception:
                pass

            stats_text = (
                f'N={n_total} animals\n'
                f'Improved: {n_improved} | Declined: {n_declined}\n'
                f'Mean: {mean_all:+.1f}% of baseline\n'
                f'Wilcoxon: {p_str}\n'
                f'Responders (N={len(pos_deltas)}): +{mean_resp:.1f}%'
            )
            ax.text(0.98, 0.98, stats_text, transform=ax.transAxes, fontsize=8,
                    va='top', ha='right', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8),
                    family='monospace')

            # Legend for injury type colors
            seen = []
            inj_legend = []
            for inj in sorted_injuries:
                if inj not in seen:
                    seen.append(inj)
                    inj_legend.append(Line2D([0], [0], color=injury_colors.get(inj, '#888888'),
                                            marker='s', linestyle='', markersize=8, label=inj))
            ax.legend(handles=inj_legend, loc='lower right', fontsize=7, title='Injury Type',
                     title_fontsize=8)

        ax.set_xlabel('Animals (sorted by recovery)', fontsize=10)
        ax.set_ylabel(f'Recovery from Nadir\n(% of pre-injury baseline)', fontsize=10)
        ax.set_title('Recovery Waterfall: Rehab - Nadir (Normalized)', fontsize=12, fontweight='bold')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_xticks([])

        # === BOTTOM LEFT: Group mean trajectories by injury type ===
        ax = axes[1, 0]

        injury_groups = {}
        for nr in norm_records:
            inj = nr['injury']
            if inj not in injury_groups:
                injury_groups[inj] = []
            injury_groups[inj].append(nr)

        for injury_type, inj_records in injury_groups.items():
            color = injury_colors.get(injury_type, '#888888')

            # Individual traces (very transparent)
            for nr in inj_records:
                xs, ys = [0], [nr['pre']]
                if nr['post_1wk'] is not None:
                    xs.append(1); ys.append(nr['post_1wk'])
                if nr['post_24wk'] is not None:
                    xs.append(2); ys.append(nr['post_24wk'])
                if nr['rehab'] is not None:
                    xs.append(3); ys.append(nr['rehab'])
                ax.plot(xs, ys, '-', color=color, alpha=0.1, linewidth=0.5, zorder=2)

            # Group mean
            gm = {0: [], 1: [], 2: [], 3: []}
            for nr in inj_records:
                gm[0].append(nr['pre'])
                if nr['post_1wk'] is not None: gm[1].append(nr['post_1wk'])
                if nr['post_24wk'] is not None: gm[2].append(nr['post_24wk'])
                if nr['rehab'] is not None: gm[3].append(nr['rehab'])

            mx, my = [], []
            for xi in [0, 1, 2, 3]:
                if gm[xi]:
                    mx.append(xi)
                    my.append(np.mean(gm[xi]))
            ax.plot(mx, my, 'o-', color=color, markersize=8, linewidth=2.5, zorder=5,
                    label=f'{injury_type} (N={len(inj_records)})')

            # SEM bars
            for xi in mx:
                if len(gm[xi]) > 1:
                    sem = stats.sem(gm[xi])
                    m = np.mean(gm[xi])
                    ax.errorbar(xi, m, yerr=sem, color=color, capsize=4, linewidth=1.5, zorder=4)

        ax.axhline(y=100, color='gray', linestyle='--', linewidth=1, alpha=0.5)
        ax.set_xticks([0, 1, 2, 3])
        ax.set_xticklabels(['Pre-Injury\n(Final 3)', '1 Wk\nPost-Injury',
                            '2-4 Wk\nPost-Injury', 'Rehab\n(Last 2)'], fontsize=9)
        ax.set_ylabel(ylabel_norm, fontsize=11)
        ax.set_title('Group Mean Trajectories by Injury Type', fontsize=12, fontweight='bold')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.legend(loc='upper right', fontsize=8)
        ax.set_xlim(-0.3, 3.3)

        # === BOTTOM RIGHT: Box/violin of normalized rehab by injury type ===
        ax = axes[1, 1]

        injury_types_sorted = sorted(injury_groups.keys(),
                                     key=lambda k: np.mean([nr['rehab'] for nr in injury_groups[k]
                                                           if nr['rehab'] is not None]))
        box_data = []
        box_labels = []
        box_colors_list = []
        for inj in injury_types_sorted:
            vals = [nr['rehab'] for nr in injury_groups[inj] if nr['rehab'] is not None]
            if vals:
                box_data.append(vals)
                box_labels.append(f'{inj}\n(N={len(vals)})')
                box_colors_list.append(injury_colors.get(inj, '#888888'))

        if box_data:
            bp = ax.boxplot(box_data, patch_artist=True, widths=0.6,
                           medianprops=dict(color='black', linewidth=2))
            for patch, color in zip(bp['boxes'], box_colors_list):
                patch.set_facecolor(color)
                patch.set_alpha(0.6)

            # Overlay individual points
            np.random.seed(42)
            for i, vals in enumerate(box_data):
                jitter = np.random.uniform(-0.15, 0.15, len(vals))
                ax.scatter(np.full(len(vals), i + 1) + jitter, vals,
                          color=box_colors_list[i], s=25, zorder=5, alpha=0.7,
                          edgecolor='white', linewidth=0.3)

            ax.set_xticklabels(box_labels, fontsize=8)
            ax.axhline(y=100, color='gray', linestyle='--', linewidth=1, alpha=0.5,
                       label='Pre-injury baseline')

            # Kruskal-Wallis across injury types
            if len(box_data) >= 2:
                try:
                    kw_stat, kw_p = stats.kruskal(*box_data)
                    kw_str = f'Kruskal-Wallis: H={kw_stat:.1f}, p={kw_p:.4f}' if kw_p >= 0.0001 else f'Kruskal-Wallis: H={kw_stat:.1f}, p<0.0001'
                    ax.set_xlabel(kw_str, fontsize=9)
                except Exception:
                    pass

        ax.set_ylabel(ylabel_norm, fontsize=11)
        ax.set_title(f'Post-Rehab Recovery by Injury Type', fontsize=12, fontweight='bold')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.legend(loc='upper left', fontsize=8)

        plt.tight_layout()
        suffix = 'eaten' if metric_suffix == '' else 'contacted'
        out_path = os.path.join(output_dir, f'mega_cohort_{suffix}.png')
        plt.savefig(out_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        print(f"  Saved: {out_path}")


def plot_combined_overview(all_group_data, output_dir):
    """Create a combined overview: all groups side by side per window."""
    group_names = list(all_group_data.keys())
    n_groups = len(group_names)

    fig, axes = plt.subplots(2, 2, figsize=(max(14, n_groups * 2.5), 11))
    fig.suptitle('Behavior Performance (Manually Scored): All Historical Groups',
                 fontsize=16, fontweight='bold')

    group_colors = plt.cm.Set2(np.linspace(0, 1, n_groups))

    for row, (metric_key, metric_name, ylabel) in enumerate([
        ('eaten', 'Retrieved', '% Pellets Retrieved'),
        ('contacted', 'Contacted', '% Pellets Contacted'),
    ]):
        for col, (wk, wlabel) in enumerate([
            ('final_3', 'Final 3 (Pre-Injury)'),
            ('last_2', 'Rehab Pillar (Last 2)'),
        ]):
            ax = axes[row, col]
            x = np.arange(n_groups)

            means, sems = [], []
            for gname in group_names:
                d = all_group_data[gname]['window_data'][wk][metric_key]
                means.append(np.mean(d) if len(d) > 0 else 0)
                sems.append(stats.sem(d) if len(d) > 1 else 0)

            bars = ax.bar(x, means, yerr=sems, color=[group_colors[i] for i in range(n_groups)],
                         capsize=4, alpha=0.85, edgecolor='black', linewidth=0.8)

            np.random.seed(42)
            for i, gname in enumerate(group_names):
                d = all_group_data[gname]['window_data'][wk][metric_key]
                if len(d) > 0:
                    jitter = np.random.uniform(-0.2, 0.2, len(d))
                    ax.scatter(x[i] + jitter, d, color='black', s=20, zorder=5, alpha=0.5,
                              edgecolor='white', linewidth=0.3)

            labels = [f"{gn}\n({all_group_data[gn]['injury_type']})" for gn in group_names]
            ax.set_xticks(x)
            ax.set_xticklabels(labels, fontsize=8, ha='center')
            ax.set_ylabel(ylabel, fontsize=11)
            ax.set_title(f'{metric_name} - {wlabel}', fontsize=12, fontweight='bold')
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

    plt.tight_layout()
    out_path = os.path.join(output_dir, 'behavior_all_groups_overview.png')
    plt.savefig(out_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  Saved overview: {out_path}")


def main():
    print("=" * 70)
    print("GENERATING HISTORICAL BEHAVIOR PLOTS")
    print("  Style: Matching fig5_behavior_layout4.png (CNT reference)")
    print("  Stats: GEE (binomial, exchangeable, animal cluster) + Holm")
    print("         Fallback: chi-sq/Fisher when GEE can't converge")
    print("  Last 2: Last 2 Pillar days in continuous rehab block")
    print(f"  Learner criterion: Final 3 avg eaten% > {LEARNER_EATEN_THRESHOLD}%")
    print("  Data: Raw pellet scores from '1 - ENTER DATA HERE' tabs")
    print("=" * 70)

    all_group_data = {}

    for group_name, (filename, injury_type) in GROUP_FILES.items():
        filepath = os.path.join(SCRIPT_DIR, filename)
        if not os.path.exists(filepath):
            print(f"\nWARNING: {filename} not found, skipping {group_name}")
            continue

        print(f"\n--- {group_name} ({injury_type}) ---")
        data, test_types, test_meta, pellet_records = read_group_data(filepath)
        windows = get_time_windows(test_types, test_meta)
        learners = get_learners(data, windows['final_3'])
        n_excluded = len(data) - len(learners)

        window_data = get_animal_window_values(data, windows, learners)
        pellet_df = build_pellet_dataframe(pellet_records, windows, learners)

        print(f"  Total animals: {len(data)}, Learners: {len(learners)}, "
              f"Non-learners excluded: {n_excluded}")
        print(f"  Total pellet observations: {len(pellet_df)}")
        print(f"  Windows:")
        for wk_idx, wk in enumerate(WINDOW_KEYS):
            tts = windows[wk]
            n_animals = len(window_data[wk]['animals'])
            n_pellets = int((pellet_df['window'] == wk_idx).sum()) if not pellet_df.empty else 0
            print(f"    {wk:20s}: {n_animals:2d} animals, {n_pellets:5d} pellets | "
                  f"{', '.join(tts)}")

        plot_group(group_name, injury_type, window_data, pellet_df, len(learners), OUTPUT_DIR)
        plot_recovery(group_name, injury_type, window_data, OUTPUT_DIR)
        plot_trajectory_waterfall(group_name, injury_type, window_data, OUTPUT_DIR)

        all_group_data[group_name] = {
            'injury_type': injury_type,
            'window_data': window_data,
            'windows': windows,
        }

    print(f"\n--- Combined Overview ---")
    plot_combined_overview(all_group_data, OUTPUT_DIR)

    print(f"\n--- Mega-Cohort Analysis ---")
    plot_mega_cohort(all_group_data, OUTPUT_DIR)

    print(f"\nAll figures saved to: {OUTPUT_DIR}")


if __name__ == '__main__':
    main()
