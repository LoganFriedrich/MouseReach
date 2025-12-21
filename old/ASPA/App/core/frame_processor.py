

from datetime import timedelta, datetime
import os
import glob
import sys
import threading
import time
import traceback
import tqdm

# # for video dark check
#   import ffmpeg

# from click import group
import numpy as np
import warnings
with warnings.catch_warnings():
    warnings.simplefilter(action='ignore', category=FutureWarning)
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

pd.options.mode.chained_assignment = None  # default='warn'

from video_player import VideoPlayer
from frame_info import FrameInfo
from swipe_info import SwipeInfo
from stats import Stats
from action_detector import ActionDetector
from df_viewer import DataFrameViewer

from qtpy.QtWidgets import QApplication

import concurrent.futures

class FrameInfoProcessor:
    def __init__(self, tblFrameInfo, tblSwipeInfo, reachFilter, tn_count=6, ui=True) -> None:        
        self.thumbnails = []
        self.highlight = None
        self.play = False
        self.playback_speed = 0.1
        self.init = False
        self.play_swipes_only = False
        self.has_ui = ui
        self.animal_group = 'D'

        self.init_frame_info(tbl=tblFrameInfo)
        self.init_swipe_info(tbl=tblSwipeInfo, reachFilter=reachFilter)
        self.init_vdo_player(tn_count=tn_count)
        self.init_stats()
        self.init_swipe_detector()

        self.init = True


    #region Stats

    def init_stats(self):
        self.stats = Stats(animal_group=self.animal_group, ui=self.has_ui)

    def get_info(self, vdo_filename):
        return self.stats.get_info(vdo_filename)

    #endregion Stats

    #region FrameInfo

    def display_trackbar(self):
        self.vdo_player.display_trackbar(num_frames=self.frame_info.num_rows)

    def init_frame_info(self, tbl):
        self.frame_info = FrameInfo(tbl=tbl, ui=self.has_ui)

        if self.has_ui:
            self.frame_info.tblFrameInfo.itemSelectionChanged.connect(self.tblFrameInfo_itemSelectionChanged)

    def tblFrameInfo_itemSelectionChanged(self):
        try:
            if not self.init:
                return
            # Get the current selected row
            row = self.frame_info.tblFrameInfo.currentRow()

            self.vdo_player.current_frame_idx = int(float(self.frame_info.tblFrameInfo.item(row, 0).text()))
            
            self.vdo_player.update_vdo_display()
            self.vdo_player.can_vp.draw()
        except:
            print(traceback.format_exc())

    #endregion FrameInfo

    #region SwipeInfo

    def init_swipe_info(self, tbl, reachFilter):
        self.swipe_info = SwipeInfo(tbl=tbl, reachFilter=reachFilter, ui=self.has_ui)

        if self.has_ui:
            self.swipe_info.tblSwipeInfo.itemSelectionChanged.connect(self.tblSwipeInfo_itemSelectionChanged)

    def tblSwipeInfo_itemSelectionChanged(self):
        try:
            if not self.init:
                return
            # Get the current selected row
            row = self.swipe_info.tblSwipeInfo.currentRow()
            self.vdo_player.current_frame_idx = int(float(self.swipe_info.tblSwipeInfo.item(row, 0).text()))
            self.vdo_player.update_vdo_display()
            self.vdo_player.can_vp.draw()

            # Highlight swipe in swipe detection viewer
            self.action_detector.highlight_swipe(x=self.vdo_player.current_frame_idx, y=int(float(self.swipe_info.tblSwipeInfo.item(row, 20).text())))
        except:
            print(traceback.format_exc())

    #endregion SwipeInfo

    #region Video Player

    def init_vdo_player(self, tn_count=6):
        if not self.has_ui:
            return
        self.vdo_player = VideoPlayer(tn_count=tn_count)

        # Connect the click event to the update_main_view function
        self.vdo_player.fig_vp.canvas.mpl_connect('button_press_event', self.thumbnail_on_click)

        # Connect the click event on the navigation buttons to the on_click function
        self.vdo_player.btn_prev.on_clicked(self.nav_on_click)
        self.vdo_player.btn_next.on_clicked(self.nav_on_click)
        self.vdo_player.btn_play.on_clicked(self.play_on_click)

    def play_on_click(self, event):
        # self.vdo_player.play = not self.vdo_player.play
        # if self.vdo_player.play:
        #     self.vdo_player.btn_play.label.set_text('Pause')
        #     self.vdo_player.vdo_thread.start()
        # else:
        #     self.vdo_player.vdo_thread.join()
        #     self.vdo_player.btn_play.label.set_text('Play')
        #     self.vdo_player.vdo_thread = threading.Thread(target=self.vdo_player.play_vdo)
        self.vdo_player.play_on_click(event)

    #region Thumbnail

    def nav_on_click(self, event):

        self.vdo_player.nav_on_click(event)

        self.frame_info.tblFrameInfo.itemSelectionChanged.disconnect(self.tblFrameInfo_itemSelectionChanged)
        self.vdo_player.pb_status = self.frame_info.select_row(self.vdo_player.current_frame_idx)
        self.frame_info.tblFrameInfo.itemSelectionChanged.connect(self.tblFrameInfo_itemSelectionChanged)

    # Function to update the main view
    def thumbnail_on_click(self, event):

        self.vdo_player.thumbnail_on_click(event)

        self.frame_info.tblFrameInfo.itemSelectionChanged.disconnect(self.tblFrameInfo_itemSelectionChanged)
        self.vdo_player.pb_status = self.frame_info.select_row(self.vdo_player.current_frame_idx)
        self.frame_info.tblFrameInfo.itemSelectionChanged.connect(self.tblFrameInfo_itemSelectionChanged)

    #endregion Thumbnail

    #endregion Video Player

    #region Tray Motion

    def detect_tray_motion(self, start, end):
        # Get rows that have body_coords between start and end
        self.tray_frames = self.frame_info.df[(self.frame_info.df['bodyparts_coords'] >= start) &
                                              (self.frame_info.df['bodyparts_coords'] <= end)]
        
        self.action_detector.detect_tray_movmt(df=self.tray_frames)

    #endregion Tray Motion

    #region Swipe Detector

    def init_swipe_detector(self):
        self.action_detector = ActionDetector(ui=self.has_ui)

        if self.has_ui:
            self.action_detector.can_feat.mpl_connect('button_press_event', self.detection_on_click)
            self.action_detector.can_feat.mpl_connect('pick_event', self.detection_on_pick)
            self.action_detector.can_feat.mpl_connect('pick_event', self.detection_on_unpick)

    def detection_on_pick(self, event):
        line = event.artist
        ind = event.ind[0]
        data = line.get_data()
        xdata, ydata = data
        x = xdata[ind]
        y = ydata[ind]

        # Increase size and make semitransparent
        line.set_data(xdata, ydata) 
        line.set_markersize(20)
        line.set_alpha(0.5)
        self.action_detector.fig_feat.canvas.draw_idle()

    def detection_on_unpick(self, event):
        line = event.artist
        data = line.get_data()
        line.set_data(data)
        line.set_markersize(8) 
        line.set_alpha(1)
        self.action_detector.fig_feat.canvas.draw_idle()

    def detection_on_click(self, event):
        x = event.xdata # x data coordinate of click
        y = event.ydata # y data coordinate of click
        print(f'Click at ({x}, {y})')
        if x is None:
            return
        self.vdo_player.current_frame_idx = int(x)

        self.thumbnail_on_click(event)

    # # Function to identify "dark" videos
    # def get_average_intensity(video_path):
    #     # Read video using ffmpeg
    #     probe = ffmpeg.probe(video_path)
    #     video_info = next(s for s in probe['streams'] if s['codec_type'] == 'video')
    #     width = int(video_info['width'])
    #     height = int(video_info['height'])
    #     num_frames = int(video_info['nb_frames'])

    #     # Set up ffmpeg command to read video frames
    #     out, err = (
    #         ffmpeg
    #         .input(video_path)
    #         .output('pipe:', format='rawvideo', pix_fmt='gray')
    #         .run(capture_stdout=True, capture_stderr=True)
    #     )

    #     # Convert video frames to numpy array
    #     frames = np.frombuffer(out, np.uint8).reshape((num_frames, height, width))

    #     # Calculate average intensity
    #     total_intensity = np.sum(frames)
    #     average_intensity = total_intensity / (num_frames * height * width)

    #     return average_intensity
    
    #     def is_video_dark(video_path, threshold):
    #       average_intensity = get_average_intensity(video_path)
    #       if average_intensity < threshold:
    #           return "Video is dark"
    #       else:
    #           return "Video is not dark"

    #     # Set threshold for "dark"
    #     dark_threshold = 50 # adjust this value to change dark threshold

    # # Directory containing the video files
    # video_directory = "path/to/your/video/directory"

    # # Iterate over video files in the directory
    # for filename in os.listdir(video_directory):
    #     if filename.endswith(".mp4") or filename.endswith(".avi"):  # Adjust file extensions as needed
    #         video_path = os.path.join(video_directory, filename)
    #         result = is_video_dark(video_path, dark_threshold)
    #         print(f"Video: {filename}, Result: {result}")


    # Function to filter the frames where the mouse is swiping
    def detect_swipes(self, start, end, pellet_pos_th=8, threshold=300, rh_y_conf=0.7, 
                      min_frames=2, min_frames_std=10, th_adjacent=3, std_motion=0.1, title=''):
        # Get reference y
        ref_y = int(self.frame_info.df['Reference_y'].iloc[start])
        ref_x = int(self.frame_info.df['Reference_x'].iloc[start])

        # Get rows that have body_coords between start and end
        self.swipe_frames = self.frame_info.df[(self.frame_info.df['bodyparts_coords'] >= start) &
                                               (self.frame_info.df['bodyparts_coords'] <= end)]
        
        self.action_detector.df_full = self.swipe_frames
        self.action_detector.pellet_pos_th = pellet_pos_th
        self.swipe_info.pellet_pos_th = pellet_pos_th
        
        # # Get mode of pellet x, y where pellet likelihood is greater than 0.9
        # pellet_x = self.swipe_frames[self.swipe_frames['Pellet_likelihood'] > 0.9]['Pellet_x'].mode().iloc[0]
        # pellet_y = self.swipe_frames[self.swipe_frames['Pellet_likelihood'] > 0.9]['Pellet_y'].mode().iloc[0]

        # print(f'Pellet mode x: {pellet_x}, Pellet y: {pellet_y}')

        # # Get median of pellet x, y where pellet likelihood is greater than 0.9
        # self.action_detector.pellet_x_median = self.swipe_frames[self.swipe_frames['Pellet_likelihood'] > 0.9]['Pellet_x'].median()
        # self.action_detector.pellet_y_median = self.swipe_frames[self.swipe_frames['Pellet_likelihood'] > 0.9]['Pellet_y'].median()

        # print(f'Pellet median x: {self.action_detector.pellet_x_median}, Pellet y: {self.action_detector.pellet_y_median}')
        
        self.action_detector.detect_tray_movmt(df=self.swipe_frames, title=title)
        
        # xlim for swipe detector
        start = self.swipe_frames['bodyparts_coords'].iloc[0]
        end = self.swipe_frames['bodyparts_coords'].iloc[-1]
        
        # Print average right hand likelihood
        # print(f"Average RH likelihood({self.swipe_frames.shape[0]}): {self.swipe_frames['RightHand_likelihood'].mean()}")
        # print(f'No. of frames with RH likelihood > 0.9: {self.swipe_frames[self.swipe_frames["RightHand_likelihood"] > 0.9].shape[0]}/{self.swipe_frames.shape[0]}')

        # Get rows where right hand likelihood is greater than 0.9 and right hand y is greater than nose y
        self.swipe_frames = self.swipe_frames[(self.swipe_frames['Nose_y'] > ref_y-80) &
                                              (self.swipe_frames['Nose_x'] > ref_x+70) &
                                              (self.swipe_frames['Nose_x'] < ref_x+100) &
                                              (self.swipe_frames['Nose_likelihood'] > 0.5) &
                                              (self.swipe_frames['RightHand_likelihood'] > 0.1)&
                                              (self.swipe_frames['RightHand_y'] - self.swipe_frames['Nose_y'] > 8)]
        
        # Print average right hand likelihood
        # print(f"Average RH likelihood after filter({self.swipe_frames.shape[0]}): {self.swipe_frames['RightHand_likelihood'].mean()}")
        # print(f'No. of frames with RH likelihood > 0.9: {self.swipe_frames[self.swipe_frames["RightHand_likelihood"] > 0.9].shape[0]}/{self.swipe_frames.shape[0]}')
        
        self.swipe_seq_dict, y_min, y_max = self.action_detector.detect_swipes_new(df=self.swipe_frames, threshold=threshold, 
                                                                                   rh_y_conf=rh_y_conf, min_frames=min_frames, 
                                                                                   min_frames_std=min_frames_std,
                                                                                   th_adjacent=th_adjacent, 
                                                                                   std_motion=std_motion, title=title)
                
        if len(self.swipe_seq_dict) == 0:
            print(f'No swipes detected in {title}')
            # return None, None
        
        # Display all swipes in swipe info tab
        swipe_df = self.swipe_frames[self.swipe_frames['bodyparts_coords'].isin(self.swipe_seq_dict.keys())]
        swipe_df = self.action_detector.add_pellet_no(swipe_df)
        self.swipe_info.df = swipe_df
        self.swipe_info.df_full = self.action_detector.df_full
        self.swipe_info.y_min = y_min
        self.swipe_info.y_max = y_max
        self.swipe_info.pellet_x_median = self.action_detector.pellet_x_median
        self.swipe_info.pellet_y_median = self.action_detector.pellet_y_median
        self.swipe_info.pellet_loc_df = self.action_detector.pellet_loc_df
        self.swipe_info.tot_frames = end - start
        self.swipe_info.attention_frames = self.frame_info.df[(self.frame_info.df['Nose_likelihood'] > 0.9) & 
                                                              (self.frame_info.df['Nose_y'] > ref_y - 80)].shape[0]

        summary, df_swipe_analysis = self.analyse_swipes_new(title=title)

        if self.has_ui:
            self.action_detector.ax.set_xlim(start-100, end+100)
            self.action_detector.ax.set_ylim(bottom=y_min, top=y_max+10)
            self.action_detector.update_title_info(summary)

        # Add all input parameters to summary
        summary.update({'pellet_pos_th': pellet_pos_th, 'threshold': threshold, 'rh_y_conf': rh_y_conf,
                        'min_frames': min_frames, 'min_frames_std': min_frames_std, 'th_adjacent': th_adjacent,
                        'std_motion': std_motion})

        return summary, df_swipe_analysis

    def analyse_swipes_new(self, title):

        # Get missed swipes
        swipes_idx_list = list(self.swipe_seq_dict)
        missed_swipes = [item for item in swipes_idx_list if item not in self.swipe_seq_dict.keys()]

        if len(missed_swipes) > 0:
            print(f"Missed swipes in {title}: {missed_swipes}")

        if self.has_ui:
            row_indices = []
            for seq in self.swipe_seq_dict.values():
                row_indices.extend(seq.index)

            self.swipe_frames = self.swipe_frames.loc[row_indices]

            self.swipe_frames = self.action_detector.add_pellet_no(self.swipe_frames)

            self.vdo_player.swipe_frames = self.swipe_frames
            self.frame_info.display_df(self.swipe_frames)


        
        summary, df = self.swipe_info.analyse_swipe_sequences(self.swipe_seq_dict)

        df = self.add_video_info_to_df(df, title)
            
        return summary, df
    
    def add_video_info_to_df(self, df, title):
        # Add video info to df to all rows
        df['Date'] = datetime.strptime(title[:8], '%Y%m%d').strftime('%d-%b-%Y')
        df['Animal ID'] = title.split('_')[1]
        df['Tray'] = title.split('_')[-1]

        # move Date, Animal ID, Tray to the front
        cols = df.columns.tolist()
        cols = cols[-3:] + cols[:-3]
        # df = df[cols]
        df = df.reindex(columns=cols)
        # print(df.head())
        return df


    def save_swipe_analysis(self):
        self.swipe_info.save_swipe_analysis()

    #endregion Swipe Detector


def convert_timedelta(tdelta):
    days, seconds = tdelta.days, tdelta.seconds
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = (seconds % 60)
    return days, hours, minutes, seconds

def process_video(file_path):
    summary = None
    df_swipe_analysis = None
    vdo_start = time.time()
    title = file_path.split('\\')[-1][:15]
    try:
        # Find csv file with same name as video file
        csv_file = glob.glob(file_path.replace('.mp4', '*.csv'))[0]
    except:
            print(f'No deeplabcut csv file found for {title}')
            return None

    try:
        frame_proc = FrameInfoProcessor(tblFrameInfo=None,
                                        tblSwipeInfo=None,
                                        reachFilter=None,
                                        tn_count=10, ui=False)
        frame_proc.frame_info.read_csv(csv_file)
        # frame_proc.stats.df = frame_proc.frame_info.df
        end = frame_proc.frame_info.df.iloc[-1, 0]

        summary = {'Date': datetime.strptime(title[:8], '%Y%m%d').strftime('%d-%b-%Y'),
                'Animal ID': title.split('_')[1],
                'Tray': title.split('_')[-1]}
        
        swipe_summary, df_swipe_analysis = frame_proc.detect_swipes(start=1, end=end,
                                                pellet_pos_th = 10,
                                                threshold=300, rh_y_conf=0.7,
                                                min_frames=2, min_frames_std=10,
                                                th_adjacent=3, std_motion=0.1,
                                                title=title)
                                        
        if swipe_summary is not None:
            summary.update(swipe_summary)
        else:
            print(f'No swipes detected for {title}')
    except:
        # write to log file
        #log_file = open(time.strftime('log_%Y%m%d%H%M%S.txt'), 'w')
        log_file = open(f'log_{title}.txt', 'w')
        log_file.write(traceback.format_exc())
        log_file.close()
        print(traceback.format_exc())        

    d, h, m, s = convert_timedelta(timedelta(seconds=(time.time() - vdo_start)))
    print(f'{title} in {h}h, {m}m, {s}s', flush=True)

    return summary, df_swipe_analysis

def update_summary_with_manual_recording(summary_file, recorded_file):
    # Check if summary.xlsx exists
    if os.path.exists(summary_file):
        df = pd.read_excel(summary_file)
        info_df = pd.pd.read_excel(recorded_file, sheet_name='1 - ENTER DATA HERE') # usecols='A:G'

        # Sort by Animal ID, Date, Tray
        df = df.sort_values(by=['Animal ID', 'Date', 'Tray'])

        # Insert columns 'Eaten', 'Displaced' from info_df where Animal ID, Date, Tray match
        df['H Eaten'] = None
        df['H Displaced'] = None
        for i in range(len(df)):
            animal_id = df['Animal ID'].iloc[i]
            date = df['Date'].iloc[i]
            tray = df['Tray'].iloc[i]
            tray_no_str = f'.{int(tray[-1])-1}' if int(tray[-1])-1 != 0 else ''
            eaten = info_df[(info_df['Animal #'] == int(animal_id[-2:])) & (info_df['Test Date'] == stats.parse_date(date))][f'Eaten{tray_no_str}'].iloc[0]
            displaced = info_df[(info_df['Animal #'] == int(animal_id[-2:])) & (info_df['Test Date'] == date)][f'Displaced{tray_no_str}'].iloc[0]
            df['H Eaten'].iloc[i] = eaten
            df['H Displaced'].iloc[i] = displaced

        # Reorder columns so that H Eaten and H Displaced are next to Eaten and Displaced
        cols = list(df.columns)
        cols = cols[:7] + cols[-2:] + cols[7:-2]
        df = df[cols]

        app = QApplication(sys.argv)
        window = DataFrameViewer(df)
        window.show()
        sys.exit(app.exec_())

def generate_group_dlc_summary(folder_path, mt=True):
    start_time = time.time()
    print(folder_path)
    save_folder = 'A:\ASPA\AI'
    # Dataframe to store summary of all videos 
    df_summary = pd.DataFrame()
    df_swipe_info = pd.DataFrame()
    swipe_info_list = []

    # Write summary dataframe to excel, use parent folder name as filename
    group = folder_path.split('\\')[-2]
    swipe_info_excel_file = os.path.join(save_folder, f'swipe_info_{group}.xlsx')
    # Create a log file
    log_file = open(f'log_{group}.txt', 'w')

    # Get list of all video files that end with .mp4
    files = glob.glob(f'{folder_path}/**/*.mp4', recursive=True)
    
    print(f'Number of files: {len(files)}')
    log_file.write(f'Number of files: {len(files)}\n')

    if len(files) == 0:
        print(f'No DLC output in {folder_path}')
        return
    
    start_time = time.time()
    err_count = 0
    if mt:
        # Run using concurrent.futures
        print(f'Running all experiments with {os.cpu_count()-4} processes', flush=True)
        with concurrent.futures.ProcessPoolExecutor() as executor:
            results = executor.map(process_video, files)
        
        for file, result in zip(files, results):
            try:
                if result[0] is not None and result[1] is not None:                    
                    df_summary = df_summary._append(result[0], ignore_index=True)
                    if not result[1].empty:
                        swipe_info_list.append(result[1].iloc[:,:13])                    
                else:
                    err_count += 1
                    print(f'No swipes in {file}')
                    log_file.write(f'No swipes in {file}\n')
            except:
                err_count += 1
                print(f'Error in processing {file}')
                print(traceback.format_exc())
                log_file.write(f'Error in processing file {file}\n')
                log_file.write(traceback.format_exc())
        log_file.close()
        df_swipe_info = pd.concat(swipe_info_list, ignore_index=True)
        
    else:
        
        for file in files:
            try:
                summary, df_swipe_analysis = process_video(file)
                df_summary = pd.concat([df_summary, pd.DataFrame([summary])], ignore_index=True)
                if not df_swipe_analysis.empty:
                    swipe_info_list.append(df_swipe_analysis.iloc[:,:13])
            except:
                err_count += 1
                print(f'Error in processing {file}')
                log_file.write(f'Error in processing file {file}\n')
                log_file.write(traceback.format_exc())

        df_swipe_info = pd.concat(swipe_info_list, ignore_index=True)

    d, h, m, s = convert_timedelta(timedelta(seconds=(time.time() - start_time)))
    print(f'Finished {len(files)-err_count}/{len(files)} experiments in {h}h, {m}m, {s}s', flush=True)
    
    # df_summary.to_excel(os.path.join(folder_path, f'summary_{group}.xlsx'), index=False,
    # Adjust column widths to fit content
    writer = pd.ExcelWriter(os.path.join(save_folder, f'summary_{group}.xlsx'), engine='xlsxwriter')
    df_summary.to_excel(writer, index=False)

    swipe_info_writer = pd.ExcelWriter(swipe_info_excel_file, engine='xlsxwriter')
    df_swipe_info.to_excel(swipe_info_writer, index=False)
    swipe_info_writer.close()

    # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']

    # Set the column width and format.
    for idx, col in enumerate(df_summary):
        max_len = max([len(str(s)) for s in df_summary[col].values] + [len(col)])
        worksheet.set_column(idx, idx, max_len)
        # color column names that end with S, D, M with green, pink, yellow
        if col.endswith('S'):
            worksheet.write(0, idx, col, workbook.add_format({'bg_color': '#98FB98'}))
        elif col.endswith('D'):
            worksheet.write(0, idx, col, workbook.add_format({'bg_color': '#FFC0CB'}))
        elif col.endswith('M'):
            worksheet.write(0, idx, col, workbook.add_format({'bg_color': '#FFFF99'}))
        elif col.endswith(')'):
            worksheet.write(0, idx, col, workbook.add_format({'bg_color': '#FFFF99'}))


    writer.close()

    print(f'Summary saved to {os.path.join(save_folder, f"summary_{group}.xlsx")}', flush=True)

def is_nan(value):
    try:
        return np.isnan(value)
    except TypeError:
        return False

def manual_ai_comparison():
    # Load excel file with manual recording
    manual_scored_file = r'A:\ASPA\Data\ManualAIComparison\manual_recorded_pellet_reach_outcomess.xlsx'
    manual_df = pd.read_excel(manual_scored_file)

    # Add a sheet to the excel file to color the cells
    wb = load_workbook(manual_scored_file)
    
    # Make a copy of Sheet1 to Sheet2
    ws = wb['Sheet1']
    ws2 = wb.copy_worksheet(ws)
    ws2.title = 'Sheet2'      

    pastel_green_fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')
    pastel_pink_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    pastel_yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

    # Replace empty strings or whitespace with np.nan
    manual_df = manual_df.replace(r'^\s*$', np.nan, regex=True)
    # Remove columns that are empty with just headers
    manual_df = manual_df.dropna(axis=1, how='all')
    column_names = manual_df.columns
    
    # Ignore first column and split column names and get unique values
    column_names = column_names[1:]
    column_names = column_names.str.split(' ', expand=True)
    
    # Get only first part from the split column names
    column_names = np.unique([col[0] for col in column_names])

    match_count_dict = {'swipe successful': 0, 'pellet displaced': 0, 'swipe missed': 0, 'no swipes': 0}
    outcome_count_dict = {'swipe successful': 0, 'pellet displaced': 0, 'swipe missed': 0, 'no swipes': 0}

    for col_idx, col in enumerate(column_names):
        vdo = glob.glob(f'**/{col}.mp4', recursive=True)[0]
        try:
            if os.path.exists(vdo):
                summary = process_video(vdo)
            else:
                print(f'{vdo} not found')
                continue
        except:
            print(traceback.format_exc())
            continue

        # Get manual scored outcome columns for the video
        manual_cols = manual_df.columns[manual_df.columns.str.contains(col)]
        outcomes = manual_df[manual_cols].values

        indices = np.where(manual_df.columns.str.contains(col))[0]
        indices = [idx+1 for idx in indices]

        reach_outcome_dict = {'S': 'swipe successful', 'D': 'pellet displaced', 'M': 'swipe missed', 'N': 'no swipes'}

        print(f'Processing {col}...')
        for pell_idx, outcome in enumerate(outcomes):
            frame_no = outcome[0]
            reach_outcome = outcome[1]

            if is_nan(frame_no) or is_nan(reach_outcome):
                continue
            
            try:
                outcome_count_dict[reach_outcome] += 1
                try:
                    ai_outcome = summary[f'{pell_idx+1}']
                except:
                    ai_outcome = '[S0|D1|M0](N)(0)' # No swipes detected
                # regex ai_outcome with format '[S0|D1|M41](D)(346)' to get reach_outcome = D, frame_no = 346
                ai_reach_outcome = re.findall(r"\(.*?\)", ai_outcome)[0][1:-1]
                ai_frame_no = int(re.findall(r"\(.*?\)", ai_outcome)[1][1:-1])
                # Get number of successful swipes, pellet displaced, swipe missed
                ai_S = int(re.findall(r"S\d+", ai_outcome)[0][1:])
                ai_D = int(re.findall(r"D\d+", ai_outcome)[0][1:])
                print(f'P{pell_idx+1} Manual:{frame_no} - {reach_outcome} - AI:{ai_frame_no} - {ai_reach_outcome}')

                # Outcome column
                if reach_outcome == reach_outcome_dict[ai_reach_outcome]:
                    # color cell green
                    ws2.cell(row=pell_idx+2, column=indices[1]).fill = pastel_green_fill
                    match_count_dict[reach_outcome] += 1
                    if ai_S + ai_D > 1:
                        ws2.cell(row=pell_idx+2, column=indices[1]).value = f'{reach_outcome} ({ai_outcome})'
                        # Make font bold
                        ws2.cell(row=pell_idx+2, column=indices[1]).font = Font(bold=True)
                else:
                    # color cell red
                    ws2.cell(row=pell_idx+2, column=indices[1]).fill = pastel_pink_fill
                    # Append AI outcome to manual scored file
                    ws2.cell(row=pell_idx+2, column=indices[1]).value = f'{reach_outcome} ({ai_reach_outcome})'


                # Frame number column
                if ai_reach_outcome == 'N':
                    ws2.cell(row=pell_idx+2, column=indices[0]).fill = pastel_green_fill
                elif np.abs(frame_no - ai_frame_no) < 50:
                    # color cell green
                    ws2.cell(row=pell_idx+2, column=indices[0]).fill = pastel_green_fill 
                else:
                    # color cell red
                    ws2.cell(row=pell_idx+2, column=indices[0]).fill = pastel_pink_fill
                    # Append AI frame number to manual scored file
                    ws2.cell(row=pell_idx+2, column=indices[0]).value = f'{frame_no} ({ai_frame_no})'

            except:
                # color cell yellow
                ws2.cell(row=pell_idx+2, column=col_idx*2+2).fill = pastel_yellow_fill
                ws2.cell(row=pell_idx+2, column=col_idx*2+3).fill = pastel_yellow_fill
                # print(f'P{pell_idx+1} Manual:{frame_no} - {reach_outcome} - AI:None')
                print(traceback.format_exc())

        
        print(f'{col} done')

    s_percent = match_count_dict['swipe successful']/outcome_count_dict['swipe successful']*100
    d_percent = match_count_dict['pellet displaced']/outcome_count_dict['pellet displaced']*100
    m_percent = match_count_dict['swipe missed']/outcome_count_dict['swipe missed']*100
    n_percent = match_count_dict['no swipes']/outcome_count_dict['no swipes']*100

    # Add to worksheet at the bottom after 23nd row
    ws2.cell(row=23, column=1).value = 'Summary'
    ws2.cell(row=24, column=1).value = 'S'
    ws2.cell(row=24, column=2).value = f'{match_count_dict["swipe successful"]}/{outcome_count_dict["swipe successful"]} ({s_percent:.2f}%)'
    ws2.cell(row=25, column=1).value = 'D'
    ws2.cell(row=25, column=2).value = f'{match_count_dict["pellet displaced"]}/{outcome_count_dict["pellet displaced"]} ({d_percent:.2f}%)'
    ws2.cell(row=26, column=1).value = 'M'
    ws2.cell(row=26, column=2).value = f'{match_count_dict["swipe missed"]}/{outcome_count_dict["swipe missed"]} ({m_percent:.2f}%)'
    ws2.cell(row=27, column=1).value = 'N'
    ws2.cell(row=27, column=2).value = f'{match_count_dict["no swipes"]}/{outcome_count_dict["no swipes"]} ({n_percent:.2f}%)'
    ws2.cell(row=28, column=1).value = 'Total'
    tot_matches = match_count_dict["swipe successful"] + match_count_dict["pellet displaced"] + match_count_dict["swipe missed"]
    tot_outcomes = outcome_count_dict["swipe successful"] + outcome_count_dict["pellet displaced"] + outcome_count_dict["swipe missed"]
    ws2.cell(row=28, column=2).value = f'{tot_matches} / {tot_outcomes} ({tot_matches/tot_outcomes*100:.2f}%)'

    # Add input parameters to the worksheet
    ws2.cell(row=23, column=5).value = 'Input Parameters'
    ws2.cell(row=24, column=5).value = 'Pellet Pos Threshold'
    ws2.cell(row=25, column=5).value = summary['pellet_pos_th']
    ws2.cell(row=24, column=6).value = 'Threshold'
    ws2.cell(row=25, column=6).value = summary['threshold']
    ws2.cell(row=24, column=7).value = 'RH Y Confidence'
    ws2.cell(row=25, column=7).value = summary['rh_y_conf']
    ws2.cell(row=24, column=8).value = 'Min Frames'
    ws2.cell(row=25, column=8).value = summary['min_frames']
    ws2.cell(row=24, column=9).value = 'Min Frames Std'
    ws2.cell(row=25, column=9).value = summary['min_frames_std']
    ws2.cell(row=24, column=10).value = 'Th Adjacent'
    ws2.cell(row=25, column=10).value = summary['th_adjacent']
    ws2.cell(row=24, column=11).value = 'Std Motion'
    ws2.cell(row=25, column=11).value = summary['std_motion']

    print(f'S: {match_count_dict["swipe successful"]}/{outcome_count_dict["swipe successful"]}({s_percent:.2f}%)' + 
          f' D: {match_count_dict["pellet displaced"]}/{outcome_count_dict["pellet displaced"]}({d_percent:.2f}%)' + 
          f' M: {match_count_dict["swipe missed"]}/{outcome_count_dict["swipe missed"]}({m_percent:.2f}%)' +
          f' N: {match_count_dict["no swipes"]}/{outcome_count_dict["no swipes"]}({n_percent:.2f}%)' +
          f' Total: {tot_matches}/{tot_outcomes}({tot_matches/tot_outcomes*100:.2f}%)')
    
    wb.save(manual_scored_file)

def generate_dlc_summary_ext(folder_path, mt=True):
    # Go to Single Animal folder for every group using os.walk
    for root, dirs, files in os.walk(folder_path):
        if 'Single_Animal' in dirs:
            generate_group_dlc_summary(folder_path=os.path.join(root, 'Single_Animal'), mt=mt)

def generate_dlc_summary():
    # folder_path = r'A:\ASPA\Data\OptD'
    # folder_path = r'D:\! DLC Output\Analyzed\K'
    # generate_dlc_summary_ext(folder_path=folder_path, mt=True)
    # folder_path = r'D:\! DLC Output\Analyzed\ABS3'

    folder_path = r'D:\! DLC Output\Analyzed'
    folders = ['H', 'I', 'L', 'M', 'OptD', 'OptG', 'K']
    for folder in folders:
        generate_dlc_summary_ext(folder_path=os.path.join(folder_path, folder))


def update_summary_headers():
    # summary_path = r'C:\Users\3449mathews\Downloads\AI Summary'
    summary_path = r'A:\AI Summary'

    # Get list of all summary files
    files = glob.glob(f'{summary_path}/*.xlsx')

    for file in files:
        # Load the data into a DataFrame
        df = pd.read_excel(file, engine='openpyxl')

        for col in df.columns:
            # Check if columns is string type
            if isinstance(df[col].iloc[1], str):
                # Remove units from column name by splitting at space
                df[col] = df[col].apply(lambda x: x.split(' ')[0] if isinstance(x, str) else x)

        # Rename the columns
        df.rename(columns={
            'Swipe length': 'Swipe length (mm)',
            'Swipe breadth': 'Swipe breadth (mm)',
            'Swipe area': 'Swipe area (mm^2)',
            'Swipe speed': 'Swipe speed (mm/s)',
            'Swiperange': 'Swipe range (seconds)'
        }, inplace=True)
        
        # Write the DataFrame back to the Excel file
        df.to_excel(file, index=False)

def strip_swipe_info_file(swipe_info_file):

    # Read the Excel file
    df = pd.read_excel(swipe_info_file)

    # Create a new DataFrame to store the transformed data
    new_df = pd.DataFrame(columns=[
        'Date',
        'Animal ID',
        'Tray',
        'Frame ID', 
        'Pellet No',
        'Reach outcome',
        'Swipe breadth (pixels)',
        'Swipe breadth (mm)',
        'Swipe length (pixels)',
        'Swipe length (mm)',
        'Swipe Area (pix^2)',
        'Swipe Area (mm^2)',
        'Swipe speed (mm/s)',
        'Swipe duration (frames)',
        'Swipe duration (s)'
    ])

    row_count = len(df)
    for index, row in df.iterrows():
        try:
            # Split the values in the row
            values = row.values
            reach_outcome = str(values[5]).split()
            reach_outcome = ' '.join(reach_outcome[:-3])
            # Extract and rearrange the desired values
            new_row = [
                values[0],  # Date
                values[1],  # Animal ID
                values[2],  # Tray
                values[3],  # Frame ID
                values[4],  # Pellet No
                reach_outcome,  # Reach outcome
                float(values[6].split('(')[0]),  # Swipe breadth (pixels)
                float(values[6].split('(')[1].replace('mm)', '')),  # Swipe breadth (mm)
                float(values[7].split('(')[0]),  # Swipe length (pixels)
                float(values[7].split('(')[1].replace('mm)', '')),  # Swipe length (mm)
                float(values[8].split('(')[0]),  # Swipe Area (pix^2)
                float(values[8].split('(')[1].replace('mm^2)', '')),  # Swipe Area (mm^2)
                float(values[9].replace('mm/s', '')),  # Swipe speed (mm/s)
                int(values[11].split('-')[1].split(' ')[0]) - int(values[11].split('-')[0]),  # Swipe duration (frames)
                float(values[11].split('(')[1].split(', ')[-1].replace('seconds)', '')),  # Swipe duration (s)
            ]
            
            # Append the new row to the new DataFrame
            new_df.loc[len(new_df)] = new_row

            if index % 1000 == 0:
                progress = (index + 1) / row_count * 100
                print(f'{swipe_info_file}: {progress:.2f}%')
        except:
            print(f'Error in processing row {index} in {swipe_info_file}')
            print(f'Row: {row.values}')
            print(traceback.format_exc())

    print(f'Finished processing {swipe_info_file}')


    # Save the new DataFrame to an Excel file
    new_df.to_excel(f'{swipe_info_file.split(".")[0]}_stripped.xlsx', index=False)

def strip_swipe_info_files(folder_path, mt=True):
    # Get list of all swipe info files (starting with 'swipe_info')
    files = glob.glob(f'{folder_path}/swipe_info*.xlsx')

    if mt:
        results = []
        # Run using concurrent.futures
        with concurrent.futures.ProcessPoolExecutor() as executor:
            futures = [executor.submit(strip_swipe_info_file, file) for file in files]

            for future in tqdm(concurrent.futures.as_completed(futures), total=len(futures)):
                results.append(future.result())
    else:
        for file in files:
            strip_swipe_info_file(file)

if __name__ == '__main__':
    # manual_ai_comparison()
    generate_dlc_summary()
    # update_summary_headers()
    # strip_swipe_info_files(r'A:\ASPA\AI', mt=False)

    



    


    
        
    




